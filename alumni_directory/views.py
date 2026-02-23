import logging
import io
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Count, F
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import get_user_model
from .models import Alumni, AlumniDocument, Achievement, ProfessionalExperience
from django.core.exceptions import PermissionDenied
from django.db import transaction, IntegrityError
from django.http import Http404
from django.core.mail import send_mail
from django.urls import reverse
from django.conf import settings
from accounts.decorators import paginate
from django.contrib import messages
from accounts.models import Profile, Experience
from .forms import AlumniForm, AlumniFilterForm, AlumniSearchForm, AlumniDocumentForm
import csv
from datetime import datetime

logger = logging.getLogger(__name__)

User = get_user_model()

def apply_selective_export_filters(request, base_queryset=None):
    """
    Apply selective export filters and generate appropriate filename
    Returns tuple: (filtered_queryset, filename, has_selective_filters)
    """
    if base_queryset is None:
        export_queryset = Alumni.objects.select_related('user').all()
    else:
        export_queryset = base_queryset

    # Apply selective export filters
    export_colleges = [c for c in request.GET.getlist('export_colleges') if c]
    if export_colleges:
        export_queryset = export_queryset.filter(college__in=export_colleges)

    export_courses = [c for c in request.GET.getlist('export_courses') if c]
    if export_courses:
        export_queryset = export_queryset.filter(course__in=export_courses)

    export_years = [y for y in request.GET.getlist('export_years') if y]
    if export_years:
        export_queryset = export_queryset.filter(graduation_year__in=export_years)

    # Year range filters
    year_from = request.GET.get('export_year_from', '').strip()
    year_to = request.GET.get('export_year_to', '').strip()
    if year_from:
        export_queryset = export_queryset.filter(graduation_year__gte=year_from)
    if year_to:
        export_queryset = export_queryset.filter(graduation_year__lte=year_to)

    export_employment_status = [e for e in request.GET.getlist('export_employment_status') if e]
    if export_employment_status:
        export_queryset = export_queryset.filter(employment_status__in=export_employment_status)

    export_verification_status = [v for v in request.GET.getlist('export_verification_status') if v]
    if export_verification_status:
        # Convert string values to boolean
        verification_filters = []
        for status in export_verification_status:
            if status.lower() == 'true':
                verification_filters.append(True)
            elif status.lower() == 'false':
                verification_filters.append(False)
        if verification_filters:
            export_queryset = export_queryset.filter(is_verified__in=verification_filters)

    # Generate dynamic filename
    filename_parts = ['alumni']

    if export_colleges:
        college_codes = '_'.join(export_colleges)
        filename_parts.append(college_codes)

    if year_from or year_to or export_years:
        if year_from and year_to:
            filename_parts.append(f"{year_from}-{year_to}")
        elif export_years:
            if len(export_years) <= 3:
                filename_parts.append('_'.join(export_years))
            else:
                filename_parts.append(f"{min(export_years)}-{max(export_years)}")

    if export_employment_status and len(export_employment_status) < 5:
        emp_codes = '_'.join([status.split('_')[0] for status in export_employment_status])
        filename_parts.append(emp_codes)

    # Add timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename_parts.append(timestamp)

    filename = '_'.join(filename_parts) + '.csv'

    # Check if any selective filters were applied
    has_selective_filters = any([export_colleges, export_courses, export_years, year_from, year_to,
                                export_employment_status, export_verification_status])

    return export_queryset, filename, has_selective_filters

def is_admin(user):
    """Check if user has admin privileges (staff, superuser, or alumni coordinator)"""
    if not user.is_authenticated:
        return False
    
    is_coordinator = hasattr(user, 'profile') and user.profile.is_alumni_coordinator
    return user.is_superuser or user.is_staff or is_coordinator

@login_required
@paginate(per_page=12)  # Show 12 alumni per page
def alumni_list(request):
    import time
    from django.db import connection
    start_time = time.time()
    initial_query_count = len(connection.queries)
    
    try:
        # Log list view access with user access pattern
        logger.debug(
            f"Alumni list view accessed: User={request.user.username if request.user.is_authenticated else 'Anonymous'}",
            extra={
                'user_id': request.user.id if request.user.is_authenticated else None,
                'is_authenticated': request.user.is_authenticated,
                'ip_address': request.META.get('REMOTE_ADDR'),
                'user_agent': request.META.get('HTTP_USER_AGENT', '')[:100],  # Truncate for logging
                'action': 'list_view_access'
            }
        )
        
        # Get all unique values for filters
        graduation_years = Alumni.objects.values_list('graduation_year', flat=True).distinct().order_by('-graduation_year')
        courses = Alumni.objects.values_list('course', flat=True).distinct().order_by('course')
        provinces = Alumni.objects.values_list('province', flat=True).distinct().order_by('province')
        
        # Get counts for statistics
        graduation_years_count = graduation_years.count()
        courses_count = courses.count()
        provinces_count = provinces.count()
        
        # Base queryset with efficient loading
        queryset = Alumni.objects.select_related('user').all()
        initial_count = queryset.count()
        
        logger.debug(
            f"Initial alumni queryset count: {initial_count}",
            extra={
                'user_id': request.user.id if request.user.is_authenticated else None,
                'initial_count': initial_count,
                'action': 'query_execution'
            }
        )
        
        # Apply filters
        search_query = request.GET.get('search', '').strip()
        if search_query:
            # Create a lookup for college display names
            college_lookup = Q()
            for code, name in Alumni.COLLEGE_CHOICES:
                if search_query.lower() in name.lower():
                    college_lookup |= Q(college=code)

            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(course__icontains=search_query) |
                Q(city__icontains=search_query) |
                Q(province__icontains=search_query) |
                Q(current_company__icontains=search_query) |
                Q(job_title__icontains=search_query) |
                college_lookup  # Include college name search
            )
        
        # Graduation Year filter
        grad_year = [y for y in request.GET.getlist('graduation_year') if y]
        if grad_year:
            queryset = queryset.filter(graduation_year__in=grad_year)
        
        # Course filter
        course = [c for c in request.GET.getlist('course') if c]
        if course:
            queryset = queryset.filter(course__in=course)
        
        # College filter
        college = [c for c in request.GET.getlist('college') if c]
        if college:
            queryset = queryset.filter(college__in=college)
        
        # Campus filter
        campus = [c for c in request.GET.getlist('campus') if c]
        if campus:
            queryset = queryset.filter(campus__in=campus)
        
        # Location filter
        province = [p for p in request.GET.getlist('province') if p]
        if province:
            queryset = queryset.filter(province__in=province)
        
        # Employment Status filter
        employment_status = [e for e in request.GET.getlist('employment_status') if e]
        if employment_status:
            queryset = queryset.filter(employment_status__in=employment_status)
        
        # Count selected filters
        selected_filters_count = sum(
            bool(x) for x in [
                grad_year, course, college, campus, 
                province, employment_status, search_query
            ]
        )
        any_filter_active = selected_filters_count > 0
        
        # Get total counts
        total_alumni = Alumni.objects.count()
        total_registered = queryset.count()
        
        context = {
            'alumni_list': queryset,  # The decorator will paginate this
            'graduation_years': graduation_years,
            'courses': courses,
            'provinces': provinces,
            'colleges': Alumni.COLLEGE_CHOICES,
            'campuses': Alumni.CAMPUS_CHOICES,
            'employment_statuses': Alumni.EMPLOYMENT_STATUS_CHOICES,
            'selected_filters': {
                'graduation_year': grad_year,
                'course': course,
                'college': college,
                'campus': campus,
                'province': province,
                'employment_status': employment_status,
                'search': search_query,
            },
            'selected_filters_count': selected_filters_count,
            'any_filter_active': any_filter_active,
            'total_alumni': total_alumni,
            'total_registered': total_registered,
            'graduation_years_count': graduation_years_count,
            'courses_count': courses_count,
            'provinces_count': provinces_count
        }
        
        template_name = 'alumni_directory/alumni_list.html'
        if request.headers.get('HX-Request'):
            template_name = 'alumni_directory/partials/alumni_list.html'
        
        # Calculate performance metrics
        elapsed_time = time.time() - start_time
        final_query_count = len(connection.queries)
        queries_executed = final_query_count - initial_query_count
        
        # Log performance metrics
        logger.debug(
            f"Alumni list view completed: Time={elapsed_time:.3f}s, Queries={queries_executed}, Results={total_registered}",
            extra={
                'user_id': request.user.id if request.user.is_authenticated else None,
                'elapsed_time': elapsed_time,
                'queries_executed': queries_executed,
                'total_results': total_registered,
                'filters_active': any_filter_active,
                'selected_filters_count': selected_filters_count,
                'action': 'list_view_complete'
            }
        )
        
        # Log slow operations warning
        if elapsed_time > 2.0:
            logger.warning(
                f"Slow alumni list view: Time={elapsed_time:.3f}s, Queries={queries_executed}",
                extra={
                    'user_id': request.user.id if request.user.is_authenticated else None,
                    'elapsed_time': elapsed_time,
                    'queries_executed': queries_executed,
                    'threshold': 2.0,
                    'action': 'slow_operation'
                }
            )
        
        return render(request, template_name, context)
        
    except Exception as e:
        elapsed_time = time.time() - start_time if 'start_time' in locals() else 0
        logger.error(
            f"Error in alumni list view: {str(e)}, Time={elapsed_time:.3f}s",
            exc_info=True,
            extra={
                'user_id': request.user.id if request.user.is_authenticated else None,
                'error_type': type(e).__name__,
                'elapsed_time': elapsed_time
            }
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'error',
                'message': 'An error occurred while loading the alumni list.'
            }, status=500)
        raise

@login_required
def alumni_detail(request, pk):
    import time
    from django.db import connection
    start_time = time.time()
    initial_query_count = len(connection.queries)
    
    try:
        # Get alumni object or 404
        alumni = get_object_or_404(Alumni, pk=pk)
        try:
            alumni_name = alumni.full_name
        except Exception:
            alumni_name = f"Alumni ID: {pk}"
            logger.warning(f"Could not get full name for alumni ID: {pk}")
        
        # Log detail view access with user access pattern
        logger.debug(
            f"Alumni detail view accessed: ID={pk}, Name={alumni_name}, User={request.user.username if request.user.is_authenticated else 'Anonymous'}",
            extra={
                'alumni_id': pk,
                'alumni_name': alumni_name,
                'user_id': request.user.id if request.user.is_authenticated else None,
                'is_authenticated': request.user.is_authenticated,
                'is_staff': request.user.is_staff if request.user.is_authenticated else False,
                'ip_address': request.META.get('REMOTE_ADDR'),
                'user_agent': request.META.get('HTTP_USER_AGENT', '')[:100],
                'action': 'detail_view_access'
            }
        )
        
        # Log sensitive data access (if viewing someone else's profile)
        if request.user.is_authenticated and hasattr(request.user, 'alumni'):
            if request.user.alumni.id != pk and not request.user.is_staff:
                logger.info(
                    f"Sensitive data access: User {request.user.id} viewing alumni {pk} profile",
                    extra={
                        'viewer_id': request.user.id,
                        'viewed_alumni_id': pk,
                        'action': 'sensitive_data_access'
                    }
                )
        
        # Get profile
        profile = None
        try:
            profile = alumni.user.profile
            logger.info(f"Profile found for alumni ID: {pk}")
        except Exception:
            # Profile doesn't exist or other error
            logger.warning(f"No profile found for alumni ID: {pk}")
            profile = None
        
        # Get documents from both AlumniDocument and accounts.Document models
        alumni_documents = list(alumni.documents.all().order_by('-uploaded_at'))
        profile_documents = []
        
        if profile:
            from accounts.models import Document
            
            # Create a wrapper class to make profile documents behave like AlumniDocuments
            class DocumentWrapper:
                def __init__(self, doc):
                    self.title = doc.title
                    self.document_type = doc.document_type
                    self.file = doc.file
                    self.uploaded_at = doc.uploaded_at
                    self.is_verified = doc.is_verified
                
                def get_document_type_display(self):
                    # Map document types from accounts.Document to display names
                    type_map = dict(Document.DOCUMENT_TYPES)
                    return type_map.get(self.document_type, self.document_type)
            
            # Wrap each profile document
            profile_docs = Document.objects.filter(profile=profile).order_by('-uploaded_at')
            profile_documents = [DocumentWrapper(doc) for doc in profile_docs]
            logger.info(f"Profile documents found: {len(profile_documents)}")
        
        # Combine documents from both sources
        combined_documents = alumni_documents + profile_documents
        
        # Calculate document stats
        total_documents = len(combined_documents)
        verified_documents = sum(1 for doc in combined_documents if hasattr(doc, 'is_verified') and doc.is_verified)
        pending_verification = total_documents - verified_documents
        
        logger.info(f"Documents for alumni ID {pk}: Total={total_documents}, Verified={verified_documents}, Pending={pending_verification}")
        
        # Debug: Check if documents exist in the database
        if total_documents == 0:
            logger.warning(f"No documents found for alumni ID: {pk}")
            # Check if there are any documents in the system at all
            all_docs_count = AlumniDocument.objects.count()
            logger.info(f"Total AlumniDocuments in system: {all_docs_count}")
            
            from accounts.models import Document
            all_profile_docs_count = Document.objects.count()
            logger.info(f"Total Profile Documents in system: {all_profile_docs_count}")
        
        # Get professional experiences using the ProfessionalExperience utility class
        unified_experience = ProfessionalExperience.get_unified_experience(alumni)
        career_path = ProfessionalExperience.get_career_path_only(alumni)
        work_experience = ProfessionalExperience.get_regular_experience_only(alumni)
        
        # Calculate profile completion percentage
        completion_fields = [
            alumni.gender,
            alumni.date_of_birth,
            alumni.phone_number,
            alumni.address,
            alumni.province,
            alumni.city,
            alumni.country,
            alumni.current_company,
            alumni.job_title,
            alumni.bio
        ]
        
        filled_fields = sum(1 for field in completion_fields if field)
        completion_percentage = int((filled_fields / len(completion_fields)) * 100)
        
        # Get achievements
        achievements = alumni.achievements_list.all().order_by('-date_achieved')
        
        # Check if current user is owner
        is_owner = request.user == alumni.user
        
        # Check if the user is staff/admin
        is_admin = request.user.is_staff
        
        context = {
            'alumni': alumni,
            'profile': profile,
            'documents': combined_documents,
            'total_documents': total_documents,
            'verified_documents': verified_documents,
            'pending_verification': pending_verification,
            'completion_percentage': completion_percentage,
            'unified_experience': unified_experience,
            'career_path': career_path,
            'work_experience': work_experience,
            'achievements': achievements,
            'is_owner': is_owner,
            'is_admin': is_admin
        }
        
        return render(request, 'alumni_directory/alumni_detail.html', context)
        
    except Exception as e:
        logger.error(f"Error in alumni_detail view: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'error',
                'message': 'An error occurred while loading the alumni details.'
            }, status=500)
        raise

@user_passes_test(is_admin, login_url='account_login')
def download_document(request, doc_id):
    try:
        document = get_object_or_404(AlumniDocument, id=doc_id)
        
        # Verify the document exists and is accessible
        if not document.file:
            logger.warning(f"Document {doc_id} has no file attached")
            return JsonResponse({
                'status': 'error',
                'message': 'Document file not found.'
            }, status=404)
        
        logger.info(f"Document download requested for ID: {doc_id}")
        
        response = JsonResponse({
            'status': 'success',
            'url': document.file.url,
            'filename': document.file.name.split('/')[-1]
        })
        
        return response
        
    except Exception as e:
        logger.error(f"Error in download_document view for doc ID {doc_id}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'An error occurred while processing the document download.'
        }, status=500)

@user_passes_test(is_admin, login_url='account_login')
def send_reminder(request, pk):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
        
    try:
        alumni = get_object_or_404(Alumni.objects.select_related('user'), pk=pk)
        
        # Get missing fields
        required_fields = [
            alumni.gender, alumni.date_of_birth, alumni.phone_number,
            alumni.province, alumni.city, alumni.address,
            alumni.graduation_year, alumni.course
        ]
        optional_fields = [
            alumni.alternate_email, alumni.linkedin_profile,
            alumni.major, alumni.honors, alumni.current_company,
            alumni.job_title, alumni.industry, alumni.skills,
            alumni.interests, alumni.bio, alumni.achievements
        ]
        
        completed_required = sum(1 for field in required_fields if field)
        completed_optional = sum(1 for field in optional_fields if field)
        
        # Get missing sections
        empty_sections = []
        if not any([alumni.phone_number, alumni.alternate_email, alumni.linkedin_profile]):
            empty_sections.append('Personal Information')
        if not any([alumni.province, alumni.city, alumni.address]):
            empty_sections.append('Location Information')
        if not any([alumni.current_company, alumni.job_title, alumni.industry]):
            empty_sections.append('Professional Information')
        if not any([alumni.major, alumni.honors]):
            empty_sections.append('Academic Information')
        if not alumni.skills:
            empty_sections.append('Skills')
        if not alumni.interests:
            empty_sections.append('Interests')
        if not alumni.documents.exists():
            empty_sections.append('Documents')
        if not any([alumni.bio, alumni.achievements]):
            empty_sections.append('Additional Information')
        
        # Log reminder sending attempt
        logger.info(
            f"Profile completion reminder requested: Alumni ID={pk}, Email={alumni.email}",
            extra={
                'alumni_id': pk,
                'alumni_email': alumni.email,
                'alumni_name': alumni.full_name,
                'completed_required': completed_required,
                'total_required': len(required_fields),
                'completed_optional': completed_optional,
                'total_optional': len(optional_fields),
                'empty_sections': empty_sections,
                'user_id': request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
                'action': 'reminder_send_attempt'
            }
        )
        
        # Send email
        subject = 'Complete Your Alumni Profile'
        message = f"""Dear {alumni.full_name},

We noticed that your alumni profile is incomplete. Please take a moment to update your profile with the following missing information:

{chr(10).join('- ' + section for section in empty_sections)}

Completing your profile helps us better serve the alumni community and keep you connected with opportunities.

Required Fields Completed: {(completed_required / len(required_fields)) * 100:.0f}%
Optional Fields Completed: {(completed_optional / len(optional_fields)) * 100:.0f}%

Click here to update your profile: {request.build_absolute_uri(reverse('accounts:profile'))}

Best regards,
The Alumni Team"""

        from core.email_utils import send_email_with_provider
        
        try:
            success = send_email_with_provider(
                subject=subject,
                message=message,
                recipient_list=[alumni.email],
                from_email=settings.DEFAULT_FROM_EMAIL,
                fail_silently=False
            )
            
            if success:
                logger.info(
                    f"Profile completion reminder sent successfully: Alumni ID={pk}, Email={alumni.email}",
                    extra={
                        'alumni_id': pk,
                        'alumni_email': alumni.email,
                        'alumni_name': alumni.full_name,
                        'user_id': request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
                        'action': 'reminder_sent'
                    }
                )
                return JsonResponse({
                    'status': 'success',
                    'message': 'Reminder sent successfully'
                })
            else:
                logger.error(
                    f"Failed to send profile completion reminder: Alumni ID={pk}, Email={alumni.email}",
                    extra={
                        'alumni_id': pk,
                        'alumni_email': alumni.email,
                        'error_type': 'email_send_failed',
                        'user_id': request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
                        'action': 'reminder_send_failed'
                    }
                )
                return JsonResponse({
                    'status': 'error',
                    'message': 'Failed to send reminder'
                })
        except Exception as e:
            logger.error(
                f"Exception sending profile completion reminder: {str(e)}",
                exc_info=True,
                extra={
                    'alumni_id': pk,
                    'alumni_email': alumni.email,
                    'error_type': type(e).__name__,
                    'user_id': request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
                    'action': 'reminder_send_exception'
                }
            )
            return JsonResponse({
                'status': 'error',
                'message': 'Failed to send reminder'
            })
        
    except Alumni.DoesNotExist:
        logger.error(f"Alumni with ID {pk} not found")
        return JsonResponse({
            'status': 'error',
            'message': 'Alumni not found'
        }, status=404)
        
    except Exception as e:
        logger.error(f"Error sending reminder to alumni ID {pk}: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': 'Failed to send reminder'
        }, status=500)

@login_required
def tabular_alumni_list(request):
    """
    Display alumni in a tabular format with specific columns.
    """
    try:
        # Base queryset with efficient loading
        queryset = Alumni.objects.select_related('user').all()
        
        # Apply filters
        search_query = request.GET.get('search', '').strip()
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) | 
                Q(user__last_name__icontains=search_query) |
                Q(course__icontains=search_query) |
                Q(city__icontains=search_query) |
                Q(province__icontains=search_query)
            )
        
        # Graduation Year filter
        grad_year = [y for y in request.GET.getlist('graduation_year') if y]
        if grad_year:
            queryset = queryset.filter(graduation_year__in=grad_year)
        
        # Course filter
        course = [c for c in request.GET.getlist('course') if c]
        if course:
            queryset = queryset.filter(course__in=course)

        # College filter
        college = [c for c in request.GET.getlist('college') if c]
        if college:
            queryset = queryset.filter(college__in=college)
            
        # Export to CSV if requested
        if request.GET.get('format') == 'csv':
            # Apply selective export filters
            export_queryset, filename, has_selective_filters = apply_selective_export_filters(request)

            # If no selective filters are applied, use the current filtered queryset
            if not has_selective_filters:
                export_queryset = queryset
                filename = f"alumni_directory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            writer = csv.writer(response)
            writer.writerow(['ID', 'Full Name', 'College', 'Year', 'Course', 'Present Occupation', 'Company', 'Employment Address'])

            for alumni in export_queryset:
                # Get current experience for occupation and company
                current_exp = None
                if hasattr(alumni.user, 'profile'):
                    current_exp = alumni.user.profile.experience.filter(is_current=True).first()

                writer.writerow([
                    alumni.id,
                    alumni.full_name,
                    alumni.college_display if alumni.college else 'Not specified',
                    alumni.graduation_year,
                    alumni.course,
                    current_exp.position if current_exp else alumni.job_title,
                    current_exp.company if current_exp else alumni.current_company,
                    current_exp.location if current_exp else f"{alumni.city}, {alumni.province}"
                ])

            return response
        
        # Pagination
        paginator = Paginator(queryset, 25)  # 25 items per page
        page_number = request.GET.get('page')
        alumni_page = paginator.get_page(page_number)
        
        # Get counts for filtering dropdown options
        graduation_years = Alumni.objects.values_list('graduation_year', flat=True).distinct().order_by('-graduation_year')
        courses = Alumni.objects.values_list('course', flat=True).distinct().order_by('course')
        
        context = {
            'alumni_list': alumni_page,
            'graduation_years': graduation_years,
            'courses': courses,
            'colleges': Alumni.COLLEGE_CHOICES,
            'search_query': search_query,
            'selected_year': grad_year[0] if grad_year else None,
            'selected_course': course[0] if course else None,
            'selected_college': college[0] if college else None,
        }
        
        return render(request, 'alumni_directory/tabular_alumni_list.html', context)
        
    except Exception as e:
        logger.error(f"Error in tabular_alumni_list view: {str(e)}")
        messages.error(request, "An error occurred while loading the alumni list.")
        return redirect('alumni_directory:alumni_list')

@user_passes_test(is_admin, login_url='account_login')
def alumni_management(request):
    """
    Alumni Management view with tabular display, import/export functionality.
    """
    try:
        # Handle CSV import
        if request.method == 'POST' and 'import_csv' in request.POST:
            csv_file = request.FILES.get('csv_file')
            if csv_file:
                # Log CSV import start
                file_size = csv_file.size
                file_name = csv_file.name
                
                logger.info(
                    f"CSV import started: File={file_name}, Size={file_size} bytes",
                    extra={
                        'file_name': file_name,
                        'file_size': file_size,
                        'user_id': request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
                        'action': 'csv_import_start'
                    }
                )
                
                try:
                    # Read CSV file
                    decoded_file = csv_file.read().decode('utf-8')
                    csv_data = csv.DictReader(decoded_file.splitlines())
                    
                    imported_count = 0
                    updated_count = 0
                    error_count = 0
                    error_details = []
                    
                    for row in csv_data:
                        # Extract data from CSV row
                        full_name = row.get('Full Name', '').strip()
                        graduation_year = row.get('Year', '').strip()
                        course = row.get('Course', '').strip()
                        occupation = row.get('Present Occupation', '').strip()
                        company = row.get('Name of Company', '').strip()
                        address = row.get('Employment Address', '').strip()
                        
                        if full_name and graduation_year:
                            # Split full name into first and last name
                            name_parts = full_name.split(' ', 1)
                            first_name = name_parts[0]
                            last_name = name_parts[1] if len(name_parts) > 1 else ''
                            
                            # Try to find existing alumni by name and graduation year
                            try:
                                alumni = Alumni.objects.get(
                                    user__first_name__iexact=first_name,
                                    user__last_name__iexact=last_name,
                                    graduation_year=graduation_year
                                )
                                # Update existing alumni
                                alumni.course = course
                                alumni.job_title = occupation
                                alumni.current_company = company
                                if address:
                                    address_parts = address.split(', ')
                                    if len(address_parts) >= 2:
                                        alumni.city = address_parts[0]
                                        alumni.province = address_parts[1]
                                alumni.save()
                                updated_count += 1
                            except Alumni.DoesNotExist:
                                # Create new user and alumni
                                try:
                                    user = User.objects.create_user(
                                        username=f"{first_name.lower()}.{last_name.lower()}.{graduation_year}",
                                        first_name=first_name,
                                        last_name=last_name,
                                        email=f"{first_name.lower()}.{last_name.lower()}@example.com"
                                    )
                                    
                                    alumni = Alumni.objects.create(
                                        user=user,
                                        graduation_year=graduation_year,
                                        course=course,
                                        job_title=occupation,
                                        current_company=company
                                    )
                                    
                                    if address:
                                        address_parts = address.split(', ')
                                        if len(address_parts) >= 2:
                                            alumni.city = address_parts[0]
                                            alumni.province = address_parts[1]
                                            alumni.save()
                                    
                                    imported_count += 1
                                except IntegrityError as ie:
                                    # Handle database constraint violation (duplicate email/username)
                                    error_count += 1
                                    error_details.append({
                                        'row': full_name,
                                        'error': 'Unable to create account. A user with this email or username already exists.'
                                    })
                                    logger.error(
                                        f"Database constraint violation during CSV import: {full_name}",
                                        extra={
                                            'row_data': full_name,
                                            'error_type': 'IntegrityError',
                                            'error_message': str(ie),
                                            'file_name': file_name,
                                            'action': 'csv_import_integrity_error'
                                        }
                                    )
                            except Exception as row_error:
                                error_count += 1
                                error_details.append({
                                    'row': full_name,
                                    'error': str(row_error)
                                })
                                logger.warning(
                                    f"Error processing CSV row: {str(row_error)}",
                                    extra={
                                        'row_data': full_name,
                                        'error_type': type(row_error).__name__,
                                        'file_name': file_name,
                                        'action': 'csv_row_error'
                                    }
                                )
                    
                    # Log CSV import completion
                    logger.info(
                        f"CSV import completed: Imported={imported_count}, Updated={updated_count}, Errors={error_count}",
                        extra={
                            'file_name': file_name,
                            'file_size': file_size,
                            'imported_count': imported_count,
                            'updated_count': updated_count,
                            'error_count': error_count,
                            'total_processed': imported_count + updated_count + error_count,
                            'user_id': request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
                            'action': 'csv_import_complete'
                        }
                    )
                    
                    if error_count > 0:
                        logger.warning(
                            f"CSV import completed with errors: {error_count} errors out of {imported_count + updated_count + error_count} rows",
                            extra={
                                'file_name': file_name,
                                'error_count': error_count,
                                'error_details': error_details[:10],  # Limit to first 10 errors
                                'user_id': request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
                                'action': 'csv_import_errors'
                            }
                        )
                    
                    messages.success(request, f"Successfully imported {imported_count} new alumni and updated {updated_count} existing records.")
                    if error_count > 0:
                        messages.warning(request, f"Warning: {error_count} rows had errors during import.")
                    
                except Exception as e:
                    logger.error(
                        f"Error importing CSV: {str(e)}",
                        exc_info=True,
                        extra={
                            'file_name': file_name if 'file_name' in locals() else None,
                            'file_size': file_size if 'file_size' in locals() else None,
                            'user_id': request.user.id if hasattr(request, 'user') and request.user.is_authenticated else None,
                            'error_type': type(e).__name__,
                            'action': 'csv_import_failed'
                        }
                    )
                    messages.error(request, f"Error importing CSV file: {str(e)}")
            else:
                messages.error(request, "Please select a CSV file to import.")
        
        # Base queryset with efficient loading
        queryset = Alumni.objects.select_related('user').all()
        
        # Apply filters
        search_query = request.GET.get('search', '').strip()
        if search_query:
            # Create a lookup for college display names
            college_lookup = Q()
            for code, name in Alumni.COLLEGE_CHOICES:
                if search_query.lower() in name.lower():
                    college_lookup |= Q(college=code)

            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(course__icontains=search_query) |
                Q(city__icontains=search_query) |
                Q(province__icontains=search_query) |
                Q(current_company__icontains=search_query) |
                Q(job_title__icontains=search_query) |
                college_lookup  # Include college name search
            )
        
        # Graduation Year filter
        grad_year = [y for y in request.GET.getlist('graduation_year') if y]
        if grad_year:
            queryset = queryset.filter(graduation_year__in=grad_year)
        
        # Course filter
        course = [c for c in request.GET.getlist('course') if c]
        if course:
            queryset = queryset.filter(course__in=course)

        # College filter
        college = [c for c in request.GET.getlist('college') if c]
        if college:
            queryset = queryset.filter(college__in=college)
            
        # Export if requested (Excel and PDF only)
        export_format = request.GET.get('format')
        if export_format in ['excel', 'pdf']:
            # Apply selective export filters
            export_queryset, filename, has_selective_filters = apply_selective_export_filters(request)

            # If no selective filters are applied, use the current filtered queryset
            if not has_selective_filters:
                export_queryset = queryset
                filename = f"alumni_management_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            # Define field configuration for exports
            field_names = ['id', 'full_name', 'college_display', 'graduation_year', 'course', 
                          'job_title', 'current_company', 'employment_address']
            field_labels = ['ID', 'Full Name', 'College', 'Year', 'Course', 
                           'Present Occupation', 'Name of Company', 'Employment Address']

            if export_format == 'excel':
                from core.export_utils import ExportMixin, LogoHeaderService
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.cell.cell import MergedCell
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Alumni Data"
                
                # Get logo path and add header
                logo_path = LogoHeaderService.get_logo_path()
                header_start_row = LogoHeaderService.add_excel_header(ws, logo_path, title="NORSU Alumni System")
                
                # Style for header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Write header row
                for col, label in enumerate(field_labels, 1):
                    cell = ws.cell(row=header_start_row, column=col, value=label)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data rows
                data_start_row = header_start_row + 1
                for row_idx, alumni in enumerate(export_queryset, data_start_row):
                    # Get current experience for occupation and company
                    current_exp = None
                    if hasattr(alumni.user, 'profile'):
                        current_exp = alumni.user.profile.experience.filter(is_current=True).first()
                    
                    row_data = [
                        alumni.id,
                        alumni.full_name,
                        alumni.college_display if alumni.college else 'Not specified',
                        alumni.graduation_year,
                        alumni.course,
                        current_exp.position if current_exp else alumni.job_title,
                        current_exp.company if current_exp else alumni.current_company,
                        current_exp.location if current_exp else f"{alumni.city}, {alumni.province}" if alumni.city and alumni.province else ""
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col, value=str(value) if value else '')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        if not isinstance(cell, MergedCell):
                            column_letter = cell.column_letter
                            break
                    
                    if column_letter is None:
                        continue
                    
                    for cell in column:
                        try:
                            if not isinstance(cell, MergedCell) and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to response
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                response.write(output.getvalue())
                
                return response
            
            elif export_format == 'pdf':
                from core.export_utils import LogoHeaderService
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
                from reportlab.pdfgen import canvas as pdf_canvas
                from django.utils import timezone
                
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
                
                # Get logo path
                logo_path = LogoHeaderService.get_logo_path()
                
                # Determine page orientation (landscape for wide tables)
                use_landscape = len(field_labels) > 6
                pagesize = landscape(A4) if use_landscape else A4
                
                # Prepare export info for footer
                export_info = f"Exported on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {export_queryset.count()}"
                
                # Custom canvas class to add header and footer
                class HeaderFooterCanvas(pdf_canvas.Canvas):
                    def __init__(self, *args, **kwargs):
                        pdf_canvas.Canvas.__init__(self, *args, **kwargs)
                        self.pages = []
                        
                    def showPage(self):
                        self.pages.append(dict(self.__dict__))
                        self._startPage()
                        
                    def save(self):
                        page_count = len(self.pages)
                        for page_num, page in enumerate(self.pages, 1):
                            self.__dict__.update(page)
                            LogoHeaderService.add_pdf_header(self, doc, logo_path, title="NORSU Alumni System")
                            self.draw_footer(page_num, page_count, export_info, pagesize)
                            pdf_canvas.Canvas.showPage(self)
                        pdf_canvas.Canvas.save(self)
                        
                    def draw_footer(self, page_num, page_count, info_text, page_size):
                        self.saveState()
                        self.setFont('Helvetica', 7)
                        self.setFillColor(colors.black)
                        text_width = self.stringWidth(info_text, 'Helvetica', 7)
                        x_position = page_size[0] - text_width - 15
                        y_position = 15
                        self.drawString(x_position, y_position, info_text)
                        self.restoreState()
                
                doc = SimpleDocTemplate(
                    response, 
                    pagesize=pagesize, 
                    rightMargin=15, 
                    leftMargin=15, 
                    topMargin=80,
                    bottomMargin=25
                )
                elements = []
                
                # Get styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=14,
                    spaceAfter=12,
                    alignment=TA_CENTER,
                    textColor=colors.black
                )
                
                # Add title
                elements.append(Paragraph("Alumni Management Report", title_style))
                elements.append(Spacer(1, 18))
                
                # Prepare table data
                table_data = []
                
                # Header row
                header_row = []
                cell_style = ParagraphStyle(
                    'HeaderCell',
                    parent=styles['Normal'],
                    fontSize=7 if use_landscape else 8,
                    textColor=colors.white,
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold'
                )
                for label in field_labels:
                    header_row.append(Paragraph(str(label), cell_style))
                table_data.append(header_row)
                
                # Data cell style
                data_cell_style = ParagraphStyle(
                    'DataCell',
                    parent=styles['Normal'],
                    fontSize=6 if use_landscape else 7,
                    textColor=colors.black,
                    alignment=TA_LEFT,
                    fontName='Helvetica',
                    leading=8 if use_landscape else 9
                )
                
                # Limit to 1000 rows
                max_rows = 1000
                export_queryset_limited = export_queryset[:max_rows]
                
                for alumni in export_queryset_limited:
                    # Get current experience
                    current_exp = None
                    if hasattr(alumni.user, 'profile'):
                        current_exp = alumni.user.profile.experience.filter(is_current=True).first()
                    
                    row_data = [
                        str(alumni.id),
                        alumni.full_name,
                        alumni.college_display if alumni.college else 'Not specified',
                        str(alumni.graduation_year),
                        alumni.course or '',
                        current_exp.position if current_exp else (alumni.job_title or ''),
                        current_exp.company if current_exp else (alumni.current_company or ''),
                        current_exp.location if current_exp else (f"{alumni.city}, {alumni.province}" if alumni.city and alumni.province else "")
                    ]
                    
                    row = [Paragraph(str(val), data_cell_style) for val in row_data]
                    table_data.append(row)
                
                # Calculate column widths
                available_width = pagesize[0] - 30
                col_widths = [0.08, 0.15, 0.12, 0.08, 0.15, 0.15, 0.15, 0.12]
                final_col_widths = [w * available_width for w in col_widths]
                
                # Create table
                table = Table(table_data, colWidths=final_col_widths, repeatRows=1)
                
                # Table styling
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.black),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 7 if use_landscape else 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 6 if use_landscape else 7),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                    ('TOPPADDING', (0, 1), (-1, -1), 5),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8E8E8')]),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                elements.append(table)
                
                # Add note if truncated
                if export_queryset.count() > max_rows:
                    elements.append(Spacer(1, 10))
                    note_style = ParagraphStyle('Note', parent=styles['Normal'], fontSize=7, textColor=colors.black, alignment=TA_LEFT)
                    elements.append(Paragraph(f"Note: Showing first {max_rows} records out of {export_queryset.count()} total records.", note_style))
                
                # Build PDF
                doc.build(elements, canvasmaker=HeaderFooterCanvas)
                return response
        
        # Pagination
        paginator = Paginator(queryset, 25)  # 25 items per page
        page_number = request.GET.get('page')
        alumni_page = paginator.get_page(page_number)
        
        # Get counts for filtering dropdown options
        graduation_years = Alumni.objects.values_list('graduation_year', flat=True).distinct().order_by('-graduation_year')
        courses = Alumni.objects.values_list('course', flat=True).distinct().order_by('course')
        
        context = {
            'alumni_list': alumni_page,
            'graduation_years': graduation_years,
            'courses': courses,
            'colleges': Alumni.COLLEGE_CHOICES,
            'search_query': search_query,
            'selected_year': grad_year[0] if grad_year else None,
            'selected_course': course[0] if course else None,
            'selected_college': college[0] if college else None,
        }
        
        return render(request, 'alumni_directory/alumni_management.html', context)
        
    except Exception as e:
        logger.error(f"Error in alumni_management view: {str(e)}")
        messages.error(request, "An error occurred while loading the alumni management page.")
        return redirect('alumni_directory:alumni_list')

@user_passes_test(is_admin, login_url='account_login')
def alumni_detail_modal(request, pk):
    """
    Return alumni detail content for modal display.
    """
    try:
        alumni = get_object_or_404(Alumni, pk=pk)
        logger.info(f"Alumni detail modal accessed for ID: {pk}, Name: {alumni.full_name}")

        # Get profile
        profile = None
        unified_experience = []
        if hasattr(alumni.user, 'profile'):
            profile = alumni.user.profile

            # Get all experience entries, appropriately sorted
            unified_experience = profile.experience.all().order_by('-is_current', '-start_date')

        # Get documents from both AlumniDocument and accounts.Document models
        alumni_documents = list(alumni.documents.all().order_by('-uploaded_at'))
        profile_documents = []

        if profile:
            from accounts.models import Document

            # Create a wrapper class to make profile documents behave like AlumniDocuments
            class DocumentWrapper:
                def __init__(self, doc):
                    self.id = doc.id
                    self.title = doc.title
                    self.document_type = doc.document_type
                    self.file = doc.file
                    self.uploaded_at = doc.uploaded_at
                    self.is_verified = doc.is_verified
                    self.file_size = self._get_file_size()

                def _get_file_size(self):
                    try:
                        if self.file and hasattr(self.file, 'size'):
                            size = self.file.size
                            if size < 1024:
                                return f"{size} B"
                            elif size < 1024 * 1024:
                                return f"{size / 1024:.1f} KB"
                            else:
                                return f"{size / (1024 * 1024):.1f} MB"
                    except:
                        pass
                    return "Unknown size"

                def get_document_type_display(self):
                    # Map document types from accounts.Document to display names
                    type_map = dict(Document.DOCUMENT_TYPES)
                    return type_map.get(self.document_type, self.document_type)

            # Wrap each profile document
            profile_docs = Document.objects.filter(profile=profile).order_by('-uploaded_at')
            profile_documents = [DocumentWrapper(doc) for doc in profile_docs]

        # Combine documents from both sources
        combined_documents = alumni_documents + profile_documents

        # Group documents by type for template
        documents_by_type = {}
        for doc in combined_documents:
            doc_type = doc.document_type
            if doc_type not in documents_by_type:
                documents_by_type[doc_type] = []
            documents_by_type[doc_type].append(doc)

        # Document types for template iteration
        document_types = [
            ('RESUME', 'Resume/CV'),
            ('CERT', 'Certification'),
            ('DIPLOMA', 'Diploma'),
            ('TOR', 'Transcript of Records'),
            ('TRANSCRIPT', 'Academic Transcript'),
            ('CERTIFICATE', 'Certificate'),
            ('OTHER', 'Other'),
        ]

        # Calculate document stats
        total_documents = len(combined_documents)
        verified_documents = sum(1 for doc in combined_documents if hasattr(doc, 'is_verified') and doc.is_verified)
        pending_verification = total_documents - verified_documents

        doc_stats = {
            'total': total_documents,
            'verified': verified_documents,
            'pending': pending_verification
        }

        # Calculate profile completion
        required_fields = [
            alumni.gender, alumni.date_of_birth, alumni.phone_number,
            alumni.province, alumni.city, alumni.address,
            alumni.graduation_year, alumni.course
        ]
        optional_fields = [
            alumni.alternate_email, alumni.linkedin_profile,
            alumni.major, alumni.honors, alumni.current_company,
            alumni.job_title, alumni.industry, alumni.skills,
            alumni.interests, alumni.bio, alumni.achievements
        ]

        completed_required = sum(1 for field in required_fields if field)
        completed_optional = sum(1 for field in optional_fields if field)

        required_percentage = (completed_required / len(required_fields)) * 100
        optional_percentage = (completed_optional / len(optional_fields)) * 100
        total_percentage = (required_percentage + optional_percentage) / 2

        profile_completion = {
            'required': required_percentage,
            'optional': optional_percentage,
            'total': total_percentage
        }

        # Determine empty sections
        empty_sections = {
            'personal': not any([alumni.phone_number, alumni.alternate_email, alumni.linkedin_profile]),
            'location': not any([alumni.province, alumni.city, alumni.address]),
            'professional': not any([alumni.current_company, alumni.job_title, alumni.industry]),
            'academic': not any([alumni.major, alumni.honors]),
            'skills': not alumni.skills,
            'interests': not alumni.interests,
            'documents': not combined_documents,
            'additional': not any([alumni.bio, alumni.achievements])
        }

        # Get achievements
        achievements = alumni.achievements_list.all().order_by('-date_achieved')

        context = {
            'alumni': alumni,
            'profile': profile,
            'unified_experience': unified_experience,
            'achievements': achievements,
            'documents': documents_by_type,
            'document_types': document_types,
            'doc_stats': doc_stats,
            'profile_completion': profile_completion,
            'empty_sections': empty_sections,
        }

        return render(request, 'alumni_directory/partials/alumni_detail_modal.html', context)

    except Exception as e:
        logger.error(f"Error in alumni_detail_modal view: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'An error occurred while loading the alumni details.'
        }, status=500)
