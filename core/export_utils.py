import csv
import io
import json
import os
import logging
from datetime import datetime
from typing import Optional
from django.http import HttpResponse
from django.db.models import QuerySet
from django.utils import timezone
from django.conf import settings
from django.contrib.staticfiles import finders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.cell import MergedCell
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

logger = logging.getLogger(__name__)


class LogoHeaderService:
    """Service for managing logo headers in exports"""
    
    # Class-level cache for logo path
    _logo_path_cache = None
    
    # Maximum logo file size (2MB to accommodate existing logo)
    MAX_LOGO_SIZE = 2 * 1024 * 1024
    
    @classmethod
    def get_logo_path(cls) -> Optional[str]:
        """
        Get the absolute path to the logo file with comprehensive error handling.
        
        Returns:
            str: Absolute path to logo file if exists and valid
            None: If logo file not found or invalid (graceful degradation)
        """
        # Return cached path if available
        if cls._logo_path_cache is not None:
            return cls._logo_path_cache
        
        logo_path = None
        
        try:
            # Try to find logo using Django's static files finder (development)
            # Try both possible filenames
            try:
                logo_path = finders.find('images/norsu-logo.png')
                if logo_path is None:
                    logo_path = finders.find('images/logo_norsu.png')
            except Exception as finder_error:
                logger.warning(
                    f"Error using static files finder: {str(finder_error)}",
                    extra={
                        'error_type': type(finder_error).__name__,
                        'fallback': 'Trying STATIC_ROOT path'
                    }
                )
            
            # If not found and STATIC_ROOT is configured, try production path
            if logo_path is None and hasattr(settings, 'STATIC_ROOT') and settings.STATIC_ROOT:
                try:
                    logo_path = os.path.join(settings.STATIC_ROOT, 'images/norsu-logo.png')
                    if not os.path.exists(logo_path):
                        logo_path = os.path.join(settings.STATIC_ROOT, 'images/logo_norsu.png')
                        if not os.path.exists(logo_path):
                            logo_path = None
                except Exception as path_error:
                    logger.error(
                        f"Error constructing logo path from STATIC_ROOT: {str(path_error)}",
                        extra={
                            'error_type': type(path_error).__name__,
                            'static_root': settings.STATIC_ROOT
                        }
                    )
                    logo_path = None
            
            # If logo file not found, log warning and return None
            if logo_path is None:
                logger.warning(
                    "Logo file not found for export. Using text-only header.",
                    extra={
                        'expected_paths': ['static/images/norsu-logo.png', 'static/images/logo_norsu.png'],
                        'static_root': getattr(settings, 'STATIC_ROOT', None),
                        'debug': settings.DEBUG,
                        'fallback': 'text-only header'
                    }
                )
                return None
            
            # Validate file exists (double-check)
            try:
                if not os.path.exists(logo_path):
                    logger.warning(
                        f"Logo file path resolved but file does not exist: {logo_path}",
                        extra={
                            'logo_path': logo_path,
                            'fallback': 'text-only header'
                        }
                    )
                    return None
            except Exception as exists_error:
                logger.error(
                    f"Error checking logo file existence: {str(exists_error)}",
                    extra={
                        'logo_path': logo_path,
                        'error_type': type(exists_error).__name__,
                        'fallback': 'text-only header'
                    }
                )
                return None
            
            # Validate file size
            try:
                file_size = os.path.getsize(logo_path)
                if file_size > cls.MAX_LOGO_SIZE:
                    logger.warning(
                        f"Logo file exceeds maximum size ({cls.MAX_LOGO_SIZE} bytes). Using text-only header.",
                        extra={
                            'logo_path': logo_path,
                            'size': file_size,
                            'max_size': cls.MAX_LOGO_SIZE,
                            'fallback': 'text-only header'
                        }
                    )
                    return None
                
                if file_size == 0:
                    logger.warning(
                        "Logo file is empty (0 bytes). Using text-only header.",
                        extra={
                            'logo_path': logo_path,
                            'fallback': 'text-only header'
                        }
                    )
                    return None
            except OSError as size_error:
                logger.error(
                    f"Error checking logo file size: {str(size_error)}",
                    extra={
                        'logo_path': logo_path,
                        'error_type': type(size_error).__name__,
                        'fallback': 'text-only header'
                    }
                )
                return None
            
            # Validate file format by attempting to open with PIL
            try:
                from PIL import Image
                with Image.open(logo_path) as img:
                    img.verify()  # Verify it's a valid image
                    # Re-open to get format info (verify() closes the file)
                with Image.open(logo_path) as img:
                    img_format = img.format
                    if img_format not in ['PNG', 'JPEG', 'JPG']:
                        logger.warning(
                            f"Logo file format '{img_format}' may not be supported. Proceeding with caution.",
                            extra={
                                'logo_path': logo_path,
                                'format': img_format,
                                'supported_formats': ['PNG', 'JPEG', 'JPG']
                            }
                        )
            except ImportError as import_error:
                logger.error(
                    f"PIL/Pillow library not available for image validation: {str(import_error)}",
                    extra={
                        'logo_path': logo_path,
                        'error_type': type(import_error).__name__,
                        'fallback': 'Skipping image validation, proceeding with logo'
                    }
                )
                # Continue without validation if PIL is not available
            except Exception as img_error:
                logger.error(
                    f"Logo file is corrupted or invalid: {str(img_error)}",
                    extra={
                        'logo_path': logo_path,
                        'error_type': type(img_error).__name__,
                        'fallback': 'text-only header'
                    }
                )
                return None
            
            # Cache the valid logo path
            cls._logo_path_cache = logo_path
            logger.info(
                f"Logo file loaded successfully: {logo_path}",
                extra={
                    'logo_path': logo_path,
                    'file_size': file_size
                }
            )
            return logo_path
            
        except Exception as e:
            logger.error(
                f"Unexpected error accessing logo file: {str(e)}",
                extra={
                    'error_type': type(e).__name__,
                    'expected_paths': ['static/images/norsu-logo.png', 'static/images/logo_norsu.png'],
                    'fallback': 'text-only header'
                }
            )
            return None
    
    @staticmethod
    def add_pdf_header(canvas_obj, doc, logo_path: Optional[str], 
                      title: str = "NORSU Alumni System") -> None:
        """
        Add logo header to PDF page with comprehensive error handling.
        
        Args:
            canvas_obj: ReportLab canvas object
            doc: ReportLab document object
            logo_path: Path to logo file (None for text-only)
            title: Title text for header
        """
        try:
            canvas_obj.saveState()
            
            # Get page dimensions
            try:
                page_width, page_height = doc.pagesize
            except Exception as page_error:
                logger.error(
                    f"Error getting page dimensions: {str(page_error)}",
                    extra={
                        'error_type': type(page_error).__name__,
                        'fallback': 'Using default A4 dimensions'
                    }
                )
                # Use default A4 dimensions as fallback
                page_width, page_height = A4
            
            # NORSU blue color
            try:
                norsu_blue = colors.HexColor('#2b3c6b')
                gray_color = colors.HexColor('#4a5568')
            except Exception as color_error:
                logger.error(
                    f"Error creating colors: {str(color_error)}",
                    extra={
                        'error_type': type(color_error).__name__,
                        'fallback': 'Using default black color'
                    }
                )
                norsu_blue = colors.black
                gray_color = colors.black
            
            # Logo position and size
            logo_x = 30
            logo_y = page_height - 70
            max_logo_size = 45  # Maximum size for logo box
            
            # Text position (next to logo)
            text_x = 85
            institution_y = page_height - 50
            system_y = page_height - 62
            
            # Draw logo if available
            if logo_path:
                try:
                    # Verify file still exists before drawing
                    if not os.path.exists(logo_path):
                        logger.warning(
                            f"Logo file no longer exists at draw time: {logo_path}",
                            extra={
                                'logo_path': logo_path,
                                'fallback': 'text-only header'
                            }
                        )
                    else:
                        # Calculate logo dimensions maintaining aspect ratio
                        try:
                            from PIL import Image
                            with Image.open(logo_path) as img:
                                img_width, img_height = img.size
                                aspect_ratio = img_width / img_height
                                
                                # Calculate dimensions to fit within max_logo_size box
                                if aspect_ratio > 1:
                                    # Wider than tall
                                    logo_width = max_logo_size
                                    logo_height = max_logo_size / aspect_ratio
                                else:
                                    # Taller than wide or square
                                    logo_height = max_logo_size
                                    logo_width = max_logo_size * aspect_ratio
                        except Exception as img_error:
                            logger.warning(
                                f"Could not determine logo dimensions, using default: {str(img_error)}",
                                extra={
                                    'logo_path': logo_path,
                                    'error_type': type(img_error).__name__,
                                    'fallback': 'Using square dimensions'
                                }
                            )
                            # Fallback to square dimensions
                            logo_width = max_logo_size
                            logo_height = max_logo_size
                        
                        canvas_obj.drawImage(
                            logo_path,
                            logo_x,
                            logo_y,
                            width=logo_width,
                            height=logo_height,
                            preserveAspectRatio=True,
                            mask='auto'
                        )
                except FileNotFoundError as file_error:
                    logger.error(
                        f"Logo file not found during PDF drawing: {str(file_error)}",
                        extra={
                            'logo_path': logo_path,
                            'error_type': type(file_error).__name__,
                            'fallback': 'text-only header'
                        }
                    )
                except IOError as io_error:
                    logger.error(
                        f"IO error reading logo file for PDF: {str(io_error)}",
                        extra={
                            'logo_path': logo_path,
                            'error_type': type(io_error).__name__,
                            'fallback': 'text-only header'
                        }
                    )
                except Exception as logo_error:
                    logger.error(
                        f"Error drawing logo on PDF canvas: {str(logo_error)}",
                        extra={
                            'logo_path': logo_path,
                            'position': (logo_x, logo_y),
                            'error_type': type(logo_error).__name__,
                            'fallback': 'text-only header'
                        }
                    )
            
            # Draw institutional name
            try:
                canvas_obj.setFont('Helvetica-Bold', 11)
                canvas_obj.setFillColor(norsu_blue)
                canvas_obj.drawString(text_x, institution_y, "Negros Oriental State University")
            except Exception as text_error:
                logger.error(
                    f"Error drawing institutional name: {str(text_error)}",
                    extra={
                        'error_type': type(text_error).__name__,
                        'fallback': 'Attempting to continue with system name'
                    }
                )
            
            # Draw system name
            try:
                canvas_obj.setFont('Helvetica', 8)
                canvas_obj.setFillColor(gray_color)
                canvas_obj.drawString(text_x, system_y, "Alumni Management System")
            except Exception as text_error:
                logger.error(
                    f"Error drawing system name: {str(text_error)}",
                    extra={
                        'error_type': type(text_error).__name__,
                        'fallback': 'Attempting to continue with separator line'
                    }
                )
            
            # Draw separator line
            try:
                line_y = page_height - 75
                canvas_obj.setStrokeColor(norsu_blue)
                canvas_obj.setLineWidth(0.5)
                canvas_obj.line(30, line_y, page_width - 30, line_y)
            except Exception as line_error:
                logger.error(
                    f"Error drawing separator line: {str(line_error)}",
                    extra={
                        'error_type': type(line_error).__name__,
                        'fallback': 'Continuing without separator line'
                    }
                )
            
            canvas_obj.restoreState()
            
        except Exception as e:
            logger.error(
                f"Critical error adding PDF header: {str(e)}",
                extra={
                    'error_type': type(e).__name__,
                    'title': title,
                    'fallback': 'Export continues without header'
                }
            )
            # Attempt to restore canvas state even on error
            try:
                canvas_obj.restoreState()
            except:
                pass  # If restoreState fails, continue anyway
    
    @staticmethod
    def add_excel_header(worksheet, logo_path: Optional[str],
                        title: str = "NORSU Alumni System") -> int:
        """
        Add logo header to Excel worksheet with comprehensive error handling.
        
        Args:
            worksheet: openpyxl worksheet object
            logo_path: Path to logo file (None for text-only)
            title: Title text for header
            
        Returns:
            int: Row number where data should start (row 4 on success, row 1 on failure)
        """
        try:
            from openpyxl.drawing.image import Image as ExcelImage
            from openpyxl.styles import Font, Alignment
            
            # NORSU blue color
            norsu_blue = '2b3c6b'
            gray_color = '4a5568'
            
            # Embed logo if available
            if logo_path:
                try:
                    # Verify file still exists before embedding
                    if not os.path.exists(logo_path):
                        logger.warning(
                            f"Logo file no longer exists at embed time: {logo_path}",
                            extra={
                                'logo_path': logo_path,
                                'fallback': 'text-only header'
                            }
                        )
                    else:
                        img = ExcelImage(logo_path)
                        
                        # Calculate logo dimensions maintaining aspect ratio
                        # Ensure logo fits within 50x50 pixel box
                        max_logo_size = 50
                        try:
                            from PIL import Image as PILImage
                            with PILImage.open(logo_path) as pil_img:
                                img_width, img_height = pil_img.size
                                aspect_ratio = img_width / img_height
                                
                                # Calculate dimensions to fit within max_logo_size box
                                if aspect_ratio > 1:
                                    # Wider than tall
                                    logo_width = max_logo_size
                                    logo_height = max_logo_size / aspect_ratio
                                else:
                                    # Taller than wide or square
                                    logo_height = max_logo_size
                                    logo_width = max_logo_size * aspect_ratio
                                
                                # Set image dimensions
                                img.width = logo_width
                                img.height = logo_height
                        except Exception as img_error:
                            logger.warning(
                                f"Could not determine logo dimensions for Excel, using default: {str(img_error)}",
                                extra={
                                    'logo_path': logo_path,
                                    'error_type': type(img_error).__name__,
                                    'fallback': 'Using square dimensions'
                                }
                            )
                            # Fallback to square dimensions
                            img.width = max_logo_size
                            img.height = max_logo_size
                        
                        img.anchor = 'A1'
                        worksheet.add_image(img)
                except FileNotFoundError as file_error:
                    logger.error(
                        f"Logo file not found during Excel embedding: {str(file_error)}",
                        extra={
                            'logo_path': logo_path,
                            'error_type': type(file_error).__name__,
                            'fallback': 'text-only header'
                        }
                    )
                except IOError as io_error:
                    logger.error(
                        f"IO error reading logo file for Excel: {str(io_error)}",
                        extra={
                            'logo_path': logo_path,
                            'error_type': type(io_error).__name__,
                            'fallback': 'text-only header'
                        }
                    )
                except ImportError as import_error:
                    logger.error(
                        f"Error importing openpyxl image module: {str(import_error)}",
                        extra={
                            'error_type': type(import_error).__name__,
                            'fallback': 'text-only header'
                        }
                    )
                except Exception as img_error:
                    logger.error(
                        f"Error embedding logo in Excel: {str(img_error)}",
                        extra={
                            'logo_path': logo_path,
                            'error_type': type(img_error).__name__,
                            'fallback': 'text-only header'
                        }
                    )
            
            # Merge cells for institutional name (B1:D1)
            try:
                worksheet.merge_cells('B1:D1')
                cell_b1 = worksheet['B1']
                cell_b1.value = "Negros Oriental State University"
                cell_b1.font = Font(bold=True, size=12, color=norsu_blue)
                cell_b1.alignment = Alignment(horizontal='left', vertical='center')
            except Exception as merge_error:
                logger.error(
                    f"Error creating institutional name cell: {str(merge_error)}",
                    extra={
                        'error_type': type(merge_error).__name__,
                        'fallback': 'Attempting to continue with system name'
                    }
                )
            
            # Merge cells for system name (B2:D2)
            try:
                worksheet.merge_cells('B2:D2')
                cell_b2 = worksheet['B2']
                cell_b2.value = "Alumni Management System"
                cell_b2.font = Font(size=9, color=gray_color)
                cell_b2.alignment = Alignment(horizontal='left', vertical='center')
            except Exception as merge_error:
                logger.error(
                    f"Error creating system name cell: {str(merge_error)}",
                    extra={
                        'error_type': type(merge_error).__name__,
                        'fallback': 'Attempting to continue with row heights'
                    }
                )
            
            # Set row heights
            try:
                worksheet.row_dimensions[1].height = 30
                worksheet.row_dimensions[2].height = 20
                worksheet.row_dimensions[3].height = 10  # Spacer row
            except Exception as height_error:
                logger.error(
                    f"Error setting row heights: {str(height_error)}",
                    extra={
                        'error_type': type(height_error).__name__,
                        'fallback': 'Using default row heights'
                    }
                )
            
            # Return row number where data headers should start
            return 4
            
        except ImportError as import_error:
            logger.error(
                f"Error importing required Excel modules: {str(import_error)}",
                extra={
                    'error_type': type(import_error).__name__,
                    'title': title,
                    'fallback': 'Returning row 1 for normal export'
                }
            )
            # Return row 1 if header fails - continue with normal export
            return 1
        except Exception as e:
            logger.error(
                f"Critical error adding Excel header: {str(e)}",
                extra={
                    'error_type': type(e).__name__,
                    'title': title,
                    'fallback': 'Returning row 1 for normal export'
                }
            )
            # Return row 1 if header fails - continue with normal export
            return 1


def get_nested_value(obj, field_name):
    """Get value from nested field (e.g., 'user__username')"""
    if '__' not in field_name:
        return getattr(obj, field_name, '')
    
    # Handle nested field access
    parts = field_name.split('__')
    current_obj = obj
    
    for part in parts:
        if current_obj is None:
            return ''
        current_obj = getattr(current_obj, part, None)
    
    return current_obj if current_obj is not None else ''

def calculate_table_width(field_labels, sample_data):
    """Calculate approximate table width to determine page orientation"""
    total_width = 0
    for i, label in enumerate(field_labels):
        # Base width on label length
        col_width = len(label) * 8
        
        # Check sample data for this column
        for row in sample_data:
            if i < len(row):
                content_length = len(str(row[i])) * 6
                col_width = max(col_width, content_length)
        
        # Cap width and add to total
        col_width = min(col_width, 120)
        total_width += col_width
    
    return total_width

def should_use_landscape(field_labels, sample_data):
    """Determine if landscape orientation is needed"""
    table_width = calculate_table_width(field_labels, sample_data)
    # If table width is more than 500 points, use landscape
    return table_width > 500

class ExportMixin:
    """Mixin class to add export functionality to views"""
    
    def export_csv(self, queryset, filename, field_names=None, field_labels=None):
        """Export queryset to CSV format"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        if field_names is None:
            # Get all field names from the model
            field_names = [field.name for field in queryset.model._meta.fields]
        
        if field_labels is None:
            field_labels = field_names
        
        writer = csv.writer(response)
        writer.writerow(field_labels)
        
        for obj in queryset:
            row = []
            for field_name in field_names:
                value = get_nested_value(obj, field_name)
                if hasattr(value, 'strftime'):  # Handle datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                row.append(str(value))
            writer.writerow(row)
        
        return response
    
    def export_excel(self, queryset, filename, field_names=None, field_labels=None, sheet_name="Data"):
        """Export queryset to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if field_names is None:
            field_names = [field.name for field in queryset.model._meta.fields]
        
        if field_labels is None:
            field_labels = field_names
        
        # Get logo path using LogoHeaderService
        logo_path = LogoHeaderService.get_logo_path()
        
        # Add logo header and get starting row number
        header_start_row = LogoHeaderService.add_excel_header(ws, logo_path, title="NORSU Alumni System")
        
        # Style for header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header row at the returned row number (row 4)
        for col, label in enumerate(field_labels, 1):
            cell = ws.cell(row=header_start_row, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows starting at row 5 (header_start_row + 1)
        data_start_row = header_start_row + 1
        for row_idx, obj in enumerate(queryset, data_start_row):
            for col, field_name in enumerate(field_names, 1):
                value = get_nested_value(obj, field_name)
                if hasattr(value, 'strftime'):  # Handle datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                ws.cell(row=row_idx, column=col, value=str(value))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            # Get column letter from first non-merged cell
            column_letter = None
            for cell in column:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            
            if column_letter is None:
                continue  # Skip if all cells in column are merged
            
            for cell in column:
                try:
                    if not isinstance(cell, MergedCell) and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response
    
    def export_pdf(self, queryset, filename, field_names=None, field_labels=None, title="Data Export"):
        """Export queryset to PDF format with black and white styling"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        if field_names is None:
            field_names = [field.name for field in queryset.model._meta.fields]
        
        if field_labels is None:
            field_labels = field_names
        
        # Get logo path using LogoHeaderService
        logo_path = LogoHeaderService.get_logo_path()
        
        # Prepare sample data to determine page orientation
        sample_data = []
        sample_queryset = queryset[:10]  # Get first 10 records for width calculation
        
        for obj in sample_queryset:
            row = []
            for field_name in field_names:
                value = get_nested_value(obj, field_name)
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                value_str = str(value)
                # Don't truncate in sample data for width calculation
                row.append(value_str)
            sample_data.append(row)
        
        # Determine page orientation
        use_landscape = should_use_landscape(field_labels, sample_data)
        
        # Create PDF document with appropriate orientation
        if use_landscape:
            pagesize = landscape(A4)
        else:
            pagesize = A4
        
        # Prepare export info for footer
        export_info = f"Exported on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {queryset.count()}"
        if use_landscape:
            export_info += " | Landscape Mode"
        
        # Custom canvas class to add header and footer
        class HeaderFooterCanvas(canvas.Canvas):
            def __init__(self, *args, **kwargs):
                canvas.Canvas.__init__(self, *args, **kwargs)
                self.pages = []
                
            def showPage(self):
                self.pages.append(dict(self.__dict__))
                self._startPage()
                
            def save(self):
                page_count = len(self.pages)
                for page_num, page in enumerate(self.pages, 1):
                    self.__dict__.update(page)
                    # Draw header on each page
                    LogoHeaderService.add_pdf_header(self, doc, logo_path, title="NORSU Alumni System")
                    # Draw footer on each page
                    self.draw_footer(page_num, page_count, export_info, pagesize)
                    canvas.Canvas.showPage(self)
                canvas.Canvas.save(self)
                
            def draw_footer(self, page_num, page_count, info_text, page_size):
                # Draw export info in bottom right corner
                self.saveState()
                self.setFont('Helvetica', 7)
                self.setFillColor(colors.black)
                
                # Position text in bottom right corner
                text_width = self.stringWidth(info_text, 'Helvetica', 7)
                x_position = page_size[0] - text_width - 15  # 15 points from right edge
                y_position = 15  # 15 points from bottom
                
                self.drawString(x_position, y_position, info_text)
                self.restoreState()
            
        doc = SimpleDocTemplate(
            response, 
            pagesize=pagesize, 
            rightMargin=15, 
            leftMargin=15, 
            topMargin=80,  # Increased from 25 to 80 to accommodate header
            bottomMargin=25
        )
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=12,
            alignment=TA_CENTER,
            textColor=colors.black
        )
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 18))
        
        # Prepare table data with Paragraph objects for text wrapping
        table_data = []
        
        # Header row with Paragraph objects
        header_row = []
        cell_style = ParagraphStyle(
            'HeaderCell',
            parent=styles['Normal'],
            fontSize=7 if use_landscape else 8,
            textColor=colors.white,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        for label in field_labels:
            header_row.append(Paragraph(str(label), cell_style))
        table_data.append(header_row)
        
        # Limit data rows to prevent memory issues and improve performance
        max_rows = 1000  # Limit to 1000 rows per page
        queryset_limited = queryset[:max_rows]
        
        # Data cell style
        data_cell_style = ParagraphStyle(
            'DataCell',
            parent=styles['Normal'],
            fontSize=6 if use_landscape else 7,
            textColor=colors.black,
            alignment=TA_LEFT,
            fontName='Helvetica',
            leading=8 if use_landscape else 9
        )
        
        for obj in queryset_limited:
            row = []
            for field_name in field_names:
                value = get_nested_value(obj, field_name)
                if hasattr(value, 'strftime'):  # Handle datetime fields
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                # Use Paragraph for automatic text wrapping
                value_str = str(value)
                row.append(Paragraph(value_str, data_cell_style))
            table_data.append(row)
        
        # Calculate column widths dynamically based on content
        available_width = pagesize[0] - 30  # Total width minus margins
        num_cols = len(field_names)
        
        # Calculate relative widths based on content
        col_widths = []
        total_weight = 0
        
        for i, field_name in enumerate(field_names):
            # Base weight on field label length
            label_weight = len(field_labels[i])
            
            # Check data content for this column (sample first 20 rows)
            max_content_length = label_weight
            for row_idx, row in enumerate(sample_data):
                if row_idx >= 20:  # Only check first 20 rows
                    break
                if i < len(row):
                    content_length = len(str(row[i]))
                    max_content_length = max(max_content_length, content_length)
            
            # Use the maximum but cap it
            col_weight = min(max_content_length, 50)
            col_widths.append(col_weight)
            total_weight += col_weight
        
        # Convert weights to actual widths
        final_col_widths = []
        for weight in col_widths:
            width = (weight / total_weight) * available_width
            # Ensure minimum width
            width = max(width, 30)
            final_col_widths.append(width)
        
        # Create table with calculated column widths
        table = Table(table_data, colWidths=final_col_widths, repeatRows=1)
        
        # Black and white table styling
        table.setStyle(TableStyle([
            # Header styling - Black background with white text
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7 if use_landscape else 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data row styling - White background with black text
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6 if use_landscape else 7),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            
            # Grid styling - Black lines
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            
            # Alternating row backgrounds - White and light gray
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8E8E8')]),
            
            # Text wrapping and alignment
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        elements.append(table)
        
        # Add note if data was truncated
        if queryset.count() > max_rows:
            elements.append(Spacer(1, 10))
            note_style = ParagraphStyle(
                'Note',
                parent=styles['Normal'],
                fontSize=7,
                textColor=colors.black,
                alignment=TA_LEFT
            )
            elements.append(Paragraph(f"Note: Showing first {max_rows} records out of {queryset.count()} total records.", note_style))
        
        # Build PDF with custom canvas
        doc.build(elements, canvasmaker=HeaderFooterCanvas)
        return response

class ModelExporter:
    """Class to handle model-specific export configurations"""
    
    @staticmethod
    def get_alumni_export_config():
        """Get export configuration for Alumni model"""
        return {
            'field_names': [
                'id', 'user__username', 'user__email', 'user__first_name', 'user__last_name',
                'college', 'graduation_year', 'course', 'current_company', 'job_title', 'is_verified',
                'created_at', 'updated_at'
            ],
            'field_labels': [
                'ID', 'Username', 'Email', 'First Name', 'Last Name',
                'College', 'Graduation Year', 'Course', 'Current Company', 'Job Title', 'Verified',
                'Created At', 'Updated At'
            ],
            'sheet_name': 'Alumni Data'
        }
    
    @staticmethod
    def get_job_export_config():
        """Get export configuration for JobPosting model"""
        return {
            'field_names': [
                'id', 'title', 'company', 'location', 'job_type', 'salary_range',
                'description', 'requirements', 'is_active', 'created_at', 'updated_at'
            ],
            'field_labels': [
                'ID', 'Title', 'Company', 'Location', 'Job Type', 'Salary Range',
                'Description', 'Requirements', 'Active', 'Created At', 'Updated At'
            ],
            'sheet_name': 'Job Postings'
        }
    
    @staticmethod
    def get_mentorship_export_config():
        """Get export configuration for MentorshipRequest model"""
        return {
            'field_names': [
                'id', 'mentor__user__username', 'mentee__username',
                'status', 'created_at', 'updated_at'
            ],
            'field_labels': [
                'ID', 'Mentor', 'Mentee', 'Status', 'Created At', 'Updated At'
            ],
            'sheet_name': 'Mentorship Requests'
        }
    
    @staticmethod
    def get_event_export_config():
        """Get export configuration for Event model"""
        return {
            'field_names': [
                'id', 'title', 'description', 'start_date', 'end_date',
                'location', 'is_active', 'created_at', 'updated_at'
            ],
            'field_labels': [
                'ID', 'Title', 'Description', 'Start Date', 'End Date',
                'Location', 'Active', 'Created At', 'Updated At'
            ],
            'sheet_name': 'Events'
        }
    
    @staticmethod
    def get_donation_export_config():
        """Get export configuration for Donation model"""
        return {
            'field_names': [
                'id', 'donor__username', 'campaign__name', 'amount',
                'status', 'donation_date', 'created_at'
            ],
            'field_labels': [
                'ID', 'Donor', 'Campaign', 'Amount', 'Status', 'Donation Date', 'Created At'
            ],
            'sheet_name': 'Donations'
        }
    
    @staticmethod
    def get_user_export_config():
        """Get export configuration for User model"""
        return {
            'field_names': [
                'id', 'username', 'email', 'first_name', 'last_name',
                'is_active', 'is_staff', 'is_superuser', 'date_joined', 'last_login'
            ],
            'field_labels': [
                'ID', 'Username', 'Email', 'First Name', 'Last Name',
                'Active', 'Staff', 'Superuser', 'Date Joined', 'Last Login'
            ],
            'sheet_name': 'Users'
        }
    
    @staticmethod
    def get_announcement_export_config():
        """Get export configuration for Announcement model"""
        return {
            'field_names': [
                'id', 'title', 'content', 'category__name', 'priority_level',
                'is_active', 'date_posted', 'last_modified'
            ],
            'field_labels': [
                'ID', 'Title', 'Content', 'Category', 'Priority',
                'Active', 'Date Posted', 'Last Modified'
            ],
            'sheet_name': 'Announcements'
        }
    
    @staticmethod
    def get_feedback_export_config():
        """Get export configuration for Feedback model"""
        return {
            'field_names': [
                'id', 'user__username', 'subject', 'message', 'rating',
                'status', 'created_at', 'updated_at'
            ],
            'field_labels': [
                'ID', 'User', 'Subject', 'Message', 'Rating',
                'Status', 'Created At', 'Updated At'
            ],
            'sheet_name': 'Feedback'
        }
    
    @staticmethod
    def get_survey_export_config():
        """Get export configuration for Survey model"""
        return {
            'field_names': [
                'id', 'title', 'description', 'status', 'created_by__username',
                'created_at', 'updated_at'
            ],
            'field_labels': [
                'ID', 'Title', 'Description', 'Status', 'Created By',
                'Created At', 'Updated At'
            ],
            'sheet_name': 'Surveys'
        }

def get_export_filename(model_name, format_type):
    """Generate standardized export filename"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"{model_name}_{timestamp}"

def export_queryset(queryset, format_type, model_name, export_config=None):
    """Main export function that handles different formats"""
    if export_config is None:
        export_config = {}
    
    filename = get_export_filename(model_name, format_type)
    
    if format_type == 'csv':
        return ExportMixin().export_csv(
            queryset, 
            filename, 
            export_config.get('field_names'),
            export_config.get('field_labels')
        )
    elif format_type == 'excel':
        return ExportMixin().export_excel(
            queryset, 
            filename, 
            export_config.get('field_names'),
            export_config.get('field_labels'),
            export_config.get('sheet_name', 'Data')
        )
    elif format_type == 'pdf':
        return ExportMixin().export_pdf(
            queryset, 
            filename, 
            export_config.get('field_names'),
            export_config.get('field_labels'),
            export_config.get('sheet_name', 'Data Export')
        )
    else:
        raise ValueError(f"Unsupported format: {format_type}")
