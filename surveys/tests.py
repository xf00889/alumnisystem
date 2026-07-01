import json
from datetime import date, timedelta
from io import BytesIO
from types import SimpleNamespace

from django.contrib.auth import get_user_model
from django.conf import settings
from django.core.management import call_command
from django.test import RequestFactory, SimpleTestCase, TestCase, override_settings
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone

from alumni_directory.models import Alumni
from core.context_processors import tracer_study_banner_context
from surveys.management.commands.seed_tracer_study import ALUMNI_TITLE
from surveys.models import QuestionOption, ResponseAnswer, Survey, SurveyQuestion, SurveyResponse
from surveys.tracer_study import _answer_key, _filled_alumni_answers, _save_alumni_response


class _AnswerList(list):
    def select_related(self, *args):
        return self


class TracerStudyQuestionKeyFallbackTests(SimpleTestCase):
    def test_question_text_maps_to_filled_form_keys_without_metadata(self):
        cases = {
            "Contact Number (Mobile)": "p1_contact",
            "Date of Birth": "p1_dob",
            "Facebook Account": "p1_fb",
            "Vision (Extent of manifestation in your professional practice)": "p4_vision",
            "Mission (Extent of manifestation in your professional practice)": "p4_mission",
            "Goals (Extent of manifestation in your professional practice)": "p4_goals",
            "Core Values (Extent of manifestation in your professional practice)": "p4_core_values",
            "Program Objectives (Extent of manifestation in your professional practice)": "p4_program_objectives",
            "Negros Oriental State University VISION": "p4_vision",
            "Negros Oriental State University MISSION": "p4_mission",
            "Negros Oriental State University GOALS": "p4_goals",
            "NORSU Corporate Values/Core Values/Graduate Attributes": "p4_core_values",
            "Course/Program Objectives": "p4_program_objectives",
        }

        for text, expected in cases.items():
            question = SimpleNamespace(id=99, question_text=text, help_text="")
            self.assertEqual(_answer_key(question), expected)

    def test_saved_answers_render_in_filled_form_values(self):
        def question(text, key="", qid=99):
            return SimpleNamespace(
                id=qid,
                question_text=text,
                help_text=json.dumps({"key": key}) if key else "",
            )

        def answer(question_obj, text="", option="", rating=None):
            return SimpleNamespace(
                question=question_obj,
                text_answer=text,
                selected_option=SimpleNamespace(option_text=option) if option else None,
                rating_value=rating,
                custom_text="",
            )

        dob = "2000-06-30"
        expected_age = date.today().year - 2000 - ((date.today().month, date.today().day) < (6, 30))
        user = SimpleNamespace(
            get_full_name=lambda: "Test Alumni",
            username="test",
            email="test@example.com",
        )
        alumni = SimpleNamespace(
            user=user,
            date_of_birth=None,
            phone_number="",
            address="",
            city="",
            province="",
            country=None,
            course="",
            major="",
            gender="",
            campus="",
            graduation_year=None,
            employment_status="",
            job_title="",
            current_company="",
        )
        response = SimpleNamespace(
            alumni=alumni,
            answers=_AnswerList([
                answer(question("Contact Number (Mobile)"), text="+63 917 123 4567"),
                answer(question("Date of Birth"), text=dob),
                answer(question("Facebook Account"), text="https://facebook.com/full.profile.url"),
                answer(question("Reasons for accepting job", "p3_reasons_accept"), option="Salaries and benefits"),
                answer(question("How long did you stay in your first job?", "p3_first_job_duration"), option="1-2 years"),
                answer(question("Negros Oriental State University VISION"), rating=5),
                answer(question("Negros Oriental State University MISSION"), rating=4),
            ]),
        )

        filled = _filled_alumni_answers(response)

        self.assertEqual(filled["p1_contact"], "+63 917 123 4567")
        self.assertEqual(filled["p1_age"], expected_age)
        self.assertEqual(filled["p1_fb"], "https://facebook.com/full.profile.url")
        self.assertTrue(filled["accept_salary"])
        self.assertTrue(filled["duration_1_2y"])
        self.assertTrue(filled["p4_vision_5"])
        self.assertTrue(filled["p4_mission_4"])

        html = render_to_string(
            "tracer_study/filled_alumni_questionnaire.html",
            {
                "survey": SimpleNamespace(id=1),
                "filled": filled,
                "header_data_uri": "",
            },
        )
        vision_row = html.split("Negros Oriental State University <strong>VISION</strong>", 1)[1].split("</tr>", 1)[0]
        mission_row = html.split("Negros Oriental State University <strong>MISSION</strong>", 1)[1].split("</tr>", 1)[0]

        self.assertIn('value="+63 917 123 4567"', html)
        self.assertIn("https://facebook.com/full.profile.url", html)
        self.assertIn("checked", vision_row)
        self.assertIn("checked", mission_row)
        self.assertIn(".rating-table td input:checked", html)
        self.assertIn("radial-gradient(circle at center, #000", html)
        self.assertIn("border-width: 3.5pt", html)


class TracerStudySeedTests(TestCase):
    def test_alumni_required_and_conditional_questions_are_reachable(self):
        get_user_model().objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password",
        )

        call_command("seed_tracer_study", verbosity=0)

        survey = Survey.objects.get(title=ALUMNI_TITLE)
        questions = {str(question.id): question for question in survey.questions.prefetch_related("options")}

        for question in questions.values():
            meta = json.loads(question.help_text)
            if question.is_required:
                self.assertEqual(meta["show_when"], {}, question.question_text)

            for trigger_id, allowed_values in meta["show_when"].items():
                trigger = questions[trigger_id]
                self.assertEqual(trigger.question_type, "multiple_choice")
                labels = [
                    option.option_text.lower()
                    for option in trigger.options.all()
                ]
                for allowed in allowed_values:
                    self.assertTrue(
                        any(str(allowed).lower() in label for label in labels),
                        f"{question.question_text}: {allowed} not in {labels}",
                    )

    def test_change_job_reason_requires_employed_and_not_first_job(self):
        get_user_model().objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password",
        )

        call_command("seed_tracer_study", verbosity=0)

        survey = Survey.objects.get(title=ALUMNI_TITLE)
        by_key = {}
        for question in survey.questions.all():
            meta = json.loads(question.help_text)
            by_key[meta["key"]] = (question, meta)

        question, meta = by_key["p3_reasons_change"]
        self.assertEqual(
            meta["show_when"],
            {
                str(by_key["p3_employed"][0].id): ["yes"],
                str(by_key["p3_first_job"][0].id): ["no"],
            },
        )
        self.assertEqual(question.question_text, "Reasons for changing jobs (select all that apply)")


class TracerStudySaveAndFilledDisplayTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user("alumni", "alumni@example.com", "pass")
        self.alumni = Alumni.objects.create(
            user=self.user,
            college="CAS",
            campus="MAIN",
            graduation_year=2026,
            course="BSIT",
            gender="M",
            province="Negros Oriental",
            city="Dumaguete",
            address="A",
        )
        self.survey = Survey.objects.create(
            title=ALUMNI_TITLE,
            description="Tracer",
            created_by=self.user,
            start_date=timezone.now(),
            end_date=timezone.now() + timedelta(days=7),
            status="active",
        )
        self.factory = RequestFactory()

    def make_question(self, key, text, question_type, order):
        return SurveyQuestion.objects.create(
            survey=self.survey,
            question_text=text,
            question_type=question_type,
            display_order=order,
            help_text=json.dumps({"key": key}),
        )

    def make_option(self, question, text):
        return QuestionOption.objects.create(
            question=question,
            option_text=text,
            display_order=question.options.count(),
        )

    def test_saved_answers_are_visible_in_filled_form_mapping(self):
        accept = self.make_question("p3_reasons_accept", "Reasons for accepting job", "checkbox", 1)
        salary = self.make_option(accept, "Salaries and benefits")
        duration = self.make_question("p3_first_job_duration", "How long did you stay in your first job?", "multiple_choice", 2)
        one_to_two = self.make_option(duration, "1-2 years")
        vision = self.make_question("p4_vision", "Vision (Extent of manifestation in your professional practice)", "rating", 3)

        request = self.factory.post("/", {
            f"question_{accept.id}_{salary.id}": "1",
            f"question_{duration.id}": str(one_to_two.id),
            f"question_{vision.id}": "5",
        })

        response, answer_count = _save_alumni_response(request, self.survey, self.alumni)

        self.assertEqual(answer_count, 3)
        self.assertEqual(ResponseAnswer.objects.filter(response=response).count(), 3)
        filled = _filled_alumni_answers(response)
        self.assertTrue(filled["accept_salary"])
        self.assertTrue(filled["duration_1_2y"])
        self.assertTrue(filled["p4_vision_5"])


@override_settings(MIDDLEWARE=[
    middleware for middleware in settings.MIDDLEWARE
    if middleware != 'setup.middleware.SetupRequiredMiddleware'
])
class TracerStudyReportResponseStatusTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_user("admin", "admin@example.com", "pass", is_staff=True)
        self.respondent_user = User.objects.create_user(
            "respondent",
            "respondent@example.com",
            "pass",
            first_name="Rina",
            last_name="Responded",
        )
        self.missing_user = User.objects.create_user(
            "missing",
            "missing@example.com",
            "pass",
            first_name="Nico",
            last_name="Missing",
        )
        self.survey = Survey.objects.create(
            title=ALUMNI_TITLE,
            description="Tracer",
            created_by=self.admin,
            start_date=timezone.now(),
            end_date=timezone.now() + timedelta(days=7),
            status="active",
        )
        self.respondent = Alumni.objects.create(
            user=self.respondent_user,
            college="CAS",
            campus="MAIN",
            graduation_year=2025,
            course="BSINT",
            gender="F",
            province="Negros Oriental",
            city="Dumaguete",
            address="A",
        )
        self.missing = Alumni.objects.create(
            user=self.missing_user,
            college="CAS",
            campus="MAIN",
            graduation_year=2026,
            course="BSCS",
            gender="M",
            province="Negros Oriental",
            city="Dumaguete",
            address="B",
        )
        SurveyResponse.objects.create(survey=self.survey, alumni=self.respondent)

    def test_report_shows_and_exports_response_status(self):
        self.client.force_login(self.admin)
        report_url = reverse("surveys:tracer_study_report", args=[self.survey.id])

        response = self.client.get(report_url)
        self.assertContains(response, "Rina Responded")
        self.assertContains(response, "Nico Missing")
        self.assertContains(response, "Export Excel")

        export = self.client.get(reverse("surveys:tracer_study_report_export", args=[self.survey.id]))
        self.assertEqual(
            export["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(export.content))
        self.assertEqual(workbook.sheetnames, ["Responded", "No Response"])
        responded_values = [cell.value for row in workbook["Responded"].iter_rows() for cell in row]
        missing_values = [cell.value for row in workbook["No Response"].iter_rows() for cell in row]
        self.assertIn("Tracer Study Response Status", responded_values)
        self.assertIn("Alumni Who Responded", responded_values)
        self.assertIn("Alumni With No Response", missing_values)
        self.assertIn("Rina Responded", responded_values)
        self.assertNotIn("Nico Missing", responded_values)
        self.assertIn("Nico Missing", missing_values)
        self.assertNotIn("Rina Responded", missing_values)


class TracerStudyBannerContextTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user("alumni", "alumni@example.com", "pass")
        self.user.profile.has_completed_registration = True
        self.user.profile.save(update_fields=["has_completed_registration"])
        self.alumni = Alumni.objects.create(
            user=self.user,
            college="CAS",
            campus="MAIN",
            graduation_year=2026,
            course="BSINT",
            gender="M",
            province="Negros Oriental",
            city="Dumaguete",
            address="A",
        )
        self.survey = Survey.objects.create(
            title=ALUMNI_TITLE,
            description="Tracer",
            created_by=self.user,
            start_date=timezone.now(),
            end_date=timezone.now() + timedelta(days=7),
            status="active",
        )
        self.factory = RequestFactory()

    def request_for_user(self):
        request = self.factory.get("/")
        request.user = self.user
        request.resolver_match = SimpleNamespace(url_name="home")
        return request

    def test_banner_disappears_after_tracer_response(self):
        self.assertTrue(tracer_study_banner_context(self.request_for_user())["show_tracer_study_banner"])

        SurveyResponse.objects.create(survey=self.survey, alumni=self.alumni)

        self.assertFalse(tracer_study_banner_context(self.request_for_user())["show_tracer_study_banner"])

    def test_base_template_offsets_content_when_banner_shows(self):
        request = self.request_for_user()
        html = render_to_string(
            "base.html",
            {
                "request": request,
                "user": self.user,
                "show_tracer_study_banner": True,
                "seo": {"title": "Test"},
            },
        )

        self.assertIn("class=\"tracer-study-banner no-print\"", html)
        self.assertIn("top: 64px", html)
        self.assertIn("margin-top: 138px", html)
