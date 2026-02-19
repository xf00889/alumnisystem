from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, 
    FormView, TemplateView
)
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.forms import modelformset_factory
from django.db.models import Count, Avg, Q, Sum
from django.db.models.functions import TruncDate
from django.db import transaction
from django.utils import timezone
import json
import logging

from alumni_directory.models import Alumni
from donations.models import Donation, Campaign
from accounts.models import Mentor
from alumni_groups.models import AlumniGroup
from .models import (
    Survey, SurveyQuestion, QuestionOption, SurveyResponse, 
    ResponseAnswer, EmploymentRecord, Achievement, Report
)
from .forms import (
    SurveyForm, SurveyQuestionForm, QuestionOptionForm, ResponseAnswerForm,
    EmploymentRecordForm, AchievementForm, ReportForm,
    SurveyQuestionFormSet, QuestionOptionFormSet
)

# Logger instance for surveys app
logger = logging.getLogger(__name__)


class StaffRequiredMixin(LoginRequiredMixin):
    """
    Mixin that requires user to be staff or superuser.
    Unlike staff_member_required decorator, this doesn't redirect to Django admin.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('surveys:survey_list_public')
        return super().dispatch(request, *args, **kwargs)


# Staff/Admin Views for Survey Management
@method_decorator(staff_member_required, name='dispatch')
class SurveyListView(ListView):
    model = Survey
    template_name = 'surveys/admin/survey_list.html'
    context_object_name = 'surveys'
    
    def get_context_data(self, **kwargs):
        import time
        from django.db import connection
        start_time = time.time()
        initial_query_count = len(connection.queries)
        
        # Log list view access
        logger.debug(
            f"Survey list view accessed: User={self.request.user.username}",
            extra={
                'user_id': self.request.user.id,
                'ip_address': self.request.META.get('REMOTE_ADDR'),
                'user_agent': self.request.META.get('HTTP_USER_AGENT', '')[:100],
                'action': 'list_view_access'
            }
        )
        
        context = super().get_context_data(**kwargs)
        
        # Get basic counts
        total_alumni = Alumni.objects.count()
        
        # Count unique alumni who have responded to any survey
        alumni_with_responses = Alumni.objects.filter(
            survey_responses__isnull=False
        ).distinct().count()
        
        # Calculate participation percentage
        participation_percentage = 0
        if total_alumni > 0:
            participation_percentage = round((alumni_with_responses / total_alumni) * 100, 1)
        
        # Count active surveys
        active_surveys = Survey.objects.filter(status='active').count()
        
        # Get response data by college
        college_data = SurveyResponse.objects.select_related('alumni').values(
            'alumni__college'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Format college data for the template
        formatted_college_data = []
        for item in college_data:
            if item['alumni__college']:
                formatted_college_data.append({
                    'college_name': item['alumni__college'],
                    'count': item['count']
                })
            
        # If no college data, add a default entry for testing/visualization
        if not formatted_college_data:
            # Add a sample entry if no real data exists
            formatted_college_data.append({
                'college_name': 'No College Data',
                'count': 1
            })
        
        # Get response data by graduation year (batch)
        batch_data = SurveyResponse.objects.values(
            'alumni__graduation_year'
        ).annotate(
            count=Count('id')
        ).order_by('alumni__graduation_year')
        
        # Get survey-specific response data
        survey_data = []
        for survey in self.get_queryset():
            responses_count = survey.responses.count()
            response_rate = 0
            if total_alumni > 0:
                response_rate = round((responses_count / total_alumni) * 100, 1)
            
            survey_data.append({
                'id': survey.id,
                'title': survey.title,
                'status': survey.status,
                'start_date': survey.start_date,
                'end_date': survey.end_date,
                'responses_count': responses_count,
                'response_rate': response_rate
            })
        
        # Add all data to context
        context.update({
            'total_alumni': total_alumni,
            'alumni_with_responses': alumni_with_responses,
            'participation_percentage': participation_percentage,
            'active_surveys': active_surveys,
            'college_data': formatted_college_data,
            'batch_data': batch_data,
            'survey_data': survey_data
        })
        
        # Log performance metrics
        elapsed_time = time.time() - start_time
        final_query_count = len(connection.queries)
        queries_executed = final_query_count - initial_query_count
        
        logger.debug(
            f"Survey list context built: Time={elapsed_time:.3f}s, Queries={queries_executed}, Surveys={len(survey_data)}",
            extra={
                'user_id': self.request.user.id,
                'elapsed_time': elapsed_time,
                'queries_executed': queries_executed,
                'surveys_count': len(survey_data),
                'total_alumni': total_alumni,
                'action': 'list_view_complete'
            }
        )
        
        # Log slow operations warning
        if elapsed_time > 2.0:
            logger.warning(
                f"Slow survey list view: Time={elapsed_time:.3f}s, Queries={queries_executed}",
                extra={
                    'user_id': self.request.user.id,
                    'elapsed_time': elapsed_time,
                    'queries_executed': queries_executed,
                    'threshold': 2.0,
                    'action': 'slow_operation'
                }
            )
        
        return context

class SurveyCreateView(StaffRequiredMixin, CreateView):
    model = Survey
    form_class = SurveyForm
    template_name = 'surveys/admin/survey_form.html'
    
    def get_success_url(self):
        """Return the URL to redirect to after successful form submission"""
        return reverse('surveys:survey_list')
    
    def dispatch(self, request, *args, **kwargs):
        """Check if user has permission to create surveys"""
        # Call parent dispatch which handles staff check
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # Log survey creation start
        logger.info(
            f"Survey creation started by user: {self.request.user.username}",
            extra={
                'user_id': self.request.user.id,
                'user_email': self.request.user.email,
                'survey_title': form.cleaned_data.get('title', ''),
                'is_external': self.request.POST.get('is_external') == 'true'
            }
        )
        
        try:
            # Set the created_by field to the current user
            form.instance.created_by = self.request.user
            
            # Get survey type and external URL
            is_external = self.request.POST.get('is_external') == 'true'
            form.instance.is_external = is_external
            
            # Save the survey first
            response = super().form_valid(form)
            
            questions_count = 0
            if not is_external:
                # Process questions data
                try:
                    questions_data = json.loads(self.request.POST.get('questions', '[]'))
                    questions_count = len(questions_data)
                    
                    for question_data in questions_data:
                        # Create question
                        question = SurveyQuestion.objects.create(
                            survey=self.object,
                            question_text=question_data['question_text'],
                            question_type=question_data['question_type'],
                            is_required=question_data['is_required'],
                            help_text=question_data.get('help_text', ''),
                            display_order=question_data['display_order']
                        )
                        
                        # Add scale type if applicable
                        if question_data.get('scale_type'):
                            question.scale_type = question_data['scale_type']
                            question.save()
                        
                        # Create options if applicable
                        if 'options' in question_data:
                            for option_data in question_data['options']:
                                QuestionOption.objects.create(
                                    question=question,
                                    option_text=option_data['option_text'],
                                    display_order=option_data['display_order']
                                )
                
                except json.JSONDecodeError as e:
                    # Log the error but don't prevent survey creation
                    logger.error(
                        f"Error processing questions data for survey creation: {str(e)}",
                        extra={
                            'user_id': self.request.user.id,
                            'survey_id': self.object.id,
                            'survey_title': self.object.title,
                            'error_type': 'JSONDecodeError',
                            'exc_info': True
                        }
                    )
            
            # Log survey creation success
            logger.info(
                f"Survey created successfully: ID={self.object.id}, Title={self.object.title}",
                extra={
                    'survey_id': self.object.id,
                    'survey_title': self.object.title,
                    'user_id': self.request.user.id,
                    'is_external': is_external,
                    'questions_count': questions_count
                }
            )
            
            # Add success message
            messages.success(self.request, f'Survey "{self.object.title}" has been created successfully!')
            
            # Return redirect response (not JSON since we removed AJAX)
            return response
            
        except Exception as e:
            logger.error(
                f"Unexpected error in survey creation: {str(e)}",
                extra={
                    'user_id': self.request.user.id,
                    'survey_title': form.cleaned_data.get('title', ''),
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            messages.error(self.request, f'Error creating survey: {str(e)}')
            raise

class SurveyUpdateView(StaffRequiredMixin, UpdateView):
    model = Survey
    form_class = SurveyForm
    template_name = 'surveys/admin/survey_form.html'
    
    def get_success_url(self):
        """Return the URL to redirect to after successful form submission"""
        return reverse('surveys:survey_list')
    
    def get_context_data(self, **kwargs):
        """Add existing questions to context for editing"""
        context = super().get_context_data(**kwargs)
        if self.object:
            # Get questions with their options
            questions = self.object.questions.all().order_by('display_order').prefetch_related('options')
            questions_data = []
            for question in questions:
                question_dict = {
                    'id': question.id,
                    'question_text': question.question_text,
                    'question_type': question.question_type,
                    'is_required': question.is_required,
                    'help_text': question.help_text or '',
                    'scale_type': question.scale_type or '',
                    'display_order': question.display_order,
                    'options': []
                }
                # Add options if they exist
                for option in question.options.all().order_by('display_order'):
                    question_dict['options'].append({
                        'id': option.id,
                        'option_text': option.option_text,
                        'display_order': option.display_order
                    })
                questions_data.append(question_dict)
            context['existing_questions'] = json.dumps(questions_data)
        else:
            context['existing_questions'] = '[]'
        return context
    
    def dispatch(self, request, *args, **kwargs):
        """Check if user has permission to update surveys"""
        # Call parent dispatch which handles staff check
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        # Log survey update start
        logger.info(
            f"Survey update started by user: {self.request.user.username}",
            extra={
                'user_id': self.request.user.id,
                'survey_id': self.object.id,
                'survey_title': self.object.title
            }
        )
        
        try:
            # Store old values for comparison
            old_title = self.object.title
            old_status = self.object.status
            
            # Get survey type
            is_external = self.request.POST.get('is_external') == 'true'
            form.instance.is_external = is_external
            
            # Save the form
            response = super().form_valid(form)
            
            questions_count = 0
            if not is_external:
                # Process questions data
                try:
                    questions_data = json.loads(self.request.POST.get('questions', '[]'))
                    questions_count = len(questions_data)
                    
                    # Delete existing questions if new ones are provided
                    if questions_count > 0:
                        self.object.questions.all().delete()
                        
                        for question_data in questions_data:
                            # Create question
                            question = SurveyQuestion.objects.create(
                                survey=self.object,
                                question_text=question_data['question_text'],
                                question_type=question_data['question_type'],
                                is_required=question_data['is_required'],
                                help_text=question_data.get('help_text', ''),
                                display_order=question_data['display_order']
                            )
                            
                            # Add scale type if applicable
                            if question_data.get('scale_type'):
                                question.scale_type = question_data['scale_type']
                                question.save()
                            
                            # Create options if applicable
                            if 'options' in question_data:
                                for option_data in question_data['options']:
                                    QuestionOption.objects.create(
                                        question=question,
                                        option_text=option_data['option_text'],
                                        display_order=option_data['display_order']
                                    )
                
                except json.JSONDecodeError as e:
                    # Log the error but don't prevent survey update
                    logger.error(
                        f"Error processing questions data for survey update: {str(e)}",
                        extra={
                            'user_id': self.request.user.id,
                            'survey_id': self.object.id,
                            'survey_title': self.object.title,
                            'error_type': 'JSONDecodeError',
                            'exc_info': True
                        }
                    )
            
            # Log changes
            changes = []
            if old_title != self.object.title:
                changes.append(f"title: '{old_title}' -> '{self.object.title}'")
            if old_status != self.object.status:
                changes.append(f"status: '{old_status}' -> '{self.object.status}'")
            if questions_count > 0:
                changes.append(f"questions updated: {questions_count} questions")
            
            # Log survey update success
            logger.info(
                f"Survey updated successfully: ID={self.object.id}, Title={self.object.title}",
                extra={
                    'survey_id': self.object.id,
                    'survey_title': self.object.title,
                    'user_id': self.request.user.id,
                    'changes': changes if changes else 'No changes detected',
                    'questions_count': questions_count
                }
            )
            
            # Add success message
            messages.success(self.request, f'Survey "{self.object.title}" has been updated successfully!')
            
            return response
            
        except Exception as e:
            logger.error(
                f"Error updating survey: {str(e)}",
                extra={
                    'user_id': self.request.user.id,
                    'survey_id': self.object.id,
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            messages.error(self.request, f'Error updating survey: {str(e)}')
            raise

@method_decorator(staff_member_required, name='dispatch')
class SurveyDeleteView(DeleteView):
    model = Survey
    success_url = reverse_lazy('surveys:survey_list')
    
    def get(self, request, *args, **kwargs):
        """Redirect GET requests back to list - deletion should only happen via POST with AJAX"""
        messages.warning(request, 'Please use the delete button to remove surveys.')
        return HttpResponseRedirect(self.success_url)
    
    def post(self, request, *args, **kwargs):
        """Handle POST request for deletion"""
        return self.delete(request, *args, **kwargs)
    
    def delete(self, request, *args, **kwargs):
        survey = self.get_object()
        
        # Log survey deletion warning with details
        logger.warning(
            f"Survey deletion requested by user: {request.user.username}",
            extra={
                'user_id': request.user.id,
                'survey_id': survey.id,
                'survey_title': survey.title,
                'survey_status': survey.status,
                'responses_count': survey.responses.count(),
                'questions_count': survey.questions.count(),
                'action': 'survey_deletion'
            }
        )
        
        try:
            # Store survey details before deletion
            survey_id = survey.id
            survey_title = survey.title
            responses_count = survey.responses.count()
            questions_count = survey.questions.count()
            
            # Perform deletion
            survey.delete()
            
            # Log successful deletion
            logger.warning(
                f"Survey deleted successfully: ID={survey_id}, Title={survey_title}",
                extra={
                    'user_id': request.user.id,
                    'survey_id': survey_id,
                    'survey_title': survey_title,
                    'responses_count': responses_count,
                    'questions_count': questions_count,
                    'action': 'survey_deletion_complete'
                }
            )
            
            # Return JSON for AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Survey "{survey_title}" has been deleted successfully.'
                })
            
            messages.success(request, f'Survey "{survey_title}" has been deleted successfully.')
            return HttpResponseRedirect(self.success_url)
            
        except Exception as e:
            logger.error(
                f"Error deleting survey: {str(e)}",
                extra={
                    'survey_id': survey.id,
                    'survey_title': survey.title,
                    'user_id': request.user.id,
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            
            # Return JSON error for AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error deleting survey: {str(e)}'
                }, status=500)
            
            messages.error(request, f'Error deleting survey: {str(e)}')
            return HttpResponseRedirect(self.success_url)

@method_decorator(staff_member_required, name='dispatch')
class SurveyDetailView(DetailView):
    model = Survey
    template_name = 'surveys/admin/survey_detail.html'
    context_object_name = 'survey'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['questions'] = self.object.questions.all().prefetch_related('options')
        response_count = self.object.responses.count()
        context['response_count'] = response_count
        
        # Log survey detail access (DEBUG level)
        logger.debug(
            f"Survey detail accessed: Survey ID={self.object.id}, Title={self.object.title}",
            extra={
                'survey_id': self.object.id,
                'survey_title': self.object.title,
                'response_count': response_count,
                'questions_count': self.object.questions.count(),
                'user_id': self.request.user.id if hasattr(self.request, 'user') else None
            }
        )
        
        return context

@method_decorator(staff_member_required, name='dispatch')
class SurveyResponsesView(DetailView):
    model = Survey
    template_name = 'surveys/admin/survey_responses.html'
    context_object_name = 'survey'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        survey = self.object
        
        # Get all responses for this survey with related alumni data
        responses = survey.responses.select_related('alumni__user').prefetch_related(
            'answers', 'answers__question', 'answers__selected_option'
        ).order_by('-submitted_at')
        
        response_count = responses.count()
        
        # Log survey responses access (INFO level)
        logger.info(
            f"Survey responses accessed: Survey ID={survey.id}, Title={survey.title}, Response Count={response_count}",
            extra={
                'survey_id': survey.id,
                'survey_title': survey.title,
                'response_count': response_count,
                'user_id': self.request.user.id if hasattr(self.request, 'user') else None
            }
        )
        
        # Process responses for easy display in template
        processed_responses = []
        for response in responses:
            response_data = {
                'id': response.id,
                'alumni_name': f"{response.alumni.user.first_name} {response.alumni.user.last_name}",
                'submitted_at': response.submitted_at,
                'answers': []
            }
            
            # Group answers by question
            for question in survey.questions.all():
                answer_objects = [a for a in response.answers.all() if a.question_id == question.id]
                
                answer_value = None
                # Handle different types of questions
                if question.question_type in ['text', 'email', 'number', 'phone', 'url', 'date', 'time']:
                    # Text-based answers (stored in text_answer field)
                    if answer_objects and answer_objects[0].text_answer:
                        answer_value = answer_objects[0].text_answer
                    
                elif question.question_type == 'file':
                    # File upload
                    if answer_objects and answer_objects[0].file_answer:
                        answer_value = answer_objects[0].file_answer.url
                    
                elif question.question_type == 'multiple_choice':
                    # Single selection
                    if answer_objects and answer_objects[0].selected_option:
                        answer_value = answer_objects[0].selected_option.option_text
                        
                elif question.question_type == 'checkbox':
                    # Multiple selections
                    if answer_objects:
                        options = [a.selected_option.option_text for a in answer_objects if a.selected_option]
                        answer_value = ", ".join(options) if options else None
                        
                elif question.question_type in ['rating', 'likert']:
                    # Rating value
                    if answer_objects and answer_objects[0].rating_value is not None:
                        answer_value = answer_objects[0].rating_value
                
                # Always add the question to the response, even if no answer
                response_data['answers'].append({
                    'question': question.question_text,
                    'question_type': question.question_type,
                    'value': answer_value
                })
            
            processed_responses.append(response_data)
        
        # Calculate statistics for questions
        questions_stats = []
        for question in survey.questions.all():
            answers = ResponseAnswer.objects.filter(question=question)
            stat = {
                'question': question.question_text,
                'type': question.question_type,
                'total_answers': answers.count(),
            }
            
            if question.question_type in ['multiple_choice', 'checkbox']:
                # Calculate distribution of selected options
                stat['option_distribution'] = ResponseAnswer.objects.filter(
                    question=question
                ).values('selected_option__option_text').annotate(
                    count=Count('selected_option')
                ).exclude(selected_option__isnull=True)
                
            elif question.question_type in ['rating', 'likert']:
                # Calculate average rating
                avg_rating = answers.exclude(rating_value__isnull=True).aggregate(
                    avg=Avg('rating_value')
                )
                stat['rating_avg'] = round(avg_rating['avg'], 1) if avg_rating['avg'] else 0
                
                # Calculate distribution of ratings
                ratings = list(answers.exclude(rating_value__isnull=True).values_list('rating_value', flat=True))
                rating_distribution = {}
                for i in range(1, 6):  # Assuming 1-5 scale
                    rating_distribution[i] = len([r for r in ratings if r == i])
                stat['rating_distribution'] = rating_distribution
            
            questions_stats.append(stat)
        
        # Add to context
        context['responses'] = processed_responses
        context['questions_stats'] = questions_stats
        context['response_stats'] = {
            'total_responses': responses.count(),
            'response_rate': round((responses.count() / Alumni.objects.count()) * 100, 1) if Alumni.objects.count() > 0 else 0
        }
        
        return context

@method_decorator(staff_member_required, name='dispatch')
class SurveyQuestionCreateView(CreateView):
    model = SurveyQuestion
    form_class = SurveyQuestionForm
    template_name = 'surveys/admin/question_form.html'
    
    def get_success_url(self):
        return reverse('surveys:survey_detail', kwargs={'pk': self.kwargs['survey_id']})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        survey = get_object_or_404(Survey, pk=self.kwargs['survey_id'])
        context['survey'] = survey
        
        if self.request.POST:
            context['option_formset'] = QuestionOptionFormSet(self.request.POST)
        else:
            context['option_formset'] = QuestionOptionFormSet()
            
        return context
    
    def form_valid(self, form):
        survey = get_object_or_404(Survey, pk=self.kwargs['survey_id'])
        form.instance.survey = survey
        
        context = self.get_context_data()
        option_formset = context['option_formset']
        
        try:
            with transaction.atomic():
                self.object = form.save()
                
                options_count = 0
                if form.instance.question_type in ['multiple_choice', 'checkbox'] and option_formset.is_valid():
                    option_formset.instance = self.object
                    option_formset.save()
                    options_count = option_formset.total_form_count()
            
            # Log question creation
            logger.info(
                f"Survey question created: Question ID={self.object.id}, Type={form.instance.question_type}, Survey ID={survey.id}",
                extra={
                    'question_id': self.object.id,
                    'question_text': form.instance.question_text[:100],  # Truncate for logging
                    'question_type': form.instance.question_type,
                    'survey_id': survey.id,
                    'survey_title': survey.title,
                    'options_count': options_count,
                    'user_id': self.request.user.id if hasattr(self.request, 'user') else None
                }
            )
            
            return HttpResponseRedirect(self.get_success_url())
            
        except Exception as e:
            logger.error(
                f"Error creating survey question: {str(e)}",
                extra={
                    'survey_id': survey.id,
                    'question_type': form.instance.question_type,
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            raise

@method_decorator(staff_member_required, name='dispatch')
class SurveyQuestionUpdateView(UpdateView):
    model = SurveyQuestion
    form_class = SurveyQuestionForm
    template_name = 'surveys/admin/question_form.html'
    
    def get_success_url(self):
        return reverse('surveys:survey_detail', kwargs={'pk': self.object.survey.id})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['survey'] = self.object.survey
        
        if self.request.POST:
            context['option_formset'] = QuestionOptionFormSet(
                self.request.POST, 
                instance=self.object
            )
        else:
            context['option_formset'] = QuestionOptionFormSet(instance=self.object)
            
        return context
    
    def form_valid(self, form):
        # Store old values for comparison
        old_question_text = self.object.question_text
        old_question_type = self.object.question_type
        
        context = self.get_context_data()
        option_formset = context['option_formset']
        
        try:
            with transaction.atomic():
                self.object = form.save()
                
                options_count = 0
                if form.instance.question_type in ['multiple_choice', 'checkbox'] and option_formset.is_valid():
                    option_formset.instance = self.object
                    option_formset.save()
                    options_count = option_formset.total_form_count()
            
            # Log question update
            changes = []
            if old_question_text != form.instance.question_text:
                changes.append('question_text')
            if old_question_type != form.instance.question_type:
                changes.append(f'question_type: {old_question_type} -> {form.instance.question_type}')
            
            logger.info(
                f"Survey question updated: Question ID={self.object.id}, Type={form.instance.question_type}",
                extra={
                    'question_id': self.object.id,
                    'question_type': form.instance.question_type,
                    'survey_id': self.object.survey.id,
                    'options_count': options_count,
                    'changes': changes if changes else 'No changes detected',
                    'user_id': self.request.user.id if hasattr(self.request, 'user') else None
                }
            )
            
            return HttpResponseRedirect(self.get_success_url())
            
        except Exception as e:
            logger.error(
                f"Error updating survey question: {str(e)}",
                extra={
                    'question_id': self.object.id,
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            raise

# Alumni-facing Views for Survey Responses
class SurveyTakeView(LoginRequiredMixin, DetailView):
    model = Survey
    template_name = 'surveys/take_survey.html'
    context_object_name = 'survey'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        questions = self.object.questions.all().prefetch_related('options')
        context['questions'] = questions
        
        # Check if already responded
        alumni = get_object_or_404(Alumni, user=self.request.user)
        context['already_responded'] = SurveyResponse.objects.filter(
            survey=self.object, 
            alumni=alumni
        ).exists()
        
        return context
    
    def post(self, request, *args, **kwargs):
        survey = self.get_object()
        alumni = get_object_or_404(Alumni, user=request.user)
        
        # Log survey submission start with POST data
        logger.info(
            f"Survey submission started by user: {request.user.username}",
            extra={
                'user_id': request.user.id,
                'alumni_id': alumni.id,
                'survey_id': survey.id,
                'survey_title': survey.title,
                'post_data': dict(request.POST),
                'files_data': list(request.FILES.keys())
            }
        )
        
        # Check if already responded
        if SurveyResponse.objects.filter(survey=survey, alumni=alumni).exists():
            logger.warning(
                f"Duplicate survey submission attempt by user: {request.user.username}",
                extra={
                    'user_id': request.user.id,
                    'alumni_id': alumni.id,
                    'survey_id': survey.id,
                    'survey_title': survey.title,
                    'action': 'duplicate_submission'
                }
            )
            messages.error(request, "You have already completed this survey.")
            return redirect('surveys:survey_list_public')
        
        try:
            # Create survey response
            response = SurveyResponse.objects.create(
                survey=survey,
                alumni=alumni,
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            answer_count = 0
            # Process answers for each question
            for question in survey.questions.all():
                answer_data = {}
                field_name = f'question_{question.id}'
                field_value = request.POST.get(field_name)
                
                logger.info(
                    f"Processing question: {question.question_text}",
                    extra={
                        'question_id': question.id,
                        'question_type': question.question_type,
                        'field_name': field_name,
                        'field_value': field_value,
                        'field_value_type': type(field_value).__name__,
                        'is_required': question.is_required,
                        'post_keys': list(request.POST.keys())
                    }
                )
                
                if question.question_type in ['text', 'email', 'number', 'phone', 'url', 'date', 'time']:
                    # All text-based answers
                    text_answer = request.POST.get(f'question_{question.id}', '').strip()
                    if text_answer:  # Only save if there's actual content
                        answer_data['text_answer'] = text_answer
                        logger.info(f"Saving text answer for Q{question.id}: {text_answer}")
                    else:
                        logger.info(f"No text answer for Q{question.id} (value was empty or whitespace)")
                        
                elif question.question_type == 'file':
                    # File upload - Note: ResponseAnswer model doesn't have file_answer field
                    # Files would need to be stored differently
                    file_answer = request.FILES.get(f'question_{question.id}')
                    if file_answer:
                        # For now, store filename in text_answer
                        answer_data['text_answer'] = file_answer.name
                        logger.debug(f"File uploaded for Q{question.id}: {file_answer.name}")
                        
                elif question.question_type == 'multiple_choice':
                    option_id = request.POST.get(f'question_{question.id}')
                    if option_id:
                        try:
                            option = QuestionOption.objects.get(id=option_id)
                            answer_data['selected_option'] = option
                            logger.debug(f"Selected option for Q{question.id}: {option.option_text}")
                        except QuestionOption.DoesNotExist:
                            logger.error(f"Option {option_id} not found for Q{question.id}")
                        
                elif question.question_type == 'checkbox':
                    # Handle multiple selections
                    selected_options = []
                    for option in question.options.all():
                        if request.POST.get(f'question_{question.id}_{option.id}'):
                            selected_options.append(option)
                    
                    logger.debug(f"Checkbox selections for Q{question.id}: {len(selected_options)} options")
                    
                    # For checkbox, create multiple answers
                    for option in selected_options:
                        ResponseAnswer.objects.create(
                            response=response,
                            question=question,
                            selected_option=option
                        )
                        answer_count += 1
                    continue  # Skip the default creation below
                        
                elif question.question_type in ['rating', 'likert']:
                    # Rating and Likert scale
                    rating = request.POST.get(f'question_{question.id}')
                    if rating:
                        answer_data['rating_value'] = int(rating)
                        logger.debug(f"Rating for Q{question.id}: {rating}")
                
                # Create the answer with collected data
                if answer_data:
                    answer = ResponseAnswer.objects.create(
                        response=response,
                        question=question,
                        **answer_data
                    )
                    answer_count += 1
                    logger.info(f"Created answer {answer.id} for Q{question.id} with data: {answer_data}")
                else:
                    logger.info(f"No answer data for Q{question.id}, skipping answer creation")
                # Note: We don't create empty answers for unanswered questions
                # The absence of an answer indicates the question wasn't answered
            
            # Log successful submission
            logger.info(
                f"Survey submission completed successfully: Survey ID={survey.id}, Response ID={response.id}",
                extra={
                    'user_id': request.user.id,
                    'alumni_id': alumni.id,
                    'survey_id': survey.id,
                    'survey_title': survey.title,
                    'response_id': response.id,
                    'answers_count': answer_count,
                    'ip_address': request.META.get('REMOTE_ADDR')
                }
            )
            
            messages.success(request, "Thank you for completing the survey!")
            return redirect('surveys:survey_list_public')
            
        except Exception as e:
            logger.error(
                f"Error submitting survey: {str(e)}",
                extra={
                    'user_id': request.user.id,
                    'alumni_id': alumni.id,
                    'survey_id': survey.id,
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            messages.error(request, "An error occurred while submitting your survey. Please try again.")
            return redirect('surveys:survey_list_public')

class SurveyListPublicView(LoginRequiredMixin, ListView):
    model = Survey
    template_name = 'surveys/survey_list.html'
    context_object_name = 'surveys'
    
    def get_queryset(self):
        # Show all surveys with 'active' status, regardless of start/end dates
        # This ensures all active surveys are visible to alumni users
        queryset = Survey.objects.filter(
            status='active'
        ).order_by('end_date')
        
        # Log survey list access (DEBUG level)
        logger.debug(
            f"Survey list accessed by user: {self.request.user.username}",
            extra={
                'user_id': self.request.user.id,
                'surveys_count': queryset.count(),
                'action': 'survey_list_public'
            }
        )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Mark surveys that user has already responded to
        alumni = get_object_or_404(Alumni, user=self.request.user)
        responded_surveys = SurveyResponse.objects.filter(
            alumni=alumni
        ).values_list('survey_id', flat=True)
        
        context['responded_surveys'] = responded_surveys
        return context

# Report generation views
@method_decorator(staff_member_required, name='dispatch')
class ReportListView(ListView):
    model = Report
    template_name = 'surveys/admin/report_list.html'
    context_object_name = 'reports'

@method_decorator(staff_member_required, name='dispatch')
class ReportCreateView(CreateView):
    model = Report
    form_class = ReportForm
    template_name = 'surveys/admin/report_form.html'
    success_url = reverse_lazy('surveys:report_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Check data availability for each report type
        data_availability = {
            'employment': {
                'has_data': EmploymentRecord.objects.exists(),
                'count': EmploymentRecord.objects.count(),
                'message': 'No employment records found. Please add employment records before generating this report.'
            },
            'geographic': {
                'has_data': Alumni.objects.filter(
                    Q(country__isnull=False) | 
                    Q(province__isnull=False) | 
                    Q(city__isnull=False)
                ).exists(),
                'count': Alumni.objects.filter(
                    Q(country__isnull=False) | 
                    Q(province__isnull=False) | 
                    Q(city__isnull=False)
                ).count(),
                'message': 'No alumni with location data found. Please update alumni location information before generating this report.'
            },
            'achievements': {
                'has_data': Achievement.objects.exists(),
                'count': Achievement.objects.count(),
                'message': 'No achievements found. Please add achievements before generating this report.'
            },
            'feedback': {
                'has_data': SurveyResponse.objects.exists(),
                'count': SurveyResponse.objects.count(),
                'message': 'No survey responses found. Please ensure surveys have been completed before generating this report.'
            },
            'donations': {
                'has_data': Donation.objects.filter(status='completed').exists(),
                'count': Donation.objects.filter(status='completed').count(),
                'message': 'No completed donations found. Please ensure donations have been completed before generating this report.'
            },
            'mentors': {
                'has_data': Mentor.objects.filter(is_active=True).exists(),
                'count': Mentor.objects.filter(is_active=True).count(),
                'message': 'No active mentors found. Please ensure mentors have been added and activated.'
            },
            'groups': {
                'has_data': AlumniGroup.objects.filter(is_active=True).exists(),
                'count': AlumniGroup.objects.filter(is_active=True).count(),
                'message': 'No active groups found. Please ensure groups have been created and activated.'
            },
            'custom': {
                'has_data': True,
                'count': 0,
                'message': 'Custom reports can be created based on specific needs.'
            }
        }
        
        context['data_availability'] = data_availability
        
        # Get available filter options for each report type
        context['filter_options'] = {
            'employment': {
                'industries': EmploymentRecord.objects.values_list('industry', flat=True).distinct().order_by('industry'),
                'years': EmploymentRecord.objects.values_list('start_date__year', flat=True).distinct().order_by('-start_date__year'),
            },
            'geographic': {
                'countries': Alumni.objects.exclude(country__isnull=True).values_list('country', flat=True).distinct().order_by('country'),
                'provinces': Alumni.objects.exclude(province__isnull=True).exclude(province='').values_list('province', flat=True).distinct().order_by('province'),
            },
            'achievements': {
                'types': Achievement.ACHIEVEMENT_TYPES,
            },
            'donations': {
                'campaigns': Campaign.objects.all().order_by('name'),
            }
        }
        
        return context
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        
        # Explicitly set default values for scheduling fields to ensure they're always set
        form.instance.is_scheduled = False
        form.instance.schedule_frequency = 'none'
        
        # Process dynamic filter fields into parameters JSON
        parameters = {}
        for field_name in self.request.POST:
            if field_name not in ['csrfmiddlewaretoken', 'title', 'description', 'report_type', 'is_scheduled', 'schedule_frequency', 'last_scheduled_run', 'next_scheduled_run']:
                value = self.request.POST[field_name]
                if value:  # Only include non-empty values
                    parameters[field_name] = value
        
        form.instance.parameters = parameters
        
        return super().form_valid(form)

@method_decorator(staff_member_required, name='dispatch')
class ReportDetailView(DetailView):
    model = Report
    template_name = 'surveys/admin/report_detail.html'
    context_object_name = 'report'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report = self.get_object()
        
        # Log report access
        logger.info(
            f"Report detail view accessed: Report ID={report.id}, Type={report.report_type}",
            extra={
                'report_id': report.id,
                'report_title': report.title,
                'report_type': report.report_type,
                'user_id': self.request.user.id if hasattr(self.request, 'user') else None
            }
        )
        
        # Generate report data based on type
        try:
            report_data = self.generate_report_data(report)
            context['report_data'] = report_data
            context['report_error'] = None
            
            # Update last_run timestamp only if successful
            report.last_run = timezone.now()
            report.save(update_fields=['last_run'])
        except Exception as e:
            # Log the error and provide error context
            import traceback
            error_message = str(e)
            error_traceback = traceback.format_exc()
            
            logger.error(
                f"Error generating report data: {error_message}",
                extra={
                    'report_id': report.id,
                    'report_title': report.title,
                    'report_type': report.report_type,
                    'error_type': type(e).__name__,
                    'traceback': error_traceback,
                    'exc_info': True
                }
            )
            
            context['report_data'] = None
            context['report_error'] = {
                'message': error_message,
                'traceback': error_traceback
            }
        
        return context
    
    def generate_report_data(self, report):
        """
        Generate report data based on report type and parameters
        """
        import time
        start_time = time.time()
        
        # Log report generation start
        logger.info(
            f"Report generation started: Report ID={report.id}, Type={report.report_type}",
            extra={
                'report_id': report.id,
                'report_title': report.title,
                'report_type': report.report_type,
                'parameters': report.parameters if hasattr(report, 'parameters') else {}
            }
        )
        
        try:
            data = {
                'title': report.title,
                'type': report.report_type,
                'chart_data': {},
                'table_data': [],
                'summary': {}
            }
        except Exception as e:
            logger.error(
                f"Error initializing report data: {str(e)}",
                extra={
                    'report_id': report.id,
                    'report_type': report.report_type,
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            raise Exception(f"Error initializing report data: {str(e)}")
        
        if report.report_type == 'employment':
            # Employment trends report
            employment_records = EmploymentRecord.objects.all()
            
            # Apply filters from parameters
            if 'industry' in report.parameters and report.parameters['industry']:
                employment_records = employment_records.filter(
                    industry__icontains=report.parameters['industry']
                )
                
            if 'start_year' in report.parameters and report.parameters['start_year']:
                start_year = int(report.parameters['start_year'])
                employment_records = employment_records.filter(
                    start_date__year__gte=start_year
                )
                
            if 'end_year' in report.parameters and report.parameters['end_year']:
                end_year = int(report.parameters['end_year'])
                employment_records = employment_records.filter(
                    start_date__year__lte=end_year
                )
            
            # Industry distribution
            industry_counts = employment_records.values('industry').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # Create combined list for easier template iteration
            industries_list = [{'label': item['industry'], 'count': item['count']} for item in industry_counts]
            data['chart_data']['industries'] = {
                'labels': [item['industry'] for item in industry_counts],
                'data': [item['count'] for item in industry_counts],
                'items': industries_list,  # Combined structure for templates
            }
            
            # Job titles
            job_title_counts = employment_records.values('job_title').annotate(
                count=Count('id')
            ).order_by('-count')[:10]
            
            # Create combined list for easier template iteration
            job_titles_list = [{'label': item['job_title'], 'count': item['count']} for item in job_title_counts]
            data['chart_data']['job_titles'] = {
                'labels': [item['job_title'] for item in job_title_counts],
                'data': [item['count'] for item in job_title_counts],
                'items': job_titles_list,  # Combined structure for templates
            }
            
            # Table data - recent employment records
            data['table_data'] = list(employment_records.values(
                'alumni__user__first_name', 'alumni__user__last_name',
                'company_name', 'job_title', 'industry', 'start_date'
            ).order_by('-start_date')[:50])
            
            # Summary statistics
            data['summary'] = {
                'total_records': employment_records.count(),
                'unique_companies': employment_records.values('company_name').distinct().count(),
                'unique_industries': employment_records.values('industry').distinct().count(),
            }
            
        elif report.report_type == 'geographic':
            # Geographic distribution report
            alumni_locations = Alumni.objects.all()
            
            # Apply filters
            if 'country' in report.parameters and report.parameters['country']:
                alumni_locations = alumni_locations.filter(
                    country__icontains=report.parameters['country']
                )
                
            if 'state' in report.parameters and report.parameters['state']:
                alumni_locations = alumni_locations.filter(
                    province__icontains=report.parameters['state']
                )
            
            # Country distribution
            country_counts = alumni_locations.exclude(
                country__isnull=True
            ).values('country').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # Create combined list for easier template iteration
            countries_filtered = [item for item in country_counts if item['country']]
            countries_list = [{'label': str(item['country']), 'count': item['count']} for item in countries_filtered]
            data['chart_data']['countries'] = {
                'labels': [str(item['country']) for item in countries_filtered],
                'data': [item['count'] for item in countries_filtered],
                'items': countries_list,  # Combined structure for templates
            }
            
            # State/province distribution (for top country)
            if country_counts and country_counts[0].get('country'):
                top_country = country_counts[0]['country']
                state_counts = alumni_locations.filter(
                    country=top_country
                ).exclude(
                    province__isnull=True
                ).exclude(
                    province=''
                ).values('province').annotate(
                    count=Count('id')
                ).order_by('-count')
                
                data['chart_data']['states'] = {
                    'labels': [item['province'] for item in state_counts if item['province']],
                    'data': [item['count'] for item in state_counts if item['province']],
                    'country': str(top_country)
                }
            
            # Table data - alumni locations
            data['table_data'] = list(alumni_locations.values(
                'user__first_name', 'user__last_name',
                'city', 'province', 'country'
            ).order_by('country', 'province')[:100])
            
            # Summary statistics
            data['summary'] = {
                'total_alumni': alumni_locations.count(),
                'countries_count': alumni_locations.exclude(country__isnull=True).values('country').distinct().count(),
                'cities_count': alumni_locations.exclude(city__isnull=True).exclude(city='').values('city').distinct().count(),
            }
            
        elif report.report_type == 'achievements':
            # Alumni achievements report
            achievements = Achievement.objects.all()
            
            # Apply filters
            if 'achievement_type' in report.parameters and report.parameters['achievement_type']:
                achievements = achievements.filter(
                    achievement_type=report.parameters['achievement_type']
                )
                
            if 'start_date' in report.parameters and report.parameters['start_date']:
                achievements = achievements.filter(
                    achievement_date__gte=report.parameters['start_date']
                )
                
            if 'end_date' in report.parameters and report.parameters['end_date']:
                achievements = achievements.filter(
                    achievement_date__lte=report.parameters['end_date']
                )
            
            # Achievement types distribution
            type_counts = achievements.values('achievement_type').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # Create combined list for easier template iteration
            types_list = [
                {
                    'label': dict(Achievement.ACHIEVEMENT_TYPES).get(item['achievement_type'], item['achievement_type']),
                    'count': item['count']
                }
                for item in type_counts
            ]
            data['chart_data']['types'] = {
                'labels': [dict(Achievement.ACHIEVEMENT_TYPES).get(item['achievement_type'], item['achievement_type']) 
                          for item in type_counts],
                'data': [item['count'] for item in type_counts],
                'items': types_list,  # Combined structure for templates
            }
            
            # Achievements by year
            year_counts = achievements.extra(
                select={'year': "EXTRACT(year FROM achievement_date)"}
            ).values('year').annotate(
                count=Count('id')
            ).order_by('year')
            
            data['chart_data']['years'] = {
                'labels': [int(item['year']) for item in year_counts],
                'data': [item['count'] for item in year_counts],
            }
            
            # Table data - recent achievements
            data['table_data'] = list(achievements.values(
                'alumni__user__first_name', 'alumni__user__last_name',
                'title', 'achievement_type', 'achievement_date', 'verified'
            ).order_by('-achievement_date')[:50])
            
            # Summary statistics
            data['summary'] = {
                'total_achievements': achievements.count(),
                'verified_count': achievements.filter(verified=True).count(),
                'top_type': dict(Achievement.ACHIEVEMENT_TYPES).get(
                    type_counts[0]['achievement_type']
                ) if type_counts else None,
            }
            
        elif report.report_type == 'feedback':
            # Survey feedback report
            surveys = Survey.objects.all()
            survey_responses = SurveyResponse.objects.all()
            
            # Get survey completion rate
            total_alumni = Alumni.objects.count()
            response_counts = survey_responses.values('survey').annotate(
                count=Count('alumni', distinct=True)
            )
            
            completion_rates = []
            for response in response_counts:
                survey = Survey.objects.get(id=response['survey'])
                completion_rates.append({
                    'survey': survey.title,
                    'responses': response['count'],
                    'rate': round((response['count'] / total_alumni) * 100, 1) if total_alumni > 0 else 0
                })
            
            # Create combined list for easier template iteration
            completion_list = [{'label': item['survey'], 'rate': item['rate']} for item in completion_rates]
            data['chart_data']['completion'] = {
                'labels': [item['survey'] for item in completion_rates],
                'data': [item['rate'] for item in completion_rates],
                'items': completion_list,  # Combined structure for templates
            }
            
            # Get rating averages for rating questions
            rating_questions = SurveyQuestion.objects.filter(question_type='rating')
            rating_averages = []
            
            for question in rating_questions:
                avg_rating = ResponseAnswer.objects.filter(
                    question=question,
                    rating_value__isnull=False
                ).aggregate(avg=Avg('rating_value'))
                
                if avg_rating['avg']:
                    rating_averages.append({
                        'question': question.question_text,
                        'survey': question.survey.title,
                        'average': round(avg_rating['avg'], 2)
                    })
            
            data['chart_data']['ratings'] = {
                'labels': [f"{item['survey']}: {item['question'][:30]}..." for item in rating_averages],
                'data': [item['average'] for item in rating_averages],
            }
            
            # Table data - recent textual feedback
            text_answers = ResponseAnswer.objects.filter(
                text_answer__isnull=False
            ).exclude(
                text_answer=''
            ).select_related(
                'question', 'response__alumni', 'response__survey'
            ).order_by('-response__submitted_at')[:30]
            
            data['table_data'] = [
                {
                    'survey': answer.response.survey.title,
                    'question': answer.question.question_text,
                    'answer': answer.text_answer,
                    'alumni': f"{answer.response.alumni.user.first_name} {answer.response.alumni.user.last_name}",
                    'date': answer.response.submitted_at
                }
                for answer in text_answers
            ]
            
            # Summary statistics
            data['summary'] = {
                'total_surveys': surveys.count(),
                'total_responses': survey_responses.count(),
                'avg_completion_rate': round(
                    sum(item['rate'] for item in completion_rates) / len(completion_rates), 1
                ) if completion_rates else 0,
            }
            
        elif report.report_type == 'donations':
            # Donations report
            donations = Donation.objects.filter(status='completed')
            
            # Apply filters
            if 'campaign' in report.parameters and report.parameters['campaign']:
                donations = donations.filter(campaign_id=report.parameters['campaign'])
            
            if 'start_date' in report.parameters and report.parameters['start_date']:
                donations = donations.filter(donation_date__gte=report.parameters['start_date'])
            
            if 'end_date' in report.parameters and report.parameters['end_date']:
                donations = donations.filter(donation_date__lte=report.parameters['end_date'])
            
            if 'period' in report.parameters and report.parameters['period']:
                period = report.parameters['period']
                now = timezone.now()
                if period == 'today':
                    donations = donations.filter(donation_date__date=now.date())
                elif period == 'month':
                    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                    donations = donations.filter(donation_date__gte=start_of_month)
                elif period == 'year':
                    start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
                    donations = donations.filter(donation_date__gte=start_of_year)
            
            # Total donations
            total_amount = donations.aggregate(total=Sum('amount'))['total'] or 0
            
            # Donations by campaign
            campaign_totals = donations.values('campaign__name').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('-total')
            
            campaigns_list = [
                {'label': item['campaign__name'], 'total': float(item['total']), 'count': item['count']}
                for item in campaign_totals
            ]
            
            data['chart_data']['campaigns'] = {
                'labels': [item['campaign__name'] for item in campaign_totals],
                'data': [float(item['total']) for item in campaign_totals],
                'items': campaigns_list,
            }
            
            # Donations by date (daily)
            daily_donations = donations.annotate(
                date=TruncDate('donation_date')
            ).values('date').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('date')
            
            daily_items = []
            for item in daily_donations:
                date_str = str(item['date']) if item['date'] else ''
                daily_items.append({
                    'label': date_str,
                    'total': float(item['total']),
                    'count': item['count']
                })
            
            data['chart_data']['daily'] = {
                'labels': [item['label'] for item in daily_items],
                'data': [item['total'] for item in daily_items],
                'items': daily_items,
            }
            
            # Table data - recent donations
            data['table_data'] = list(donations.values(
                'donor__first_name', 'donor__last_name', 'donor_name',
                'campaign__name', 'amount', 'donation_date', 'payment_method'
            ).order_by('-donation_date')[:100])
            
            # Summary statistics
            data['summary'] = {
                'total_donations': donations.count(),
                'total_amount': float(total_amount),
                'campaigns_count': donations.values('campaign').distinct().count(),
                'average_donation': float(total_amount / donations.count()) if donations.count() > 0 else 0,
            }
            
        elif report.report_type == 'mentors':
            # Mentors report
            mentors = Mentor.objects.filter(is_active=True)
            
            # Availability distribution
            availability_counts = mentors.values('availability_status').annotate(
                count=Count('id')
            ).order_by('-count')
            
            availability_list = [
                {'label': item['availability_status'], 'count': item['count']}
                for item in availability_counts
            ]
            
            data['chart_data']['availability'] = {
                'labels': [item['availability_status'] for item in availability_counts],
                'data': [item['count'] for item in availability_counts],
                'items': availability_list,
            }
            
            # Verification status
            verified_count = mentors.filter(is_verified=True).count()
            unverified_count = mentors.filter(is_verified=False).count()
            
            data['chart_data']['verification'] = {
                'labels': ['Verified', 'Unverified'],
                'data': [verified_count, unverified_count],
                'items': [
                    {'label': 'Verified', 'count': verified_count},
                    {'label': 'Unverified', 'count': unverified_count},
                ],
            }
            
            # Table data - mentors list
            data['table_data'] = list(mentors.values(
                'user__first_name', 'user__last_name',
                'availability_status', 'is_verified', 'current_mentees', 
                'max_mentees', 'accepting_mentees', 'created_at'
            ).order_by('-created_at')[:100])
            
            # Summary statistics
            total_mentees = sum(mentor.current_mentees for mentor in mentors)
            data['summary'] = {
                'total_mentors': mentors.count(),
                'verified_mentors': verified_count,
                'unverified_mentors': unverified_count,
                'active_mentors': mentors.filter(accepting_mentees=True).count(),
                'total_mentees': total_mentees,
                'average_mentees': round(total_mentees / mentors.count(), 2) if mentors.count() > 0 else 0,
            }
            
        elif report.report_type == 'groups':
            # Groups report
            groups = AlumniGroup.objects.filter(is_active=True)
            
            # Group type distribution
            type_counts = groups.values('group_type').annotate(
                count=Count('id')
            ).order_by('-count')
            
            type_list = [
                {'label': dict(AlumniGroup.GROUP_TYPES).get(item['group_type'], item['group_type']), 'count': item['count']}
                for item in type_counts
            ]
            
            data['chart_data']['types'] = {
                'labels': [dict(AlumniGroup.GROUP_TYPES).get(item['group_type'], item['group_type']) for item in type_counts],
                'data': [item['count'] for item in type_counts],
                'items': type_list,
            }
            
            # Visibility distribution
            visibility_counts = groups.values('visibility').annotate(
                count=Count('id')
            ).order_by('-count')
            
            visibility_list = [
                {'label': dict(AlumniGroup.VISIBILITY_CHOICES).get(item['visibility'], item['visibility']), 'count': item['count']}
                for item in visibility_counts
            ]
            
            data['chart_data']['visibility'] = {
                'labels': [dict(AlumniGroup.VISIBILITY_CHOICES).get(item['visibility'], item['visibility']) for item in visibility_counts],
                'data': [item['count'] for item in visibility_counts],
                'items': visibility_list,
            }
            
            # Table data - groups list with member counts
            groups_data = []
            for group in groups.select_related('created_by').prefetch_related('memberships')[:100]:
                try:
                    # Safely get member count
                    member_count = group.memberships.count() if hasattr(group, 'memberships') else 0
                except Exception:
                    member_count = 0
                
                groups_data.append({
                    'name': group.name,
                    'group_type': group.group_type,
                    'visibility': group.visibility,
                    'created_by__first_name': group.created_by.first_name if group.created_by else '',
                    'created_by__last_name': group.created_by.last_name if group.created_by else '',
                    'member_count': member_count,
                    'created_at': group.created_at,
                })
            data['table_data'] = groups_data
            
            # Summary statistics - get member counts safely
            try:
                # Use already calculated member counts from groups_data
                total_members = sum(group.get('member_count', 0) for group in groups_data)
            except Exception:
                total_members = 0
            
            data['summary'] = {
                'total_groups': groups.count(),
                'public_groups': groups.filter(visibility='PUBLIC').count(),
                'private_groups': groups.filter(visibility='PRIVATE').count(),
                'restricted_groups': groups.filter(visibility='RESTRICTED').count(),
                'total_members': total_members,
                'average_members': round(total_members / groups.count(), 2) if groups.count() > 0 else 0,
            }
            
        elif report.report_type == 'custom':
            # Placeholder for custom reports
            data['summary'] = {
                'message': 'Custom reports can be implemented based on specific needs'
            }
        
        # Calculate generation time
        elapsed_time = time.time() - start_time
        
        # Log report generation success
        logger.info(
            f"Report generation completed successfully: Report ID={report.id}, Type={report.report_type}, Time={elapsed_time:.2f}s",
            extra={
                'report_id': report.id,
                'report_title': report.title,
                'report_type': report.report_type,
                'generation_time': elapsed_time,
                'table_data_count': len(data.get('table_data', []))
            }
        )
        
        # Log slow generation warning
        if elapsed_time > 5.0:
            logger.warning(
                f"Report generation took longer than 5 seconds: Report ID={report.id}, Time={elapsed_time:.2f}s",
                extra={
                    'report_id': report.id,
                    'report_type': report.report_type,
                    'generation_time': elapsed_time,
                    'threshold': 5.0
                }
            )
        
        return data


@login_required
@staff_member_required
def survey_export_responses(request, pk):
    """Export survey responses to Excel format with NORSU header"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.utils.text import slugify
    from core.export_utils import LogoHeaderService
    
    survey = get_object_or_404(Survey, pk=pk)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey Responses"
    
    # Add NORSU header with logo
    logo_path = LogoHeaderService.get_logo_path()
    start_row = LogoHeaderService.add_excel_header(
        ws, 
        logo_path, 
        title=f"{survey.title} - Responses"
    )
    
    # Get all questions for the survey
    questions = survey.questions.all().order_by('display_order')
    
    # Define header style
    header_fill = PatternFill(start_color='2b3c6b', end_color='2b3c6b', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write header row
    headers = ['Response ID', 'Alumni Name', 'Email', 'Submitted At']
    for question in questions:
        headers.append(f"Q{question.display_order + 1}: {question.question_text}")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Get all responses with related data
    responses = survey.responses.select_related('alumni__user').prefetch_related(
        'answers', 'answers__question', 'answers__selected_option'
    ).order_by('-submitted_at')
    
    # Define data style
    data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Write data rows
    current_row = start_row + 1
    for response in responses:
        # Basic response info
        ws.cell(row=current_row, column=1, value=response.id)
        ws.cell(row=current_row, column=2, value=f"{response.alumni.user.first_name} {response.alumni.user.last_name}")
        ws.cell(row=current_row, column=3, value=response.alumni.user.email)
        ws.cell(row=current_row, column=4, value=response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Add answers for each question
        col_num = 5
        for question in questions:
            answer_objects = [a for a in response.answers.all() if a.question_id == question.id]
            
            answer_value = ''
            if question.question_type in ['text', 'email', 'url', 'date', 'time']:
                if answer_objects and answer_objects[0].text_answer:
                    answer_value = answer_objects[0].text_answer
            elif question.question_type == 'number':
                # Store as number to avoid scientific notation
                if answer_objects and answer_objects[0].text_answer:
                    try:
                        answer_value = float(answer_objects[0].text_answer)
                    except (ValueError, TypeError):
                        answer_value = answer_objects[0].text_answer
            elif question.question_type == 'phone':
                # Store phone numbers as text with apostrophe prefix to prevent formatting
                if answer_objects and answer_objects[0].text_answer:
                    answer_value = answer_objects[0].text_answer
                    # Format as text to preserve leading zeros and prevent scientific notation
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = answer_value
                    cell.number_format = '@'  # Text format
                    cell.alignment = data_alignment
                    col_num += 1
                    continue
            elif question.question_type == 'multiple_choice':
                if answer_objects and answer_objects[0].selected_option:
                    answer_value = answer_objects[0].selected_option.option_text
            elif question.question_type == 'checkbox':
                if answer_objects:
                    options = [a.selected_option.option_text for a in answer_objects if a.selected_option]
                    answer_value = "; ".join(options) if options else ''
            elif question.question_type in ['rating', 'likert']:
                if answer_objects and answer_objects[0].rating_value is not None:
                    answer_value = answer_objects[0].rating_value
            elif question.question_type == 'file':
                # File uploads are stored in text_answer field as file path/URL
                if answer_objects and answer_objects[0].text_answer:
                    answer_value = answer_objects[0].text_answer
            
            cell = ws.cell(row=current_row, column=col_num, value=answer_value)
            cell.alignment = data_alignment
            col_num += 1
        
        current_row += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        # Check header length
        if col_num <= len(headers):
            max_length = len(str(headers[col_num - 1]))
        
        # Check data length (sample first 100 rows for performance)
        for row_num in range(start_row + 1, min(start_row + 101, current_row)):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width (max 50 characters)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create the HttpResponse object with Excel content
    http_response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    http_response['Content-Disposition'] = f'attachment; filename="{slugify(survey.title)}_responses.xlsx"'
    
    # Save workbook to response
    wb.save(http_response)
    
    # Log export action
    logger.info(
        f"Survey responses exported: Survey ID={survey.id}, Title={survey.title}, Response Count={responses.count()}",
        extra={
            'survey_id': survey.id,
            'survey_title': survey.title,
            'response_count': responses.count(),
            'user_id': request.user.id,
            'export_format': 'excel'
        }
    )
    
    return http_response


@staff_member_required
def report_export_pdf(request, pk):
    """
    Export report as PDF
    """
    import time
    start_time = time.time()
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfgen import canvas
    from core.export_utils import LogoHeaderService
    
    report = get_object_or_404(Report, pk=pk)
    
    # Log PDF export start
    logger.info(
        f"PDF export started: Report ID={report.id}, Type={report.report_type}",
        extra={
            'report_id': report.id,
            'report_title': report.title,
            'report_type': report.report_type,
            'user_id': request.user.id,
            'action': 'pdf_export'
        }
    )
    
    try:
        # Generate report data
        detail_view = ReportDetailView()
        detail_view.object = report
        report_data = detail_view.generate_report_data(report)
        
        # Get logo path for header
        logo_path = LogoHeaderService.get_logo_path()
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        filename = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Create custom canvas class with logo header
        class HeaderCanvas(canvas.Canvas):
            def __init__(self, *args, **kwargs):
                canvas.Canvas.__init__(self, *args, **kwargs)
            
            def showPage(self):
                # Add logo header to each page
                LogoHeaderService.add_pdf_header(
                    self, 
                    None,  # doc parameter not needed for this implementation
                    logo_path,
                    title="NORSU Alumni System - Survey Report"
                )
                canvas.Canvas.showPage(self)
        
        # Create PDF document with increased top margin
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=80,  # Increased from 30 to 80 for logo header
            bottomMargin=30
        )
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#417690')
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            textColor=colors.HexColor('#417690')
        )
        
        # Add title
        elements.append(Paragraph(report.title, title_style))
        elements.append(Spacer(1, 12))
        
        # Add report info
        info_data = [
            ['Type:', report.get_report_type_display()],
            ['Created by:', report.created_by.get_full_name() or report.created_by.username],
            ['Created at:', report.created_at.strftime('%Y-%m-%d %H:%M')],
            ['Last run:', report.last_run.strftime('%Y-%m-%d %H:%M') if report.last_run else 'Never'],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 15))
        
        # Add description if available
        if report.description:
            elements.append(Paragraph('Description:', styles['Heading3']))
            elements.append(Paragraph(report.description, styles['Normal']))
            elements.append(Spacer(1, 12))
        
        # Add summary statistics
        if report_data.get('summary'):
            elements.append(Paragraph('Summary Statistics', heading_style))
            summary_data = [['Metric', 'Value']]
            for key, value in report_data['summary'].items():
                summary_data.append([key.replace('_', ' ').title(), str(value)])
            
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 15))
        
        # Add chart data as tables
        if report_data.get('chart_data'):
            chart_data = report_data['chart_data']
            
            if report.report_type == 'employment':
                if chart_data.get('industries') and chart_data['industries'].get('items'):
                    elements.append(Paragraph('Industry Distribution', heading_style))
                    industries_data = [['Industry', 'Count']]
                    for item in chart_data['industries']['items']:
                        industries_data.append([item['label'], str(item['count'])])
                    
                    industries_table = Table(industries_data, colWidths=[4*inch, 2*inch])
                    industries_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ]))
                    elements.append(industries_table)
                    elements.append(Spacer(1, 15))
            
            elif report.report_type == 'geographic':
                if chart_data.get('countries') and chart_data['countries'].get('items'):
                    elements.append(Paragraph('Country Distribution', heading_style))
                    countries_data = [['Country', 'Count']]
                    for item in chart_data['countries']['items']:
                        countries_data.append([item['label'], str(item['count'])])
                    
                    countries_table = Table(countries_data, colWidths=[4*inch, 2*inch])
                    countries_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ]))
                    elements.append(countries_table)
                    elements.append(Spacer(1, 15))
            
            elif report.report_type == 'achievements':
                if chart_data.get('types') and chart_data['types'].get('items'):
                    elements.append(Paragraph('Achievement Types Distribution', heading_style))
                    types_data = [['Type', 'Count']]
                    for item in chart_data['types']['items']:
                        types_data.append([item['label'], str(item['count'])])
                    
                    types_table = Table(types_data, colWidths=[4*inch, 2*inch])
                    types_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ]))
                    elements.append(types_table)
                    elements.append(Spacer(1, 15))
            
            elif report.report_type == 'feedback':
                if chart_data.get('completion') and chart_data['completion'].get('items'):
                    elements.append(Paragraph('Survey Completion Rates', heading_style))
                    completion_data = [['Survey', 'Completion Rate (%)']]
                    for item in chart_data['completion']['items']:
                        completion_data.append([item['label'], f"{item['rate']:.1f}%"])
                    
                    completion_table = Table(completion_data, colWidths=[4*inch, 2*inch])
                    completion_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ]))
                    elements.append(completion_table)
                    elements.append(Spacer(1, 15))
            
            elif report.report_type == 'donations':
                if chart_data.get('campaigns') and chart_data['campaigns'].get('items'):
                    elements.append(Paragraph('Donations by Campaign', heading_style))
                    campaigns_data = [['Campaign', 'Total Amount', 'Count']]
                    for item in chart_data['campaigns']['items']:
                        campaigns_data.append([item['label'], f"₱{item['total']:,.2f}", str(item['count'])])
                    
                    campaigns_table = Table(campaigns_data, colWidths=[3*inch, 2*inch, 1*inch])
                    campaigns_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ]))
                    elements.append(campaigns_table)
                    elements.append(Spacer(1, 15))
            
            elif report.report_type == 'mentors':
                if chart_data.get('availability') and chart_data['availability'].get('items'):
                    elements.append(Paragraph('Mentors by Availability', heading_style))
                    availability_data = [['Availability Status', 'Count']]
                    for item in chart_data['availability']['items']:
                        availability_data.append([item['label'], str(item['count'])])
                    
                    availability_table = Table(availability_data, colWidths=[4*inch, 2*inch])
                    availability_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ]))
                    elements.append(availability_table)
                    elements.append(Spacer(1, 15))
            
            elif report.report_type == 'groups':
                if chart_data.get('types') and chart_data['types'].get('items'):
                    elements.append(Paragraph('Groups by Type', heading_style))
                    types_data = [['Group Type', 'Count']]
                    for item in chart_data['types']['items']:
                        types_data.append([item['label'], str(item['count'])])
                    
                    types_table = Table(types_data, colWidths=[4*inch, 2*inch])
                    types_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ]))
                    elements.append(types_table)
                    elements.append(Spacer(1, 15))
        
        # Add detailed table data
        if report_data.get('table_data'):
            elements.append(PageBreak())
            elements.append(Paragraph('Detailed Data', heading_style))
            
            if report.report_type == 'employment':
                table_headers = [['Name', 'Company', 'Job Title', 'Industry', 'Start Date']]
                for record in report_data['table_data'][:50]:  # Limit to 50 rows
                    name = f"{record.get('alumni__user__first_name', '')} {record.get('alumni__user__last_name', '')}".strip() or '-'
                    table_headers.append([
                        name[:30],  # Truncate long names
                        str(record.get('company_name', '-'))[:30],
                        str(record.get('job_title', '-'))[:30],
                        str(record.get('industry', '-'))[:25],
                        str(record.get('start_date', '-'))[:10] if record.get('start_date') else '-'
                    ])
            
            elif report.report_type == 'geographic':
                table_headers = [['Name', 'City', 'State/Province', 'Country']]
                for record in report_data['table_data'][:100]:  # Limit to 100 rows
                    name = f"{record.get('user__first_name', '')} {record.get('user__last_name', '')}".strip() or '-'
                    table_headers.append([
                        name[:30],
                        str(record.get('city', '-'))[:25],
                        str(record.get('province', '-'))[:25],
                        str(record.get('country', '-'))[:25]
                    ])
            
            elif report.report_type == 'achievements':
                table_headers = [['Name', 'Title', 'Type', 'Date', 'Verified']]
                for record in report_data['table_data'][:50]:  # Limit to 50 rows
                    name = f"{record.get('alumni__user__first_name', '')} {record.get('alumni__user__last_name', '')}".strip() or '-'
                    table_headers.append([
                        name[:25],
                        str(record.get('title', '-'))[:30],
                        str(record.get('achievement_type', '-'))[:20],
                        str(record.get('achievement_date', '-'))[:10] if record.get('achievement_date') else '-',
                        'Yes' if record.get('verified') else 'No'
                    ])
            
            elif report.report_type == 'feedback':
                table_headers = [['Survey', 'Question', 'Answer', 'Alumni', 'Date']]
                for record in report_data['table_data'][:30]:  # Limit to 30 rows
                    table_headers.append([
                        str(record.get('survey', '-'))[:25],
                        str(record.get('question', '-'))[:30],
                        str(record.get('answer', '-'))[:40],
                        str(record.get('alumni', '-'))[:25],
                        str(record.get('date', '-'))[:10] if record.get('date') else '-'
                    ])
            
            elif report.report_type == 'donations':
                table_headers = [['Donor', 'Campaign', 'Amount', 'Date', 'Payment Method']]
                for record in report_data['table_data'][:100]:  # Limit to 100 rows
                    donor_name = ''
                    if record.get('donor__first_name'):
                        donor_name = f"{record.get('donor__first_name', '')} {record.get('donor__last_name', '')}".strip()
                    elif record.get('donor_name'):
                        donor_name = record.get('donor_name', '')
                    else:
                        donor_name = 'Anonymous'
                    
                    table_headers.append([
                        donor_name[:25],
                        str(record.get('campaign__name', '-'))[:25],
                        f"₱{float(record.get('amount', 0)):,.2f}",
                        str(record.get('donation_date', '-'))[:10] if record.get('donation_date') else '-',
                        str(record.get('payment_method', '-'))[:20]
                    ])
            
            elif report.report_type == 'mentors':
                table_headers = [['Name', 'Availability', 'Verified', 'Current Mentees', 'Max Mentees', 'Accepting']]
                for record in report_data['table_data'][:100]:  # Limit to 100 rows
                    name = f"{record.get('user__first_name', '')} {record.get('user__last_name', '')}".strip() or '-'
                    table_headers.append([
                        name[:25],
                        str(record.get('availability_status', '-'))[:20],
                        'Yes' if record.get('is_verified') else 'No',
                        str(record.get('current_mentees', 0)),
                        str(record.get('max_mentees', '-')),
                        'Yes' if record.get('accepting_mentees') else 'No'
                    ])
            
            elif report.report_type == 'groups':
                table_headers = [['Group Name', 'Type', 'Visibility', 'Created By', 'Members', 'Created']]
                for record in report_data['table_data'][:100]:  # Limit to 100 rows
                    name = f"{record.get('created_by__first_name', '')} {record.get('created_by__last_name', '')}".strip() or '-'
                    table_headers.append([
                        str(record.get('name', '-'))[:30],
                        str(record.get('group_type', '-'))[:15],
                        str(record.get('visibility', '-'))[:15],
                        name[:25],
                        str(record.get('member_count', 0)),
                        str(record.get('created_at', '-'))[:10] if record.get('created_at') else '-'
                    ])
            
            else:
                # Custom report - use all keys
                if report_data['table_data']:
                    keys = list(report_data['table_data'][0].keys())
                    table_headers = [[k.replace('_', ' ').title() for k in keys]]
                    for record in report_data['table_data'][:50]:
                        table_headers.append([str(record.get(k, '-'))[:30] for k in keys])
            
            if len(table_headers) > 1:  # Only create table if there's data
                # Use landscape for wide tables
                use_landscape = len(table_headers[0]) > 4
                
                # Calculate column widths
                num_cols = len(table_headers[0])
                if use_landscape:
                    available_width = landscape(A4)[0] - 60  # Account for margins
                else:
                    available_width = A4[0] - 60
                col_width = available_width / num_cols
                col_widths = [col_width] * num_cols
                
                data_table = Table(table_headers, colWidths=col_widths)
                data_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#417690')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elements.append(data_table)
                
                if len(report_data['table_data']) > 50:
                    elements.append(Spacer(1, 10))
                    elements.append(Paragraph(
                        f"Note: Showing first 50 records of {len(report_data['table_data'])} total.",
                        styles['Normal']
                    ))
    
        # Build PDF with custom canvas
        doc.build(elements, canvasmaker=HeaderCanvas)
        
        # Calculate export time
        elapsed_time = time.time() - start_time
        
        # Log PDF export success
        logger.info(
            f"PDF export completed successfully: Report ID={report.id}, Type={report.report_type}, Time={elapsed_time:.2f}s",
            extra={
                'report_id': report.id,
                'report_title': report.title,
                'report_type': report.report_type,
                'user_id': request.user.id,
                'filename': filename,
                'export_time': elapsed_time,
                'action': 'pdf_export_complete'
            }
        )
        
        return response
        
    except Exception as e:
        elapsed_time = time.time() - start_time
        logger.error(
            f"Error exporting PDF: {str(e)}",
            extra={
                'report_id': report.id,
                'report_title': report.title,
                'report_type': report.report_type,
                'user_id': request.user.id,
                'error_type': type(e).__name__,
                'export_time': elapsed_time,
                'exc_info': True
            }
        )
        raise


# Employment and Achievement record management for alumni
class EmploymentRecordCreateView(LoginRequiredMixin, CreateView):
    model = EmploymentRecord
    form_class = EmploymentRecordForm
    template_name = 'surveys/employment_form.html'
    success_url = reverse_lazy('surveys:employment_list')
    
    def form_valid(self, form):
        alumni = get_object_or_404(Alumni, user=self.request.user)
        form.instance.alumni = alumni
        
        try:
            response = super().form_valid(form)
            
            # Log employment record creation
            logger.info(
                f"Employment record created: Record ID={self.object.id}, Company={form.instance.company_name}",
                extra={
                    'employment_record_id': self.object.id,
                    'company_name': form.instance.company_name,
                    'job_title': form.instance.job_title,
                    'industry': form.instance.industry,
                    'alumni_id': alumni.id,
                    'user_id': self.request.user.id
                }
            )
            
            return response
            
        except Exception as e:
            logger.error(
                f"Error creating employment record: {str(e)}",
                extra={
                    'alumni_id': alumni.id,
                    'user_id': self.request.user.id,
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            raise

class EmploymentRecordUpdateView(LoginRequiredMixin, UpdateView):
    model = EmploymentRecord
    form_class = EmploymentRecordForm
    template_name = 'surveys/employment_form.html'
    success_url = reverse_lazy('surveys:employment_list')
    
    def get_queryset(self):
        alumni = get_object_or_404(Alumni, user=self.request.user)
        return EmploymentRecord.objects.filter(alumni=alumni)

class EmploymentRecordListView(LoginRequiredMixin, ListView):
    model = EmploymentRecord
    template_name = 'surveys/employment_list.html'
    context_object_name = 'employment_records'
    
    def get_queryset(self):
        alumni = get_object_or_404(Alumni, user=self.request.user)
        return EmploymentRecord.objects.filter(alumni=alumni).order_by('-start_date')

class AchievementCreateView(LoginRequiredMixin, CreateView):
    model = Achievement
    form_class = AchievementForm
    template_name = 'surveys/achievement_form.html'
    success_url = reverse_lazy('surveys:achievement_list')
    
    def form_valid(self, form):
        alumni = get_object_or_404(Alumni, user=self.request.user)
        form.instance.alumni = alumni
        
        try:
            response = super().form_valid(form)
            
            # Log achievement creation
            logger.info(
                f"Achievement created: Achievement ID={self.object.id}, Type={form.instance.achievement_type}",
                extra={
                    'achievement_id': self.object.id,
                    'achievement_title': form.instance.title,
                    'achievement_type': form.instance.achievement_type,
                    'achievement_date': str(form.instance.achievement_date),
                    'alumni_id': alumni.id,
                    'user_id': self.request.user.id
                }
            )
            
            return response
            
        except Exception as e:
            logger.error(
                f"Error creating achievement: {str(e)}",
                extra={
                    'alumni_id': alumni.id,
                    'user_id': self.request.user.id,
                    'error_type': type(e).__name__,
                    'exc_info': True
                }
            )
            raise

class AchievementUpdateView(LoginRequiredMixin, UpdateView):
    model = Achievement
    form_class = AchievementForm
    template_name = 'surveys/achievement_form.html'
    success_url = reverse_lazy('surveys:achievement_list')
    
    def get_queryset(self):
        alumni = get_object_or_404(Alumni, user=self.request.user)
        return Achievement.objects.filter(alumni=alumni)

class AchievementListView(LoginRequiredMixin, ListView):
    model = Achievement
    template_name = 'surveys/achievement_list.html'
    context_object_name = 'achievements'
    
    def get_queryset(self):
        alumni = get_object_or_404(Alumni, user=self.request.user)
        return Achievement.objects.filter(alumni=alumni).order_by('-achievement_date')
