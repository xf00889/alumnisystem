"""
Dedicated views for the NORSU Graduate Tracer Study.

The questionnaire data is still stored in the generic ``Survey`` model (so it
can be reported on like any other survey), but the user-facing form lives at
its own URL with its own template, so it doesn't feel like a generic survey.

URLs:
* ``/tracer-study/``         - Alumni (login required, one response per alumni)
* ``/tracer-study/employer/`` - Employers (public; auto-creates an Employer
                                record on first submission and reuses it
                                thereafter)
"""
from datetime import date
from io import BytesIO
from pathlib import Path
import asyncio
import base64
import csv
import json
import logging
import os
import re
import shutil
import subprocess
import tempfile
import time
import urllib.request
import zipfile

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.crypto import constant_time_compare, salted_hmac
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_http_methods
from django.views.generic import View

from alumni_directory.models import Alumni
from .models import (
    Employer,
    EmployerResponse,
    EmployerResponseAnswer,
    QuestionOption,
    Report,
    ResponseAnswer,
    Survey,
    SurveyQuestion,
    SurveyResponse,
)


ALUMNI_TITLE = "NORSU Graduate Tracer Study (ALUMNI QUESTIONNAIRE)"
EMPLOYER_TITLE = "NORSU Graduate Tracer Study (EMPLOYER QUESTIONNAIRE)"
PART_IV_EXTENT_KEYS = {
    "p4_vision",
    "p4_mission",
    "p4_goals",
    "p4_core_values",
    "p4_program_objectives",
}
PART_IV_EXTENT_LABELS = {
    "p4_vision": "Negros Oriental State University VISION",
    "p4_mission": "Negros Oriental State University MISSION",
    "p4_goals": "Negros Oriental State University GOALS",
    "p4_core_values": "NORSU Corporate Values/Core Values/Graduate Attributes",
    "p4_program_objectives": "Course/Program Objectives",
}
logger = logging.getLogger(__name__)
_NORSU_HEADER_DATA_URI = None
_FILLED_RESPONSE_SALT = "tracer-study-filled-response"


def _norsu_header_data_uri():
    global _NORSU_HEADER_DATA_URI
    if _NORSU_HEADER_DATA_URI is not None:
        return _NORSU_HEADER_DATA_URI

    header_path = Path(settings.BASE_DIR) / "static" / "images" / "norsu-header.png"
    try:
        _NORSU_HEADER_DATA_URI = "data:image/png;base64," + base64.b64encode(header_path.read_bytes()).decode("ascii")
    except OSError:
        logger.warning("NORSU tracer letterhead image not found: %s", header_path)
        _NORSU_HEADER_DATA_URI = ""
    return _NORSU_HEADER_DATA_URI


def _can_view_tracer_reports(user):
    return (
        user.is_staff
        or user.is_superuser
        or (hasattr(user, "profile") and user.profile.is_alumni_coordinator)
    )


def _hashed_response_id(response_id):
    digest = salted_hmac(
        _FILLED_RESPONSE_SALT,
        str(response_id),
        algorithm="sha256",
    ).hexdigest()
    return f"r-{digest}"


def _response_id_from_token(response_token):
    response_ids = SurveyResponse.objects.filter(
        survey__title=ALUMNI_TITLE,
    ).values_list("id", flat=True)
    for response_id in response_ids:
        if constant_time_compare(_hashed_response_id(response_id), response_token):
            return response_id
    raise Http404


def _question_key_from_text(question_text):
    text_key = (question_text or "").strip().lower()
    if "contact number" in text_key or "mobile" in text_key:
        return "p1_contact"
    if "date of birth" in text_key or text_key == "birthdate":
        return "p1_dob"
    if "facebook" in text_key:
        return "p1_fb"
    if "twitter" in text_key or "x / twitter" in text_key:
        return "p1_twitter"
    if "presently employed" in text_key:
        return "p3_employed"
    if "employment status" in text_key:
        return "p3_status"
    if "vision" in text_key:
        return "p4_vision"
    if "mission" in text_key:
        return "p4_mission"
    if "goals" in text_key:
        return "p4_goals"
    if "core values" in text_key or "graduate attributes" in text_key:
        return "p4_core_values"
    if "program objectives" in text_key or "course/program objectives" in text_key:
        return "p4_program_objectives"
    return ""


def _alumni_response_rows(survey):
    responses = {
        response.alumni_id: response
        for response in SurveyResponse.objects.filter(survey=survey).select_related("alumni__user")
    }
    rows = []
    for alumni in Alumni.objects.select_related("user").order_by("user__last_name", "user__first_name"):
        response = responses.get(alumni.id)
        rows.append({
            "id": alumni.id,
            "name": alumni.full_name or alumni.user.username,
            "email": alumni.user.email,
            "course": alumni.course,
            "graduation_year": alumni.graduation_year,
            "status": "Responded" if response else "Not Responded",
            "submitted_at": response.submitted_at if response else None,
            "response_id": response.id if response else None,
            "response_token": _hashed_response_id(response.id) if response else None,
        })
    return rows


def _answer_key(question):
    try:
        meta = json.loads(question.help_text) if question.help_text else {}
    except (ValueError, TypeError):
        meta = {}
    if meta.get("key"):
        return meta["key"]
    return _question_key_from_text(question.question_text) or str(question.id)


def _filled_alumni_answers(response):
    by_key = {}
    for answer in response.answers.select_related("question", "selected_option"):
        key = _answer_key(answer.question)
        item = by_key.setdefault(key, {"text": "", "rating": "", "selected": set(), "other": ""})
        if answer.text_answer:
            item["text"] = answer.text_answer
        if answer.rating_value:
            item["rating"] = str(answer.rating_value)
        if answer.selected_option:
            item["selected"].add(answer.selected_option.option_text.lower())
        if answer.custom_text:
            item["other"] = answer.custom_text

    def text(key):
        return by_key.get(key, {}).get("text", "")

    def other(key):
        return by_key.get(key, {}).get("other", "")

    def checked(key, *needles):
        selected = by_key.get(key, {}).get("selected", set())
        return any(any(needle.lower() in value for value in selected) for needle in needles)

    def checked_or_source(key, source_value, *needles):
        if checked(key, *needles):
            return True
        selected = by_key.get(key, {}).get("selected", set())
        if selected:
            return False
        source_value = (source_value or "").lower()
        return any(needle.lower() in source_value for needle in needles)

    def rating(key, value):
        return by_key.get(key, {}).get("rating") == str(value)

    def birth_date():
        submitted_dob = text("p1_dob")
        if submitted_dob:
            try:
                return date.fromisoformat(submitted_dob)
            except ValueError:
                pass
        return response.alumni.date_of_birth or getattr(
            getattr(response.alumni.user, "profile", None),
            "birth_date",
            None,
        )

    alumni = response.alumni
    source = _alumni_prefill_source(alumni)
    dob = birth_date()

    return {
        "p1_name": text("p1_name") or source.get("p1_name") or "",
        "p1_address": text("p1_address") or source.get("p1_address") or "",
        "p1_email": text("p1_email") or source.get("p1_email") or "",
        "p1_contact": text("p1_contact") or source.get("p1_contact") or "",
        "p1_age": _compute_age(dob) or "",
        "p1_fb": text("p1_fb") or source.get("p1_fb") or "",
        "p1_twitter": text("p1_twitter") or source.get("p1_twitter") or "",
        "gender_male": checked("p1_gender", "male"),
        "gender_female": checked("p1_gender", "female"),
        "gender_other": checked("p1_gender", "other"),
        "gender_other_text": other("p1_gender"),
        "marital_single": checked("p1_marital", "single"),
        "marital_married": checked("p1_marital", "married"),
        "marital_other": checked("p1_marital", "other"),
        "marital_other_text": other("p1_marital"),
        "p2_course": text("p2_course") or source.get("p2_course") or "",
        "p2_year": text("p2_year") or source.get("p2_year") or "",
        "campus_main1": checked("p2_campus", "main campus i"),
        "campus_main2": checked("p2_campus", "main campus ii"),
        "campus_bayawan": checked("p2_campus", "bayawan"),
        "campus_siaton": checked("p2_campus", "siaton"),
        "campus_guihulngan": checked("p2_campus", "guihulngan"),
        "campus_bais": checked("p2_campus", "bais"),
        "campus_mabinay": checked("p2_campus", "mabinay"),
        "campus_pamplona": checked("p2_campus", "pamplona"),
        "honor_academic": checked("p2_honors", "academic"),
        "honor_leadership": checked("p2_honors", "leadership"),
        "honor_special": checked("p2_honors", "special"),
        "honor_other": checked("p2_honors", "other"),
        "honor_other_text": other("p2_honors"),
        "p2_exam_name": text("p2_exam_name"),
        "p2_exam_year": text("p2_exam_year"),
        "employed_yes": checked_or_source("p3_employed", source.get("p3_employed"), "yes"),
        "employed_no": checked_or_source("p3_employed", source.get("p3_employed"), "no"),
        "employed_never": checked_or_source("p3_employed", source.get("p3_employed"), "never"),
        "employed_other": checked_or_source("p3_employed", source.get("p3_employed"), "other"),
        "employed_other_text": other("p3_employed"),
        "unemployed_study": checked("p3_reasons_unemployed", "study", "further"),
        "unemployed_experience": checked("p3_reasons_unemployed", "experience"),
        "unemployed_family": checked("p3_reasons_unemployed", "family"),
        "unemployed_opportunity": checked("p3_reasons_unemployed", "opportunity"),
        "unemployed_health": checked("p3_reasons_unemployed", "health"),
        "unemployed_no_look": checked("p3_reasons_unemployed", "did not look"),
        "unemployed_other": checked("p3_reasons_unemployed", "other"),
        "unemployed_other_text": other("p3_reasons_unemployed"),
        "status_regular": checked("p3_status", "regular", "permanent"),
        "status_contractual": checked("p3_status", "contractual"),
        "status_temporary": checked("p3_status", "temporary"),
        "status_self": checked("p3_status", "self"),
        "status_casual": checked("p3_status", "casual"),
        "status_other": checked("p3_status", "other"),
        "status_other_text": other("p3_status"),
        "p3_position": text("p3_position"),
        "p3_company": text("p3_company"),
        "p3_company_address": text("p3_company_address"),
        "org_government": checked("p3_org_type", "government"),
        "org_private": checked("p3_org_type", "private"),
        "place_local": checked("p3_place", "local"),
        "place_abroad": checked("p3_place", "abroad"),
        "first_job_yes": checked("p3_first_job", "yes"),
        "first_job_no": checked("p3_first_job", "no"),
        "stay_salary": checked("p3_reasons_stay", "salar"),
        "stay_proximity": checked("p3_reasons_stay", "proximity"),
        "stay_challenge": checked("p3_reasons_stay", "challenge"),
        "stay_peer": checked("p3_reasons_stay", "peer"),
        "stay_skill": checked("p3_reasons_stay", "skill"),
        "stay_family": checked("p3_reasons_stay", "family"),
        "stay_course": checked("p3_reasons_stay", "course"),
        "stay_other": checked("p3_reasons_stay", "other"),
        "stay_other_text": other("p3_reasons_stay"),
        "related_yes": checked("p3_related_course", "yes"),
        "related_no": checked("p3_related_course", "no"),
        "accept_salary": checked("p3_reasons_accept", "salar"),
        "accept_skill": checked("p3_reasons_accept", "skill"),
        "accept_challenge": checked("p3_reasons_accept", "challenge"),
        "accept_proximity": checked("p3_reasons_accept", "proximity"),
        "accept_other": checked("p3_reasons_accept", "other"),
        "accept_other_text": other("p3_reasons_accept"),
        "change_salary": checked("p3_reasons_change", "salar"),
        "change_skill": checked("p3_reasons_change", "skill"),
        "change_challenge": checked("p3_reasons_change", "challenge"),
        "change_proximity": checked("p3_reasons_change", "proximity"),
        "change_other": checked("p3_reasons_change", "other"),
        "change_other_text": other("p3_reasons_change"),
        "duration_lt_1m": checked("p3_first_job_duration", "less than a month"),
        "duration_1_6m": checked("p3_first_job_duration", "1-6", "1 to 6"),
        "duration_7_11m": checked("p3_first_job_duration", "7-11", "7 to 11"),
        "duration_1_2y": checked("p3_first_job_duration", "1-2", "1 year"),
        "duration_2_3y": checked("p3_first_job_duration", "2-3", "2 years"),
        "duration_3_4y": checked("p3_first_job_duration", "3-4", "3 years"),
        "duration_other": checked("p3_first_job_duration", "other"),
        "duration_other_text": other("p3_first_job_duration"),
        "find_ad": checked("p3_find_first_job", "advert"),
        "find_school": checked("p3_find_first_job", "school"),
        "find_walkin": checked("p3_find_first_job", "walk"),
        "find_family": checked("p3_find_first_job", "family"),
        "find_referral": checked("p3_find_first_job", "recommended"),
        "find_jobfair": checked("p3_find_first_job", "job fair", "peso"),
        "find_friends": checked("p3_find_first_job", "friends"),
        "find_other": checked("p3_find_first_job", "other"),
        "find_other_text": other("p3_find_first_job"),
        "land_lt_1m": checked("p3_time_to_first_job", "less than a month"),
        "land_1_6m": checked("p3_time_to_first_job", "1-6", "1 to 6"),
        "land_7_11m": checked("p3_time_to_first_job", "7-11", "7 to 11"),
        "land_1_2y": checked("p3_time_to_first_job", "1-2", "1 year"),
        "land_2_3y": checked("p3_time_to_first_job", "2-3", "2 years"),
        "land_3_4y": checked("p3_time_to_first_job", "3-4", "3 years"),
        "land_other": checked("p3_time_to_first_job", "other"),
        "land_other_text": other("p3_time_to_first_job"),
        "curriculum_yes": checked("p3_curriculum_relevant", "yes"),
        "curriculum_no": checked("p3_curriculum_relevant", "no"),
        "competency_communication": checked("p3_useful_competencies", "communication"),
        "competency_problem": checked("p3_useful_competencies", "problem"),
        "competency_human": checked("p3_useful_competencies", "human"),
        "competency_critical": checked("p3_useful_competencies", "critical"),
        "competency_entrepreneurial": checked("p3_useful_competencies", "entrepreneurial"),
        "competency_other": checked("p3_useful_competencies", "other"),
        "competency_other_text": other("p3_useful_competencies"),
        "p3_apply_learning": text("p3_apply_learning"),
        "p3_suggestions": text("p3_suggestions"),
        **{f"p4_vision_{i}": rating("p4_vision", i) for i in range(1, 6)},
        **{f"p4_mission_{i}": rating("p4_mission", i) for i in range(1, 6)},
        **{f"p4_goals_{i}": rating("p4_goals", i) for i in range(1, 6)},
        **{f"p4_core_values_{i}": rating("p4_core_values", i) for i in range(1, 6)},
        **{f"p4_program_objectives_{i}": rating("p4_program_objectives", i) for i in range(1, 6)},
    }


def _safe_export_name(value, fallback):
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(value or "").strip())
    value = re.sub(r"\s+", " ", value).strip(" ._")
    return value or fallback


def _tracer_campus_folder(alumni):
    return {
        "MAIN": "NORSU MAIN",
        "BAIS1": "NORSU BAIS",
        "BAIS2": "NORSU BAIS",
        "BSC": "NORSU-BSC",
        "PAM": "NORSU PAMPLONA",
        "SIATON": "NORSU SIATON",
        "GUI": "NORSU GUIHULNGAN",
        "MAB": "NORSU MABINAY",
    }.get(alumni.campus, alumni.get_campus_display() or "UNKNOWN CAMPUS")


def _tracer_response_pdf_filename(response):
    user = response.alumni.user
    last_name = _safe_export_name(user.last_name or user.username, "LastName")
    first_name = _safe_export_name(user.first_name or user.username, "FirstName")
    college = _safe_export_name(response.alumni.college, "College")
    return f"TracerStudy_{last_name}_{first_name}_{college}.pdf"


def _tracer_browser_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.selenium_manager import SeleniumManager

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    chrome_binary = os.getenv("CHROME_BIN") or os.getenv("GOOGLE_CHROME_BIN")
    if chrome_binary:
        options.binary_location = chrome_binary
    remote_url = os.getenv("SELENIUM_REMOTE_URL")
    if remote_url:
        return webdriver.Remote(command_executor=remote_url, options=options)
    service = None
    if not chrome_binary:
        paths = SeleniumManager().binary_paths([
            "--browser",
            "chrome",
            "--browser-version",
            os.getenv("SELENIUM_BROWSER_VERSION", "stable"),
        ])
        options.binary_location = paths["browser_path"]
        service = Service(executable_path=paths["driver_path"])
    return webdriver.Chrome(service=service, options=options)


def _tracer_response_filled_form_html(response):
    return render_to_string(
        "tracer_study/filled_alumni_questionnaire.html",
        {
            "response": response,
            "survey": response.survey,
            "filled": _filled_alumni_answers(response),
            "header_data_uri": _norsu_header_data_uri(),
        },
    )


def _tracer_response_template_pdf_bytes(response, driver):
    html = _tracer_response_filled_form_html(response)
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".html", encoding="utf-8", delete=False) as tmp:
            tmp.write(html)
            tmp_path = tmp.name
        driver.get(Path(tmp_path).resolve().as_uri())
        driver.execute_cdp_cmd("Emulation.setEmulatedMedia", {"media": "print"})
        pdf = driver.execute_cdp_cmd(
            "Page.printToPDF",
            {
                "printBackground": True,
                "preferCSSPageSize": True,
                "displayHeaderFooter": False,
            },
        )
        return base64.b64decode(pdf["data"])
    finally:
        if tmp_path:
            try:
                Path(tmp_path).unlink()
            except OSError:
                pass


def _tracer_chrome_binary():
    for env_name in ("CHROME_BIN", "GOOGLE_CHROME_BIN", "CHROMIUM_BIN"):
        binary = os.getenv(env_name)
        if binary and Path(binary).exists():
            return binary

    for name in (
        "google-chrome",
        "google-chrome-stable",
        "chromium",
        "chromium-browser",
        "chrome",
        "msedge",
        "microsoft-edge",
    ):
        binary = shutil.which(name)
        if binary:
            return binary

    for path in (
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
        "/usr/bin/chromium",
        "/usr/bin/chromium-browser",
        "/snap/bin/chromium",
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    ):
        if Path(path).exists():
            return path

    return ""


def _tracer_chrome_env():
    env = os.environ.copy()
    default_paths = "/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin"
    env["PATH"] = f"{env.get('PATH', '')}:{default_paths}" if env.get("PATH") else default_paths
    return env


def _tracer_chrome_work_root():
    configured = os.getenv("TRACER_PDF_WORK_DIR")
    candidates = [Path(configured)] if configured else []
    candidates.extend(
        [
            Path(os.getenv("HOME", str(Path.home()))) / "norsu-tracer-pdf",
            Path(settings.BASE_DIR) / "tmp" / "tracer-pdf",
        ]
    )

    for root in candidates:
        try:
            root.mkdir(parents=True, exist_ok=True)
            probe = root / ".write-test"
            probe.write_text("", encoding="utf-8")
            probe.unlink()
            return root
        except OSError:
            continue

    logger.warning("Tracer PDF work directories are not writable; falling back to system temp")
    return None


def _tracer_chrome_pdf_env(work_path, runtime_dir):
    env = _tracer_chrome_env()
    env.setdefault("HOME", str(work_path))
    env["XDG_RUNTIME_DIR"] = str(runtime_dir)
    return env


def _tracer_chrome_base_args(chrome_binary, headless_arg, user_data_dir):
    return [
        chrome_binary,
        headless_arg,
        "--disable-gpu",
        "--no-sandbox",
        "--disable-setuid-sandbox",
        "--disable-dev-shm-usage",
        "--no-first-run",
        "--no-default-browser-check",
        f"--user-data-dir={user_data_dir}",
    ]


async def _tracer_chrome_cdp_print_pdf(websocket_url):
    import websockets

    next_id = 0

    async with websockets.connect(websocket_url) as websocket:
        async def command(method, params=None):
            nonlocal next_id
            next_id += 1
            request_id = next_id
            await websocket.send(json.dumps({"id": request_id, "method": method, "params": params or {}}))
            while True:
                message = json.loads(await websocket.recv())
                if message.get("id") != request_id:
                    continue
                if "error" in message:
                    raise RuntimeError(message["error"])
                return message.get("result", {})

        await command("Page.enable")
        for _ in range(50):
            result = await command("Runtime.evaluate", {"expression": "document.readyState", "returnByValue": True})
            if result.get("result", {}).get("value") == "complete":
                break
            await asyncio.sleep(0.1)
        await command("Emulation.setEmulatedMedia", {"media": "print"})
        pdf = await command(
            "Page.printToPDF",
            {
                "printBackground": True,
                "preferCSSPageSize": True,
                "displayHeaderFooter": False,
            },
        )
        return base64.b64decode(pdf["data"])


def _tracer_response_chrome_cdp_pdf_bytes(response):
    chrome_binary = _tracer_chrome_binary()
    if not chrome_binary:
        raise RuntimeError("Chrome/Chromium binary not found")

    with tempfile.TemporaryDirectory(dir=_tracer_chrome_work_root()) as work_dir:
        work_path = Path(work_dir)
        html_path = work_path / "response.html"
        user_data_dir = work_path / "chrome-profile"
        runtime_dir = work_path / "xdg-runtime"
        runtime_dir.mkdir(mode=0o700)
        html_path.write_text(_tracer_response_filled_form_html(response), encoding="utf-8")

        file_url = html_path.resolve().as_uri()
        last_error = None
        for headless_arg in ("--headless=new", "--headless"):
            command = _tracer_chrome_base_args(chrome_binary, headless_arg, user_data_dir) + [
                "--remote-debugging-port=0",
                file_url,
            ]
            process = subprocess.Popen(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                env=_tracer_chrome_pdf_env(work_path, runtime_dir),
            )
            try:
                port_file = user_data_dir / "DevToolsActivePort"
                for _ in range(100):
                    if process.poll() is not None:
                        break
                    if port_file.exists():
                        break
                    time.sleep(0.1)

                if not port_file.exists():
                    if process.poll() is not None and process.stderr:
                        stderr = process.stderr.read().decode("utf-8", errors="replace")
                        last_error = stderr or "Chrome DevTools port was not created"
                    else:
                        last_error = "Chrome DevTools port was not created"
                    continue

                port = port_file.read_text(encoding="utf-8").splitlines()[0].strip()
                with urllib.request.urlopen(f"http://127.0.0.1:{port}/json/list", timeout=10) as response_file:
                    targets = json.loads(response_file.read().decode("utf-8"))
                page = next((target for target in targets if target.get("type") == "page"), None)
                if not page or not page.get("webSocketDebuggerUrl"):
                    last_error = "Chrome DevTools page target was not available"
                    continue

                pdf = asyncio.run(_tracer_chrome_cdp_print_pdf(page["webSocketDebuggerUrl"]))
                if b"%PDF" in pdf[:1024]:
                    return pdf
                last_error = "Chrome DevTools PDF output was not a PDF"
            finally:
                if process.poll() is None:
                    process.terminate()
                    try:
                        process.wait(timeout=5)
                    except subprocess.TimeoutExpired:
                        process.kill()

        raise RuntimeError(last_error or "Chrome DevTools PDF export failed")


def _tracer_response_chrome_cli_pdf_bytes(response):
    chrome_binary = _tracer_chrome_binary()
    if not chrome_binary:
        raise RuntimeError("Chrome/Chromium binary not found")

    try:
        return _tracer_response_chrome_cdp_pdf_bytes(response)
    except Exception as exc:
        logger.warning("Tracer filled-form PDF Chrome DevTools renderer unavailable; using Chrome CLI renderer: %s", exc)

    with tempfile.TemporaryDirectory(dir=_tracer_chrome_work_root()) as work_dir:
        work_path = Path(work_dir)
        html_path = work_path / "response.html"
        pdf_path = work_path / "response.pdf"
        user_data_dir = work_path / "chrome-profile"
        runtime_dir = work_path / "xdg-runtime"
        runtime_dir.mkdir(mode=0o700)

        with html_path.open("w", encoding="utf-8") as html_file:
            html_file.write(_tracer_response_filled_form_html(response))

        file_url = Path(html_path).resolve().as_uri()
        last_error = None
        for headless_arg in ("--headless=new", "--headless"):
            command = _tracer_chrome_base_args(chrome_binary, headless_arg, user_data_dir) + [
                "--no-pdf-header-footer",
                "--print-to-pdf-no-header",
                f"--print-to-pdf={pdf_path}",
                file_url,
            ]
            result = subprocess.run(command, capture_output=True, timeout=90, env=_tracer_chrome_pdf_env(work_path, runtime_dir))
            output = ((result.stderr or b"") + (result.stdout or b"")).decode("utf-8", errors="replace")
            written_path = re.search(r"\d+\s+bytes written to file\s+(.+?\.pdf)", output)
            for candidate in {pdf_path, written_path.group(1) if written_path else ""}:
                if candidate and Path(candidate).exists() and Path(candidate).stat().st_size:
                    pdf = Path(candidate).read_bytes()
                    if b"%PDF" in pdf[:1024] or written_path:
                        return pdf
            last_error = output
        raise RuntimeError(last_error or "Chrome/Chromium PDF export failed")


def _tracer_study_forms_zip_response(survey):
    responses = (
        SurveyResponse.objects.filter(survey=survey)
        .select_related("alumni__user")
        .prefetch_related("answers__question", "answers__selected_option")
        .order_by("alumni__campus", "alumni__college", "alumni__course", "alumni__user__last_name", "alumni__user__first_name")
    )
    buffer = BytesIO()
    used_paths = set()
    driver = None
    try:
        driver = _tracer_browser_driver()
    except Exception as exc:
        logger.warning("Tracer filled-form PDF Selenium renderer unavailable; using Chrome CLI renderer: %s", exc)
        if not _tracer_chrome_binary():
            return HttpResponse(
                "Exact tracer filled-form ZIP export requires Chrome/Chromium or SELENIUM_REMOTE_URL on the server. "
                "Install Chromium, set CHROME_BIN, or set SELENIUM_REMOTE_URL.",
                status=503,
                content_type="text/plain",
            )

    try:
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for response in responses:
                alumni = response.alumni
                campus = _safe_export_name(_tracer_campus_folder(alumni), "UNKNOWN CAMPUS")
                college = _safe_export_name(alumni.college, "UNKNOWN COLLEGE")
                program = _safe_export_name(alumni.course, "UNKNOWN PROGRAM")
                filename = _tracer_response_pdf_filename(response)
                path = f"{campus}/{college}/{program}/{filename}"
                base, ext = path[:-4], ".pdf"
                counter = 2
                while path.lower() in used_paths:
                    path = f"{base}_{counter}{ext}"
                    counter += 1
                used_paths.add(path.lower())
                pdf = (
                    _tracer_response_template_pdf_bytes(response, driver)
                    if driver
                    else _tracer_response_chrome_cli_pdf_bytes(response)
                )
                zip_file.writestr(path, pdf)

            if not used_paths:
                zip_file.writestr("README.txt", "No tracer study responses found.")
    finally:
        if driver:
            driver.quit()

    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="tracer-study-filled-forms.zip"'
    return response


def _get_active_survey(title):
    return get_object_or_404(Survey, title=title, status="active")


# ---------------------------------------------------------------------------
# Alumni profile -> question prefill
# ---------------------------------------------------------------------------
#
# Map ``Alumni`` (and its user/profile) fields to the tracer-study question
# ``key``s defined in ``seed_tracer_study.ALUMNI_QUESTIONS``. The prefill is
# only used when the field has a value in the system; missing/blank values
# leave the form field empty so the alumni can fill it in themselves.
#
# Returned dict format: ``{question_id: prefill_value}`` where prefill_value
# is a string for text-like types and an ``option.id`` (int) for
# ``multiple_choice``.

# Alumni.gender choices -> questionnaire value_keys for p1_gender
_GENDER_VALUE_KEY = {"M": "male", "F": "female", "O": "other"}

# Alumni.campus choices -> questionnaire value_keys for p2_campus.
# Alumni combines Main I + II into a single MAIN choice; default to Main I.
_CAMPUS_VALUE_KEY = {
    "MAIN": "main1",
    "BAIS1": "bais", "BAIS2": "bais",
    "BSC": "bayawan", "SIATON": "siaton", "GUI": "guihulngan",
    "PAM": "pamplona", "MAB": "mabinay",
}

# Profile.education.school choices -> questionnaire value_keys for p2_campus.
# Used as a secondary source when Alumni.campus is blank.
_EDUCATION_CAMPUS_VALUE_KEY = {
    "NORSU-MAIN": "main1",
    "NORSU-G": "guihulngan",
    "NORSU-BC": "bais",     # Bais City — Bais Campuses
    "NORSU-BSC": "bayawan",
    "NORSU-MB": "mabinay",
    "NORSU-SC": "siaton",
    "NORSU-PC": "pamplona",
    # "OTHER" / blank / null: no match
}

# Alumni.employment_status -> p3_employed value_key
_EMPLOYED_VALUE_KEY = {
    "EMPLOYED_FULL": "yes", "EMPLOYED_PART": "yes",
    "SELF_EMPLOYED": "yes", "INTERN": "yes",
    "UNEMPLOYED": "no", "STUDENT": "no", "RETIRED": "no",
}


def _compute_age(birth_date):
    if not birth_date:
        return None
    today = date.today()
    return today.year - birth_date.year - (
        (today.month, today.day) < (birth_date.month, birth_date.day)
    )


def _campus_from_education(profile):
    """Pick a campus value_key from the most recent / primary Education row.

    Returns None if profile is missing or has no education rows with a
    mappable school code.
    """
    if profile is None:
        return None
    edu_qs = profile.education.exclude(school__isnull=True).exclude(school="")
    # Prefer the primary record; otherwise the most recent graduation year.
    primary = edu_qs.filter(is_primary=True).first()
    if primary is None:
        primary = edu_qs.order_by("-graduation_year", "-id").first()
    if primary is None:
        return None
    return _EDUCATION_CAMPUS_VALUE_KEY.get(primary.school)


def _alumni_prefill_source(alumni):
    """Return {question_key: source_value} for any prefillable field.

    The source value is the raw Alumni/Profile data; the view layer is
    responsible for turning it into a concrete question value (option id
    vs text) when it builds the per-question prefill dict.
    """
    user = alumni.user
    profile = getattr(user, "profile", None)

    dob = alumni.date_of_birth
    if not dob and profile is not None:
        dob = getattr(profile, "birth_date", None)

    address_parts = filter(None, [
        alumni.address, alumni.city, alumni.province,
        str(alumni.country.name) if alumni.country else None,
    ])
    address = ", ".join(address_parts) if address_parts else None

    course = alumni.course or None
    if course and alumni.major:
        course = f"{course} - {alumni.major}"

    # Campus: prefer Alumni.campus, fall back to Profile.education.school.
    campus = _CAMPUS_VALUE_KEY.get(alumni.campus) or _campus_from_education(profile)

    return {
        "p1_name": (user.get_full_name() or user.username) or None,
        "p1_address": address,
        "p1_email": user.email or None,
        "p1_contact": str(alumni.phone_number) if alumni.phone_number else None,
        "p1_dob": dob.isoformat() if dob else None,
        "p1_gender": _GENDER_VALUE_KEY.get(alumni.gender),
        "p1_fb": (getattr(profile, "facebook_profile", "") or "") or None,
        "p1_twitter": (getattr(profile, "twitter_profile", "") or "") or None,
        "p2_course": course,
        "p2_year": str(alumni.graduation_year) if alumni.graduation_year else None,
        "p2_campus": campus,
        "p3_employed": _EMPLOYED_VALUE_KEY.get(alumni.employment_status),
        "p3_position": alumni.job_title or None,
        "p3_company": alumni.current_company or None,
    }


def _build_prefill_for_alumni(alumni, survey):
    """Resolve the prefill source into {question_id: prefill_value}.

    The same Alumni record can power many question types; this function
    walks the survey's questions, looks up the source value by ``key``,
    and resolves multiple-choice value_keys to option ids.
    """
    source = _alumni_prefill_source(alumni)
    if not source:
        return {}

    result = {}
    for question in survey.questions.all().prefetch_related("options"):
        try:
            meta = json.loads(question.help_text) if question.help_text else {}
        except (ValueError, TypeError):
            meta = {}
        key = meta.get("key") or _question_key_from_text(question.question_text)
        if not key or key not in source:
            continue
        value = source[key]
        if value is None or value == "":
            continue

        if question.question_type in ("text", "email", "phone", "number", "date", "url"):
            result[question.id] = str(value)
        elif question.question_type == "multiple_choice":
            opt = _resolve_option_id(question, key, str(value))
            if opt is not None:
                result[question.id] = opt.id
        # checkbox and rating/likert are intentionally not prefilled: the
        # source data doesn't carry structured multi-select or subjective
        # answers, and we'd rather not seed a partial selection.
    return result


# Per-question matchers for multiple-choice prefill. ``QuestionOption`` has
# no ``value_key`` column (the seeder dropped it silently), so we look up
# options by lowercased option_text for some questions and by a keyword
# mapping for others. The mapping is keyed by the question's prefill key
# (the same key used in ``ALUMNI_QUESTIONS`` in ``seed_tracer_study``).

_CAMPUS_KEYWORDS = {
    "main1": ["main campus i"],
    "main2": ["main campus ii"],
    "bayawan": ["bayawan"],
    "siaton": ["siaton"],
    "guihulngan": ["guihulngan"],
    "bais": ["bais"],
    "mabinay": ["mabinay"],
    "pamplona": ["pamplona"],
}


def _resolve_option_id(question, question_key, value_key):
    """Find the option in ``question`` that matches ``value_key``.

    Returns the QuestionOption or None. Uses a per-question strategy:
    exact case-insensitive option_text match, with a fallback to substring
    matching for known variants (e.g. "other (please specify)" -> "other").
    """
    options = list(question.options.all())
    if not options:
        return None

    if question_key == "p2_campus":
        for kw in _CAMPUS_KEYWORDS.get(value_key, []):
            for opt in options:
                if kw in opt.option_text.lower():
                    return opt
        return None

    # Default: case-insensitive match against option_text. "yes" / "no" /
    # "never" map cleanly; "other" matches "Other (please specify)" via
    # substring fallback.
    for opt in options:
        if opt.option_text.strip().lower() == value_key.lower():
            return opt
    for opt in options:
        if value_key.lower() in opt.option_text.lower():
            return opt
    return None


def _already_submitted_alumni(survey, alumni):
    return SurveyResponse.objects.filter(survey=survey, alumni=alumni).exists()


def _save_alumni_response(request, survey, alumni):
    """Create a SurveyResponse + ResponseAnswer rows for the alumni tracer study.

    Returns the new ``SurveyResponse`` instance. Mirrors the post logic in
    ``SurveyTakeView.post`` but writes to the dedicated tracer-study URL on
    success instead of the generic survey list.
    """
    response = SurveyResponse.objects.create(
        survey=survey,
        alumni=alumni,
        ip_address=request.META.get("REMOTE_ADDR"),
    )

    answer_count = 0
    for question in survey.questions.all().prefetch_related("options"):
        field_name = f"question_{question.id}"
        answer_data = {}

        if question.question_type in (
            "text", "email", "number", "phone", "url", "date", "time",
        ):
            value = request.POST.get(field_name, "").strip()
            if value:
                answer_data["text_answer"] = value

        elif question.question_type == "multiple_choice":
            option_id = request.POST.get(field_name)
            if option_id:
                try:
                    option = QuestionOption.objects.get(id=option_id)
                    answer_data["selected_option"] = option
                    if option.allow_custom:
                        custom = request.POST.get(
                            f"{field_name}_other", ""
                        ).strip()
                        if custom:
                            answer_data["custom_text"] = custom
                except QuestionOption.DoesNotExist:
                    pass

        elif question.question_type == "checkbox":
            for option in question.options.all():
                if request.POST.get(f"{field_name}_{option.id}"):
                    custom = ""
                    if option.allow_custom:
                        custom = request.POST.get(
                            f"{field_name}_other_{option.id}", ""
                        ).strip()
                    ResponseAnswer.objects.create(
                        response=response,
                        question=question,
                        selected_option=option,
                        custom_text=custom,
                    )
                    answer_count += 1
            continue

        elif question.question_type in ("rating", "likert"):
            rating = request.POST.get(field_name)
            if rating:
                try:
                    answer_data["rating_value"] = int(rating)
                except (TypeError, ValueError):
                    pass

        if answer_data:
            ResponseAnswer.objects.create(
                response=response, question=question, **answer_data
            )
            answer_count += 1

    return response, answer_count


def _find_or_create_employer(company_name, position):
    employer, _ = Employer.objects.get_or_create(
        company_name=company_name.strip(),
        position=position.strip(),
    )
    return employer


def _save_employer_response(request, survey, employer):
    response = EmployerResponse.objects.create(
        survey=survey,
        employer=employer,
        ip_address=request.META.get("REMOTE_ADDR"),
    )

    answer_count = 0
    for question in survey.questions.all().prefetch_related("options"):
        field_name = f"question_{question.id}"
        answer_data = {}

        if question.question_type in (
            "text", "email", "number", "phone", "url", "date", "time",
        ):
            value = request.POST.get(field_name, "").strip()
            if value:
                answer_data["text_answer"] = value

        elif question.question_type in ("rating", "likert"):
            rating = request.POST.get(field_name)
            if rating:
                try:
                    answer_data["rating_value"] = int(rating)
                except (TypeError, ValueError):
                    pass

        elif question.question_type in ("multiple_choice", "checkbox"):
            # Employer questionnaire currently has no MC/checkbox questions,
            # but support them defensively.
            option_id = request.POST.get(field_name)
            if option_id:
                try:
                    answer_data["selected_option"] = QuestionOption.objects.get(
                        id=option_id
                    )
                except QuestionOption.DoesNotExist:
                    pass

        if answer_data:
            EmployerResponseAnswer.objects.create(
                response=response, question=question, **answer_data
            )
            answer_count += 1

    return response, answer_count


@require_http_methods(["GET", "POST"])
@login_required
def tracer_study_alumni(request):
    survey = _get_active_survey(ALUMNI_TITLE)
    # Authed users without an alumni profile fall into one of two cases:
    #   * Staff / superuser / alumni coordinator without an Alumni row —
    #     they have no business taking the survey themselves; send them
    #     to the reports dashboard where the data lives.
    #   * Regular authed user who hasn't completed alumni registration —
    #     send them to the post-registration page so they can finish
    #     creating their Alumni profile.
    if not Alumni.objects.filter(user=request.user).exists():
        is_staff_like = (
            request.user.is_staff
            or request.user.is_superuser
            or (hasattr(request.user, "profile")
                and request.user.profile.is_alumni_coordinator)
        )
        if is_staff_like:
            return redirect("surveys:tracer_study_reports")
        messages.info(
            request,
            "Please complete your alumni registration before taking the Tracer Study.",
        )
        return redirect("accounts:post_registration")
    alumni = Alumni.objects.get(user=request.user)

    if _already_submitted_alumni(survey, alumni):
        return render(
            request,
            "tracer_study/already_submitted.html",
            {"survey": survey, "audience": "alumni"},
        )

    if request.method == "POST":
        try:
            with transaction.atomic():
                _save_alumni_response(request, survey, alumni)
        except Exception as exc:  # pragma: no cover - defensive
            messages.error(
                request,
                "An error occurred while submitting your response. Please try again.",
            )
            return redirect("surveys:tracer_study_alumni")

        return render(
            request,
            "tracer_study/thank_you.html",
            {"survey": survey, "audience": "alumni"},
        )

    questions = list(survey.questions.all().prefetch_related("options"))
    form_questions = []
    extent_questions = []
    for question in questions:
        try:
            meta = json.loads(question.help_text) if question.help_text else {}
        except (TypeError, ValueError):
            meta = {}
        key = meta.get("key") or _question_key_from_text(question.question_text)
        if key in PART_IV_EXTENT_KEYS:
            extent_questions.append({
                "question": question,
                "label": PART_IV_EXTENT_LABELS.get(key, question.question_text),
            })
        else:
            form_questions.append(question)
    prefill_answers = _build_prefill_for_alumni(alumni, survey)
    return render(
        request,
        "tracer_study/alumni_form.html",
        {
            "survey": survey,
            "questions": form_questions,
            "extent_questions": extent_questions,
            "question_count": len(questions),
            "alumni": alumni,
            "prefill_answers": prefill_answers,
        },
    )


@require_http_methods(["GET", "POST"])
def tracer_study_employer(request):
    survey = _get_active_survey(EMPLOYER_TITLE)

    if request.method == "POST":
        company = request.POST.get("company_name", "").strip()
        position = request.POST.get("position", "").strip()
        if not company or not position:
            messages.error(
                request,
                "Company Name and Position are required before submitting.",
            )
            return redirect("surveys:tracer_study_employer")

        if EmployerResponse.objects.filter(survey=survey, employer__company_name=company).exists():
            return render(
                request,
                "tracer_study/already_submitted.html",
                {"survey": survey, "audience": "employer"},
            )

        try:
            with transaction.atomic():
                employer = _find_or_create_employer(company, position)
                # Re-check the duplicate guard inside the transaction.
                if EmployerResponse.objects.filter(
                    survey=survey, employer=employer
                ).exists():
                    return render(
                        request,
                        "tracer_study/already_submitted.html",
                        {"survey": survey, "audience": "employer"},
                    )
                _save_employer_response(request, survey, employer)
        except Exception as exc:  # pragma: no cover - defensive
            import logging
            import traceback
            logging.getLogger(__name__).error(
                "Employer tracer study submit failed: %s\n%s",
                exc, traceback.format_exc(),
            )
            messages.error(
                request,
                "An error occurred while submitting your response. Please try again.",
            )
            return redirect("surveys:tracer_study_employer")

        return render(
            request,
            "tracer_study/thank_you.html",
            {"survey": survey, "audience": "employer"},
        )

    questions = survey.questions.all().prefetch_related("options")
    return render(
        request,
        "tracer_study/employer_form.html",
        {"survey": survey, "questions": questions},
    )


# ---------------------------------------------------------------------------
# Tracer Study Report
# ---------------------------------------------------------------------------
#
# A dedicated report view that aggregates the answers for the alumni or
# employer tracer study and renders them as a clean, downloadable page. The
# existing ``Report`` model is used so this report also appears in the admin
# Reports list (Report.report_type='feedback' with parameters={"survey_id": X}
# to identify it).


def _tracer_study_survey_or_404(survey_id):
    """Return the tracer study survey for the given id (or 404).

    Only the two seeded tracer study surveys are allowed — any other survey
    id returns 404 so the URL can't be used as a generic report proxy.
    """
    survey = get_object_or_404(Survey, pk=survey_id)
    if survey.title not in (ALUMNI_TITLE, EMPLOYER_TITLE):
        from django.http import Http404
        raise Http404("Not a tracer study survey.")
    return survey


SCALE_LABELS_EXTENT = {
    1: "1 - Very Low",
    2: "2 - Low",
    3: "3 - Moderate",
    4: "4 - High",
    5: "5 - Very High",
}

SCALE_LABELS_FBR = {
    1: "1 - FBR (Far Below Requirement)",
    2: "2 - BR (Below Requirement)",
    3: "3 - MR (Meeting Requirement)",
    4: "4 - AR (Above Requirement)",
    5: "5 - FBR (Far Above Requirement)",
}


TRACER_CHART_COLORS = [
    "#2b3c6b",
    "#0f766e",
    "#b45309",
    "#7c3aed",
    "#be123c",
    "#2563eb",
    "#15803d",
    "#a16207",
]


def _tracer_question_chart(question, aggregation):
    """Build Chart.js-friendly data for closed-ended tracer questions."""
    if aggregation.get("kind") not in {"choice", "checkbox", "rating"}:
        return None

    rows = aggregation.get("rows") or []
    labels = [row["label"] for row in rows]
    values = [row["count"] for row in rows]
    if not any(values):
        return None

    if aggregation["kind"] == "choice":
        chart_type = "doughnut"
    else:
        chart_type = "bar"

    return {
        "id": f"chart-q-{question.id}",
        "kind": aggregation["kind"],
        "type": chart_type,
        "labels": labels,
        "values": values,
        "colors": [TRACER_CHART_COLORS[i % len(TRACER_CHART_COLORS)] for i in range(len(labels))],
    }


def _aggregate_question(survey, question, alumni_model, employer_model):
    """Return an aggregation dict for a single question.

    Dispatches on ``question.question_type`` and uses the appropriate answer
    table (``ResponseAnswer`` for alumni, ``EmployerResponseAnswer`` for
    employer).
    """
    audience = "alumni" if survey.title == ALUMNI_TITLE else "employer"
    answer_model = ResponseAnswer if audience == "alumni" else EmployerResponseAnswer
    response_model = SurveyResponse if audience == "alumni" else EmployerResponse

    qs = answer_model.objects.filter(question=question)
    total_responses = response_model.objects.filter(survey=survey).count()

    type_label = {
        "text": "Text", "email": "Email", "number": "Number",
        "phone": "Phone", "url": "URL", "date": "Date", "time": "Time",
        "multiple_choice": "Single Choice", "checkbox": "Multiple Choice",
        "rating": "Rating", "likert": "Likert",
    }.get(question.question_type, question.question_type)

    if question.question_type in ("text", "email", "number", "phone", "url", "date", "time"):
        rows = list(
            qs.exclude(Q(text_answer__isnull=True) | Q(text_answer=""))
            .values_list("text_answer", flat=True)
        )
        return {
            "kind": "text",
            "type_label": type_label,
            "answers": rows,
            "count": len(rows),
            "total_responses": total_responses,
        }

    if question.question_type == "multiple_choice":
        option_counts = {
            o.id: o.option_text for o in question.options.all()
        }
        counts = dict(
            qs.exclude(selected_option__isnull=True)
              .values_list("selected_option_id")
              .annotate(n=Count("id"))
        )
        total = sum(counts.values()) or 0
        rows = []
        for opt_id, label in option_counts.items():
            n = counts.get(opt_id, 0)
            pct = (n / total * 100.0) if total else 0.0
            rows.append({"label": label, "count": n, "percent": pct})
        # Free-text "Other" answers
        other = list(
            qs.exclude(Q(custom_text__isnull=True) | Q(custom_text=""))
            .values_list("custom_text", flat=True)
        )
        if other:
            rows.append({"label": "Other (free text)", "count": len(other), "percent": 0.0, "other": other})
        return {
            "kind": "choice",
            "type_label": type_label,
            "rows": rows,
            "total_selections": total,
            "total_responses": total_responses,
        }

    if question.question_type == "checkbox":
        option_counts = {
            o.id: o.option_text for o in question.options.all()
        }
        counts = dict(
            qs.values_list("selected_option_id")
              .annotate(n=Count("id"))
        )
        total = sum(counts.values()) or 0
        rows = []
        for opt_id, label in option_counts.items():
            n = counts.get(opt_id, 0)
            pct = (n / total_responses * 100.0) if total_responses else 0.0
            rows.append({"label": label, "count": n, "percent": pct})
        other = list(
            qs.exclude(Q(custom_text__isnull=True) | Q(custom_text=""))
            .values_list("custom_text", flat=True)
        )
        if other:
            rows.append({"label": "Other (free text)", "count": len(other), "percent": 0.0, "other": other})
        return {
            "kind": "checkbox",
            "type_label": type_label,
            "rows": rows,
            "total_selections": total,
            "total_responses": total_responses,
        }

    if question.question_type in ("rating", "likert"):
        label_map = (
            SCALE_LABELS_EXTENT if question.scale_type == "extent" else SCALE_LABELS_FBR
        )
        buckets = {i: 0 for i in range(1, 6)}
        rating_rows = qs.exclude(rating_value__isnull=True).values_list("rating_value", flat=True)
        n = 0
        total_score = 0
        for v in rating_rows:
            if v in buckets:
                buckets[v] += 1
                n += 1
                total_score += v
        avg = (total_score / n) if n else 0.0
        rows = []
        for i in range(1, 6):
            cnt = buckets[i]
            pct = (cnt / n * 100.0) if n else 0.0
            rows.append({"label": label_map.get(i, str(i)), "count": cnt, "percent": pct})
        return {
            "kind": "rating",
            "type_label": type_label,
            "rows": rows,
            "average": avg,
            "count": n,
            "total_responses": total_responses,
        }

    return {"kind": "unknown", "type_label": type_label}


@login_required
def tracer_study_reports(request):
    """Landing page for the two tracer study reports."""
    if not _can_view_tracer_reports(request.user):
        return redirect("surveys:tracer_study_alumni")

    surveys = []
    for title, audience in ((ALUMNI_TITLE, "alumni"), (EMPLOYER_TITLE, "employer")):
        try:
            s = Survey.objects.get(title=title)
        except Survey.DoesNotExist:
            continue
        if audience == "alumni":
            count = SurveyResponse.objects.filter(survey=s).count()
        else:
            count = EmployerResponse.objects.filter(survey=s).count()
        surveys.append({"survey": s, "audience": audience, "count": count})

    return render(
        request,
        "tracer_study/reports_index.html",
        {"surveys": surveys},
    )


@login_required
def tracer_study_report(request, survey_id):
    """Aggregate report for a tracer study survey."""
    if not _can_view_tracer_reports(request.user):
        return redirect("surveys:tracer_study_alumni")

    survey = _tracer_study_survey_or_404(survey_id)
    audience = "alumni" if survey.title == ALUMNI_TITLE else "employer"

    # Upsert a Report row so this report also appears in the admin Reports
    # list and can be exported in future. parameters={"survey_id": X} is the
    # tag the existing Survey Feedback report logic uses to scope by survey.
    report, _ = Report.objects.get_or_create(
        title=f"Tracer Study Report — {audience.title()}",
        report_type="feedback",
        created_by=request.user,
        defaults={"parameters": {"survey_id": survey.id, "audience": audience}},
    )
    # Keep parameters in sync if survey_id / audience drift.
    p = report.parameters or {}
    if p.get("survey_id") != survey.id or p.get("audience") != audience:
        p["survey_id"] = survey.id
        p["audience"] = audience
        report.parameters = p
        report.save(update_fields=["parameters"])
    report.last_run = timezone.now()
    report.save(update_fields=["last_run"])

    response_model = SurveyResponse if audience == "alumni" else EmployerResponse
    total_responses = response_model.objects.filter(survey=survey).count()
    response_rows = []
    responded_rows = []
    missing_rows = []
    response_rate = None
    if audience == "alumni":
        response_rows = _alumni_response_rows(survey)
        responded_rows = [row for row in response_rows if row["submitted_at"]]
        missing_rows = [row for row in response_rows if not row["submitted_at"]]
        response_rate = (len(responded_rows) / len(response_rows) * 100) if response_rows else 0

    # Group questions by part (stored in help_text JSON)
    questions = (
        survey.questions.all()
        .prefetch_related("options")
        .order_by("display_order")
    )
    import json
    sections = []
    chart_payload = []
    current = None
    for q in questions:
        try:
            meta = json.loads(q.help_text) if q.help_text else {}
        except (ValueError, TypeError):
            meta = {}
        part = meta.get("part") or ""
        if current is None or current["part"] != part:
            current = {"part": part, "questions": []}
            sections.append(current)
        agg = _aggregate_question(survey, q, SurveyResponse, EmployerResponse)
        chart = _tracer_question_chart(q, agg)
        if chart:
            chart_payload.append(chart)
        current["questions"].append({
            "question": q,
            "aggregation": agg,
            "chart": chart,
        })

    return render(
        request,
        "tracer_study/report.html",
        {
            "survey": survey,
            "audience": audience,
            "total_responses": total_responses,
            "total_expected": len(response_rows),
            "responded_rows": responded_rows,
            "missing_rows": missing_rows,
            "response_rate": response_rate,
            "sections": sections,
            "chart_payload": chart_payload,
            "report": report,
        },
    )


@login_required
def tracer_study_filled_alumni_response_legacy(request, response_id):
    return redirect(
        "surveys:tracer_study_filled_alumni_response",
        response_token=_hashed_response_id(response_id),
    )


@login_required
def tracer_study_filled_alumni_response(request, response_token):
    if not _can_view_tracer_reports(request.user):
        return redirect("surveys:tracer_study_alumni")

    response_id = _response_id_from_token(response_token)
    response = get_object_or_404(
        SurveyResponse.objects.select_related("survey", "alumni__user").prefetch_related(
            "answers__question",
            "answers__selected_option",
        ),
        pk=response_id,
        survey__title=ALUMNI_TITLE,
    )
    return render(
        request,
        "tracer_study/filled_alumni_questionnaire.html",
        {
            "response": response,
            "survey": response.survey,
            "filled": _filled_alumni_answers(response),
            "header_data_uri": _norsu_header_data_uri(),
        },
    )


@login_required
def tracer_study_report_export(request, survey_id, format_type=None):
    if not _can_view_tracer_reports(request.user):
        return redirect("surveys:tracer_study_alumni")

    survey = _tracer_study_survey_or_404(survey_id)
    if survey.title != ALUMNI_TITLE:
        raise Http404

    format_type = (format_type or request.GET.get("format") or "excel").lower()
    if format_type == "xlsx":
        format_type = "excel"

    if format_type == "zip":
        return _tracer_study_forms_zip_response(survey)

    rows = _alumni_response_rows(survey)
    responded_rows = [row for row in rows if row["submitted_at"]]
    missing_rows = [row for row in rows if not row["submitted_at"]]
    headers = ["Alumni ID", "Name", "Email", "Program", "Graduation Year", "Status", "Submitted At"]

    def row_values(row):
        return [
            row["id"],
            row["name"],
            row["email"],
            row["course"],
            row["graduation_year"],
            row["status"],
            row["submitted_at"].strftime("%Y-%m-%d %H:%M:%S") if row["submitted_at"] else "",
        ]

    if format_type == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="tracer-study-response-status.csv"'
        writer = csv.writer(response)
        writer.writerow(["Tracer Study Response Status"])
        writer.writerow(["Survey", survey.title])
        writer.writerow(["Generated", timezone.now().strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow(["Responded", len(responded_rows), "No Response", len(missing_rows)])
        writer.writerow([])
        for title, sheet_rows in (("Alumni Who Responded", responded_rows), ("Alumni With No Response", missing_rows)):
            writer.writerow([title])
            writer.writerow(headers)
            writer.writerows(row_values(row) for row in sheet_rows)
            writer.writerow([])
        return response

    if format_type == "pdf":
        from core.export_utils import LogoHeaderService
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="tracer-study-response-status.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=90)
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Tracer Study Response Status", styles["Title"]),
            Paragraph(f"Survey: {survey.title}", styles["Normal"]),
            Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]),
            Paragraph(f"Responded: {len(responded_rows)} | No Response: {len(missing_rows)}", styles["Normal"]),
            Spacer(1, 12),
        ]
        table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b3c6b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ])
        for title, table_rows in (("Alumni Who Responded", responded_rows), ("Alumni With No Response", missing_rows)):
            story.append(Paragraph(title, styles["Heading2"]))
            data = [headers] + [row_values(row) for row in table_rows]
            table = Table(data or [headers], repeatRows=1)
            table.setStyle(table_style)
            story.extend([table, Spacer(1, 12)])
        doc.build(
            story,
            onFirstPage=lambda canvas, doc: LogoHeaderService.add_pdf_header(canvas, doc, LogoHeaderService.get_logo_path(), "Tracer Study Response Status"),
            onLaterPages=lambda canvas, doc: LogoHeaderService.add_pdf_header(canvas, doc, LogoHeaderService.get_logo_path(), "Tracer Study Response Status"),
        )
        return response

    if format_type != "excel":
        raise Http404

    from core.export_utils import LogoHeaderService
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill(start_color="2b3c6b", end_color="2b3c6b", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="2b3c6b")
    section_font = Font(bold=True, size=12, color="2b3c6b")

    def write_sheet(ws, sheet_rows, sheet_title):
        start_row = LogoHeaderService.add_excel_header(
            ws,
            LogoHeaderService.get_logo_path(),
            title="Tracer Study Response Status",
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        ws.cell(start_row, 1, "Tracer Study Response Status").font = title_font
        ws.cell(start_row, 1).alignment = Alignment(horizontal="center")
        ws.cell(start_row + 1, 1, f"Survey: {survey.title}")
        ws.cell(start_row + 2, 1, f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(start_row + 3, 1, f"Responded: {len(responded_rows)}")
        ws.cell(start_row + 3, 2, f"No Response: {len(missing_rows)}")

        table_title_row = start_row + 5
        ws.merge_cells(start_row=table_title_row, start_column=1, end_row=table_title_row, end_column=len(headers))
        ws.cell(table_title_row, 1, sheet_title).font = section_font

        header_row = table_title_row + 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(header_row, col_num, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = f"A{header_row + 1}"
        for row in sheet_rows:
            ws.append(row_values(row))
        for col_num in range(1, len(headers) + 1):
            max_length = max(len(str(ws.cell(row_num, col_num).value or "")) for row_num in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 45)

    ws = wb.active
    ws.title = "Responded"
    write_sheet(ws, responded_rows, "Alumni Who Responded")
    write_sheet(wb.create_sheet("No Response"), missing_rows, "Alumni With No Response")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="tracer-study-response-status.xlsx"'
    wb.save(response)
    return response
