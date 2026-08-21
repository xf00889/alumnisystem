import re
from datetime import timedelta
from io import BytesIO

from django.conf import settings
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from alumni_directory.models import Alumni
from surveys.models import (
    Employer,
    EmployerResponse,
    EmployerResponseAnswer,
    Report,
    ResponseAnswer,
    Survey,
    SurveyQuestion,
    SurveyResponse,
)
from surveys.tracer_metadata import (
    ALUMNI_TITLE,
    EMPLOYER_TITLE,
    build_tracer_metadata,
    resolve_report_survey,
)
from surveys.views import ReferencedSurveyUnavailable, ReportDetailView


TEST_MIDDLEWARE = [
    middleware
    for middleware in settings.MIDDLEWARE
    if middleware != "setup.middleware.SetupRequiredMiddleware"
]


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class AlumniClarificationRequirementsTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_user(
            "clarification-admin",
            "clarification-admin@example.com",
            "pass",
            is_staff=True,
        )
        self.now = timezone.now()
        self.client.force_login(self.admin)

    def make_survey(self, *, title=ALUMNI_TITLE, label="SY 2026–2027", **kwargs):
        description = kwargs.pop("description", f"{label} — Tracer study")
        defaults = {
            "title": title,
            "description": description,
            "created_by": self.admin,
            "start_date": self.now - timedelta(days=1),
            "end_date": self.now + timedelta(days=30),
            "status": "active",
        }
        defaults.update(kwargs)
        return Survey.objects.create(**defaults)

    def make_alumni(
        self,
        suffix,
        *,
        college="CAS",
        course="BSCS",
        graduation_year=2026,
    ):
        User = get_user_model()
        user = User.objects.create_user(
            f"clarification-{suffix}",
            f"clarification-{suffix}@example.com",
            "pass",
            first_name="Alumni",
            last_name=suffix.title(),
        )
        return Alumni.objects.create(
            user=user,
            college=college,
            campus="MAIN",
            course=course,
            graduation_year=graduation_year,
            gender="F",
            province="Negros Oriental",
            city="Dumaguete",
            address="Test address",
        )

    def make_question(self, survey, text, *, question_type="text", order=0):
        return SurveyQuestion.objects.create(
            survey=survey,
            question_text=text,
            question_type=question_type,
            display_order=order,
        )

    def make_report(self, survey=None, *, parameters=None, title="Feedback Report"):
        if parameters is None:
            parameters = {"survey_id": survey.pk} if survey else {}
        return Report.objects.create(
            title=title,
            description="Scoped feedback",
            report_type="feedback",
            parameters=parameters,
            created_by=self.admin,
        )

    def generate(self, report):
        return ReportDetailView().generate_report_data(report)

    def test_metadata_parsing_and_fallbacks(self):
        unrestricted = self.make_survey(label="SY 2026–2027")
        metadata = build_tracer_metadata(unrestricted)
        self.assertEqual(metadata["cycle_label"], "SY 2026–2027")
        self.assertEqual(metadata["audience_label"], "Alumni")
        self.assertEqual(metadata["target_college"], "All Colleges")
        self.assertEqual(metadata["target_program"], "All Programs")
        self.assertEqual(metadata["graduation_year_range"], "All Graduation Years")
        self.assertEqual(metadata["status"], "Open")

        unrestricted.target_college = "CAS"
        unrestricted.target_program = "STALE PROGRAM"
        unrestricted.target_graduation_year_from = 1990
        unrestricted.target_graduation_year_to = 1991
        unrestricted_metadata = build_tracer_metadata(unrestricted)
        self.assertEqual(unrestricted_metadata["target_college"], "All Colleges")
        self.assertEqual(unrestricted_metadata["target_program"], "All Programs")
        self.assertEqual(
            unrestricted_metadata["graduation_year_range"], "All Graduation Years"
        )

        restricted = self.make_survey(
            label="Cycle B",
            display_to_all=False,
            target_college="CAS",
            target_program="BSCS",
            target_graduation_year_from=2020,
            target_graduation_year_to=2024,
        )
        restricted_metadata = build_tracer_metadata(restricted)
        self.assertEqual(
            restricted_metadata["target_college"], "College of Arts and Sciences"
        )
        self.assertEqual(restricted_metadata["target_program"], "BSCS")
        self.assertEqual(restricted_metadata["graduation_year_range"], "2020–2024")

        missing_cycle = self.make_survey(description="Legacy tracer study")
        self.assertEqual(
            build_tracer_metadata(missing_cycle)["cycle_label"], "Not specified"
        )

        employer = self.make_survey(title=EMPLOYER_TITLE, label="Employer Cycle")
        employer_metadata = build_tracer_metadata(employer)
        self.assertEqual(employer_metadata["audience_label"], "Employer")
        self.assertEqual(employer_metadata["target_college"], "Not applicable")
        self.assertEqual(employer_metadata["target_program"], "Not applicable")
        self.assertEqual(
            employer_metadata["graduation_year_range"], "Not applicable"
        )

    def test_report_survey_resolution_fails_closed(self):
        survey = self.make_survey()
        report = self.make_report(survey)
        resolved, has_reference, missing = resolve_report_survey(report)
        self.assertEqual(resolved, survey)
        self.assertTrue(has_reference)
        self.assertFalse(missing)

        report.parameters = {"survey_id": "invalid"}
        report.save(update_fields=["parameters"])
        self.assertEqual(resolve_report_survey(report), (None, True, True))
        with self.assertRaises(ReferencedSurveyUnavailable):
            self.generate(report)

        report.parameters = {"survey_id": 999999999}
        report.save(update_fields=["parameters"])
        with self.assertRaises(ReferencedSurveyUnavailable):
            self.generate(report)

    def test_creation_form_contains_exact_cycle_and_year_guidance(self):
        response = self.client.get(reverse("core:create_tracer_study"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Names this study period, e.g. the school year. This label does not limit who can respond.",
        )
        self.assertContains(
            response,
            "The endpoints are inclusive, must be supplied together, and are independent of the Cycle Label.",
        )

    def test_survey_management_distinguishes_cycles_and_audiences(self):
        first = self.make_survey(
            label="SY 2025–2026",
            display_to_all=False,
            target_college="CAS",
            target_program="BSCS",
            target_graduation_year_from=2025,
            target_graduation_year_to=2026,
        )
        second = self.make_survey(
            label="SY 2026–2027",
            display_to_all=False,
            target_college="CAS",
            target_program="BSCS",
            target_graduation_year_from=2026,
            target_graduation_year_to=2027,
        )
        employer_survey = self.make_survey(
            title=EMPLOYER_TITLE,
            label="Employer 2026",
        )
        eligible_one = self.make_alumni("eligible-one", graduation_year=2026)
        self.make_alumni("eligible-two", graduation_year=2027)
        historical = self.make_alumni("historical", graduation_year=2024)
        SurveyResponse.objects.create(survey=second, alumni=eligible_one)
        SurveyResponse.objects.create(survey=second, alumni=historical)
        employer = Employer.objects.create(company_name="Example Corp", position="HR")
        EmployerResponse.objects.create(survey=employer_survey, employer=employer)

        response = self.client.get(reverse("surveys:survey_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cycle: SY 2025–2026", count=2)
        self.assertContains(response, "Cycle: SY 2026–2027", count=2)
        self.assertContains(response, "Audience: Employer", count=2)
        self.assertContains(
            response,
            "Target: College of Arts and Sciences · BSCS",
            count=4,
        )

        by_id = {item["id"]: item for item in response.context["survey_data"]}
        self.assertEqual(by_id[first.pk]["responses_count"], 0)
        self.assertEqual(by_id[second.pk]["responses_count"], 1)
        self.assertEqual(by_id[second.pk]["response_rate"], 50.0)
        self.assertEqual(by_id[employer_survey.pk]["responses_count"], 1)
        self.assertFalse(by_id[employer_survey.pk]["response_rate_available"])

    def test_alumni_feedback_is_survey_and_cohort_scoped(self):
        target = self.make_survey(
            label="Target Cycle",
            display_to_all=False,
            target_college="CAS",
            target_program="BSCS",
            target_graduation_year_from=2026,
            target_graduation_year_to=2027,
        )
        other = self.make_survey(label="Other Cycle")
        target_question = self.make_question(target, "Target feedback")
        other_question = self.make_question(other, "Other feedback")
        eligible = self.make_alumni("target", graduation_year=2026)
        self.make_alumni("target-missing", graduation_year=2027)
        historical = self.make_alumni("old", graduation_year=2024)

        eligible_response = SurveyResponse.objects.create(survey=target, alumni=eligible)
        historical_response = SurveyResponse.objects.create(survey=target, alumni=historical)
        other_response = SurveyResponse.objects.create(survey=other, alumni=historical)
        ResponseAnswer.objects.create(
            response=eligible_response,
            question=target_question,
            text_answer="INCLUDED TARGET ANSWER",
        )
        ResponseAnswer.objects.create(
            response=historical_response,
            question=target_question,
            text_answer="OUT OF COHORT ANSWER",
        )
        ResponseAnswer.objects.create(
            response=other_response,
            question=other_question,
            text_answer="OTHER SURVEY ANSWER",
        )

        data = self.generate(self.make_report(target))
        rendered_answers = [row["answer"] for row in data["table_data"]]
        self.assertEqual(rendered_answers, ["INCLUDED TARGET ANSWER"])
        self.assertEqual(data["summary"]["total_responses"], 1)
        self.assertEqual(data["summary"]["expected_respondents"], 2)
        self.assertEqual(data["summary"]["response_rate"], "50.0%")
        self.assertEqual(data["survey_context"]["cycle_label"], "Target Cycle")

    def test_employer_and_legacy_feedback_behaviors(self):
        alumni_survey = self.make_survey(label="Alumni Cycle")
        alumni_question = self.make_question(alumni_survey, "Alumni feedback")
        alumni = self.make_alumni("legacy")
        alumni_response = SurveyResponse.objects.create(
            survey=alumni_survey, alumni=alumni
        )
        ResponseAnswer.objects.create(
            response=alumni_response,
            question=alumni_question,
            text_answer="ALUMNI GLOBAL ANSWER",
        )

        employer_survey = self.make_survey(
            title=EMPLOYER_TITLE, label="Employer Cycle"
        )
        employer_question = self.make_question(employer_survey, "Employer feedback")
        employer = Employer.objects.create(
            company_name="Scoped Employer", position="Manager"
        )
        employer_response = EmployerResponse.objects.create(
            survey=employer_survey, employer=employer
        )
        EmployerResponseAnswer.objects.create(
            response=employer_response,
            question=employer_question,
            text_answer="EMPLOYER SCOPED ANSWER",
        )

        employer_data = self.generate(self.make_report(employer_survey))
        self.assertEqual(employer_data["respondent_label"], "Employer")
        self.assertEqual(employer_data["summary"]["total_responses"], 1)
        self.assertEqual(
            employer_data["table_data"][0]["answer"], "EMPLOYER SCOPED ANSWER"
        )
        self.assertEqual(
            employer_data["table_data"][0]["respondent"], "Scoped Employer"
        )

        global_data = self.generate(self.make_report(parameters={}))
        self.assertFalse(global_data["is_survey_scoped"])
        self.assertIn(
            "ALUMNI GLOBAL ANSWER",
            [row["answer"] for row in global_data["table_data"]],
        )
        self.assertNotIn(
            "EMPLOYER SCOPED ANSWER",
            [row["answer"] for row in global_data["table_data"]],
        )

    def test_feedback_detail_uses_full_text_and_25_row_pagination(self):
        survey = self.make_survey(label="Paginated Cycle")
        alumni = self.make_alumni("pagination")
        survey_response = SurveyResponse.objects.create(survey=survey, alumni=alumni)
        long_answer = "FULL START " + ("wrapped feedback " * 80) + " FULL END"
        for index in range(26):
            question = self.make_question(
                survey, f"Feedback question {index}", order=index
            )
            ResponseAnswer.objects.create(
                response=survey_response,
                question=question,
                text_answer=long_answer if index == 0 else f"Answer {index}",
            )
        report = self.make_report(survey, parameters={
            "survey_id": survey.pk,
            "audience": "alumni",
            "technical_option": "visible",
        })

        page_one = self.client.get(reverse("surveys:report_detail", args=[report.pk]))
        self.assertEqual(page_one.status_code, 200)
        self.assertEqual(len(page_one.context["report_data"]["table_data"]), 25)
        self.assertEqual(page_one.context["feedback_page"].paginator.count, 26)
        self.assertContains(page_one, "Survey Context")
        self.assertContains(page_one, "Paginated Cycle")
        self.assertEqual(
            page_one.context["display_parameters"],
            {"technical_option": "visible"},
        )
        self.assertNotContains(page_one, "survey_id")
        self.assertNotContains(page_one, "audience</span>")
        self.assertContains(page_one, "FULL END")

        page_two = self.client.get(
            reverse("surveys:report_detail", args=[report.pk]),
            {"feedback_page": 2},
        )
        self.assertEqual(len(page_two.context["report_data"]["table_data"]), 1)

    def test_feedback_pdf_is_landscape_multipage_and_accepts_long_urls(self):
        survey = self.make_survey(label="PDF Cycle")
        alumni = self.make_alumni("pdf")
        response = SurveyResponse.objects.create(survey=survey, alumni=alumni)
        long_url = "https://example.com/" + ("very-long-path-segment/" * 20)
        for index in range(30):
            question = self.make_question(survey, f"PDF question {index}", order=index)
            ResponseAnswer.objects.create(
                response=response,
                question=question,
                text_answer=f"{long_url}?row={index}&source=clarification",
            )
        report = self.make_report(survey, title="Long Feedback PDF")

        pdf = self.client.get(reverse("surveys:report_export_pdf", args=[report.pk]))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))
        media_boxes = re.findall(
            rb"/MediaBox\s*\[\s*0\s+0\s+([0-9.]+)\s+([0-9.]+)\s*\]",
            pdf.content,
        )
        self.assertTrue(media_boxes)
        width, height = map(float, media_boxes[0])
        self.assertGreater(width, height)
        page_counts = [int(value) for value in re.findall(rb"/Count\s+(\d+)", pdf.content)]
        self.assertTrue(page_counts)
        self.assertGreater(max(page_counts), 1)

    def test_response_status_exports_include_metadata_and_preserve_rows(self):
        survey = self.make_survey(
            label="SY 2026–2027",
            display_to_all=False,
            target_college="CAS",
            target_program="BSCS",
            target_graduation_year_from=2025,
            target_graduation_year_to=2027,
        )
        responded = self.make_alumni("export-responded", graduation_year=2026)
        missing = self.make_alumni("export-missing", graduation_year=2027)
        SurveyResponse.objects.create(survey=survey, alumni=responded)
        export_url = reverse(
            "surveys:tracer_study_report_export", args=[survey.pk]
        )
        query = {
            "college": "CAS",
            "year_from": 2026,
            "year_to": 2027,
        }

        csv_response = self.client.get(export_url, {**query, "format": "csv"})
        csv_text = csv_response.content.decode("utf-8")
        self.assertIn("Cycle Label,SY 2026–2027", csv_text)
        self.assertIn("Audience,Alumni", csv_text)
        self.assertIn("Target College,College of Arts and Sciences", csv_text)
        self.assertIn("Target Program,BSCS", csv_text)
        self.assertIn("Graduation Years,2025–2027", csv_text)
        self.assertIn("Applied Filters", csv_text)
        self.assertIn("Graduation year from: 2026", csv_text)
        self.assertIn("Alumni Export-Responded", csv_text)
        self.assertIn("Alumni Export-Missing", csv_text)

        pdf_response = self.client.get(export_url, {**query, "format": "pdf"})
        self.assertEqual(pdf_response.status_code, 200)
        self.assertTrue(pdf_response.content.startswith(b"%PDF"))

        excel_response = self.client.get(export_url, {**query, "format": "excel"})
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(excel_response.content))
        self.assertEqual(workbook.sheetnames, ["Responded", "No Response", "Summary"])
        for sheet_name in workbook.sheetnames:
            values = [
                cell.value
                for row in workbook[sheet_name].iter_rows()
                for cell in row
            ]
            self.assertIn("Cycle Label", values)
            self.assertIn("SY 2026–2027", values)
            self.assertIn("Target College", values)
            self.assertIn("College of Arts and Sciences", values)
            self.assertIn("Target Program", values)
            self.assertIn("BSCS", values)
            self.assertIn("Graduation Years", values)
            self.assertIn("2025–2027", values)
            self.assertIn("Applied Filters", values)

        responded_values = [
            cell.value
            for row in workbook["Responded"].iter_rows()
            for cell in row
        ]
        missing_values = [
            cell.value
            for row in workbook["No Response"].iter_rows()
            for cell in row
        ]
        self.assertIn("Alumni Export-Responded", responded_values)
        self.assertIn("Alumni Export-Missing", missing_values)
