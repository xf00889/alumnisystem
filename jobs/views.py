from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.core.cache import cache
from django.contrib import messages
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db import transaction
from django.core.mail import send_mail
from django.conf import settings
from django_ratelimit.decorators import ratelimit
from .models import JobPosting, JobApplication, RequiredDocument, ScrapedJob
from .forms import JobPostingForm, JobApplicationForm, RequiredDocumentFormSet, JobPreferenceForm
from .scraper_forms import JobScraperForm
from .scraper_utils import scraper
from .utils import calculate_job_match_score, get_skill_recommendations
from accounts.models import Profile, SkillMatch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import logging
import json
import os
from datetime import datetime

logger = logging.getLogger(__name__)

def is_hr_or_admin(user):
    """Check if user is HR, admin, or alumni coordinator"""
    if not user.is_authenticated:
        return False
    
    # Superusers and staff always have access
    if user.is_superuser or user.is_staff:
        return True
    
    # Check HR status or alumni coordinator status
    try:
        return user.profile.is_hr or user.profile.is_alumni_coordinator
    except:
        return False

def careers(request):
    """Public careers page for job listings"""
    jobs = JobPosting.objects.filter(is_active=True).order_by('-posted_date')
    
    # Get filter parameters (for future implementation)
    job_type = request.GET.get('job_type')
    location = request.GET.get('location')
    category = request.GET.get('category')
    search_query = request.GET.get('q')
    
    # Apply filters (prepared for future implementation)
    if job_type:
        jobs = jobs.filter(job_type=job_type)
    if location:
        jobs = jobs.filter(location__icontains=location)
    if category:
        jobs = jobs.filter(category=category)
    if search_query:
        jobs = jobs.filter(
            Q(job_title__icontains=search_query) |
            Q(company_name__icontains=search_query) |
            Q(job_description__icontains=search_query)
        )
    
    # Get featured jobs
    featured_jobs = jobs.filter(is_featured=True)[:3]
    
    # Get recent jobs (excluding featured ones)
    recent_jobs = jobs.exclude(is_featured=True)[:10]
    
    # Get job statistics
    stats = {
        'total_jobs': jobs.count(),
        'featured_jobs': featured_jobs.count(),
        'recent_jobs': recent_jobs.count(),
    }
    
    # Prepare filter options for future use
    filter_options = {
        'job_types': JobPosting.JOB_TYPE_CHOICES,
        'categories': JobPosting.CATEGORY_CHOICES,
        'locations': JobPosting.objects.filter(is_active=True).values_list('location', flat=True).distinct().order_by('location'),
    }
    
    # Add breadcrumbs for SEO
    breadcrumbs = [
        {'name': 'Home', 'url': '/'},
        {'name': 'Careers', 'url': '/jobs/careers/'}
    ]
    
    context = {
        'featured_jobs': featured_jobs,
        'recent_jobs': recent_jobs,
        'jobs': jobs,  # Pass jobs queryset for structured data
        'stats': stats,
        'filter_options': filter_options,
        'current_filters': {
            'job_type': job_type,
            'location': location,
            'category': category,
            'search_query': search_query,
        },
        'page_title': 'Career Opportunities - NORSU Alumni',
        'page_description': 'Discover exciting career opportunities and job postings from NORSU alumni network. Find your next career move with our exclusive job board.',
        'breadcrumbs': breadcrumbs,
    }
    return render(request, 'jobs/careers.html', context)

def get_job_details(request, job_id):
    """Get job details as JSON for modal display"""
    try:
        job = get_object_or_404(JobPosting, id=job_id, is_active=True)
        
        # Prepare job data for JSON response
        job_data = {
            'id': job.id,
            'job_title': job.job_title,
            'company_name': job.company_name,
            'location': job.location,
            'job_type': job.get_job_type_display(),
            'job_description': job.job_description,
            'requirements': job.requirements,
            'responsibilities': job.responsibilities,
            'experience_level': job.get_experience_level_display(),
            'skills_required': job.skills_required,
            'education_requirements': job.education_requirements,
            'benefits': job.benefits,
            'salary_range': job.salary_range,
            'application_link': job.application_link,
            'posted_date': job.posted_date.strftime('%B %d, %Y'),
            'is_featured': job.is_featured,
            'category': job.get_category_display(),
            'source_type': job.source_type,
            'source_type_display': job.get_source_type_display(),
            'accepts_internal_applications': job.accepts_internal_applications,
        }
        
        return JsonResponse({
            'success': True,
            'job': job_data,
            'user': {
                'is_authenticated': request.user.is_authenticated,
                'is_alumni': request.user.is_authenticated,  # For now, assume authenticated users are alumni
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting job details: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to load job details'
        }, status=500)

def check_job_application_eligibility(request, job_id):
    """Check if user can apply for a specific job based on source_type restrictions"""
    try:
        job = get_object_or_404(JobPosting, id=job_id, is_active=True)
        
        # Check source_type restrictions
        if job.source_type == 'INTERNAL':
            if not request.user.is_authenticated:
                return JsonResponse({
                    'success': False,
                    'eligible': False,
                    'message': 'Please log in or register to apply for this job.',
                    'redirect_url': '/accounts/login/'
                })
        
        return JsonResponse({
            'success': True,
            'eligible': True,
            'message': 'You are eligible to apply for this job.'
        })
        
    except Exception as e:
        logger.error(f"Error checking job application eligibility: {str(e)}")
        return JsonResponse({
            'success': False,
            'eligible': False,
            'message': 'Unable to check application eligibility.'
        }, status=500)

def job_list(request):
    jobs = JobPosting.objects.filter(is_active=True)
    
    # Get scraped jobs
    scraped_jobs = ScrapedJob.objects.filter(is_active=True).order_by('-scraped_at')
    
    # Scraper/external view is disabled; keep job board on posted jobs only.
    view_type = 'posted'
    
    # Check if user is admin/HR using existing function
    is_admin_or_hr_user = is_hr_or_admin(request.user)
    
    # Initialize preference-related variables
    show_preference_modal = False
    preferences_configured = False
    matching_jobs_count = 0
    active_filters = []
    preference_service = None
    
    # For alumni users (not admin/HR), check preferences and apply filtering
    if request.user.is_authenticated and not is_admin_or_hr_user:
        from .preference_filter import PreferenceFilterService
        
        # Get PreferenceFilterService instance
        preference_service = PreferenceFilterService(request.user)
        
        # Check if preferences are configured
        preferences_configured = preference_service.preferences.is_configured
        
        # Determine if modal should be shown
        show_preference_modal = preference_service.should_show_modal(request.session)
        
        # Handle "show_all" toggle from GET parameter
        show_all = request.GET.get('show_all', 'false').lower() == 'true'
        
        # Apply preference filtering if configured and not showing all
        if preferences_configured and not show_all:
            # Get filtered jobs using service
            job_matches = preference_service.get_filtered_jobs()
            
            # Extract jobs and match data
            jobs = JobPosting.objects.filter(
                id__in=[match['job'].id for match in job_matches]
            )
            
            # Calculate matching_jobs_count
            matching_jobs_count = len(job_matches)
            
            # Create match_data dictionary for template
            match_data = {match['job'].id: match for match in job_matches}
            
            # Build active_filters list for display
            prefs = preference_service.preferences
            
            if prefs.job_types:
                job_type_labels = dict(JobPosting.JOB_TYPE_CHOICES)
                for jt in prefs.job_types:
                    active_filters.append({
                        'key': f'job_type_{jt}',
                        'label': job_type_labels.get(jt, jt)
                    })
            
            if prefs.remote_only:
                active_filters.append({
                    'key': 'remote_only',
                    'label': 'Remote Only'
                })
            
            if prefs.location_text and not prefs.willing_to_relocate:
                active_filters.append({
                    'key': 'location',
                    'label': f'Location: {prefs.location_text}'
                })
            
            if prefs.minimum_salary:
                active_filters.append({
                    'key': 'minimum_salary',
                    'label': f'Min Salary: ₱{prefs.minimum_salary:,}'
                })
            
            if prefs.source_type != 'BOTH':
                source_type_labels = dict(JobPosting.SOURCE_TYPE_CHOICES)
                active_filters.append({
                    'key': 'source_type',
                    'label': source_type_labels.get(prefs.source_type, prefs.source_type)
                })
            
            if prefs.skill_matching_enabled:
                active_filters.append({
                    'key': 'skill_matching',
                    'label': f'Skill Match ≥{prefs.skill_match_threshold}%'
                })
        else:
            # No filtering applied
            match_data = {}
    else:
        # Admin/HR or unauthenticated users see all jobs without filtering
        match_data = {}
    
    # Search functionality
    query = request.GET.get('q')
    if query:
        if view_type == 'scraped':
            scraped_jobs = scraped_jobs.filter(
                Q(search_keyword__icontains=query) |
                Q(search_location__icontains=query)
            )
        else:
            jobs = jobs.filter(
                Q(job_title__icontains=query) |
                Q(company_name__icontains=query) |
                Q(location__icontains=query) |
                Q(job_description__icontains=query)
            )
    
    # Filtering
    job_type = request.GET.get('job_type')
    source_type = request.GET.get('source_type')
    
    if job_type and view_type == 'posted':
        jobs = jobs.filter(job_type=job_type)
    if source_type:
        if view_type == 'scraped':
            scraped_jobs = scraped_jobs.filter(source=source_type)
        else:
            jobs = jobs.filter(source_type=source_type)
    
    # Skill-based view toggle (only for posted jobs)
    skill_based_view = request.GET.get('skill_based', 'false').lower() == 'true'
    
    # Get user profile if authenticated and skill-based view is enabled
    user_profile = None
    job_matches = []
    recommended_skills = []
    
    if request.user.is_authenticated and skill_based_view and view_type == 'posted':
        try:
            user_profile = Profile.objects.get(user=request.user)
            
            # Get job matches from database
            skill_matches = SkillMatch.objects.filter(
                profile=user_profile, 
                job__in=jobs
            ).select_related('job')
            
            # If no matches exist yet or all are outdated, calculate them on the fly
            if not skill_matches.exists():
                for job in jobs:
                    score, matched_skills, missing_skills = calculate_job_match_score(
                        user_profile, job, required_skills_only=False
                    )
                    if score > 0:
                        job_matches.append({
                            'job': job,
                            'score': score,
                            'matched_skills': matched_skills,
                            'missing_skills': missing_skills
                        })
                
                # Sort by score descending
                job_matches.sort(key=lambda x: x['score'], reverse=True)
            else:
                # Convert existing matches to the expected format
                for match in skill_matches:
                    job_matches.append({
                        'job': match.job,
                        'score': match.match_score,
                        'matched_skills': json.loads(match.matched_skills) if match.matched_skills else {},
                        'missing_skills': json.loads(match.missing_skills) if match.missing_skills else {}
                    })
            
            # Sort by score descending
            job_matches.sort(key=lambda x: x['score'], reverse=True)
            
            # Get skill recommendations
            recommended_skills = get_skill_recommendations(user_profile)
        except Profile.DoesNotExist:
            pass
        
    # Sorting
    sort = request.GET.get('sort')
    if skill_based_view and job_matches and view_type == 'posted':
        # Already sorted by match score above
        pass
    elif sort == 'oldest':
        jobs = jobs.order_by('posted_date')
    else:  # newest first is default
        jobs = jobs.order_by('-posted_date')
    
    # Featured jobs
    featured_jobs = jobs.filter(is_featured=True)[:5]
    
    # For skill-based view, use the job_matches list
    scraped_jobs_total_count = 0
    if skill_based_view and user_profile and view_type == 'posted':
        paginator = Paginator([match['job'] for match in job_matches], 10)
        page = request.GET.get('page')
        jobs_page = paginator.get_page(page)
        
        # Create a map for easy lookup of match data (merge with preference match_data if exists)
        skill_match_data = {match['job'].id: match for match in job_matches}
        if match_data:
            # Merge skill matching data with preference matching data
            for job_id, skill_match in skill_match_data.items():
                if job_id in match_data:
                    match_data[job_id]['skill_score'] = skill_match['score']
                    match_data[job_id]['matched_skills'] = skill_match.get('matched_skills', {})
                    match_data[job_id]['missing_skills'] = skill_match.get('missing_skills', {})
                else:
                    match_data[job_id] = skill_match
        else:
            match_data = skill_match_data
    elif view_type == 'scraped':
        # Keep only scrape runs that actually have job entries
        scraped_jobs_with_data = [entry for entry in scraped_jobs if entry.jobs_count > 0]
        scraped_jobs_total_count = sum(entry.jobs_count for entry in scraped_jobs_with_data)

        # Pagination for scraped jobs
        paginator = Paginator(scraped_jobs_with_data, 10)
        page = request.GET.get('page')
        jobs_page = paginator.get_page(page)
    else:
        # Pagination for regular view
        paginator = Paginator(jobs, 10)
        page = request.GET.get('page')
        jobs_page = paginator.get_page(page)
    
    # Get scraped source choices
    scraped_source_choices = getattr(ScrapedJob, 'SOURCE_CHOICES', [])
    
    context = {
        'jobs': jobs_page,
        'featured_jobs': featured_jobs,
        'current_query': query,
        'current_job_type': job_type,
        'current_source_type': source_type,
        'current_sort': sort,
        'job_types': JobPosting.JOB_TYPE_CHOICES,
        'source_types': JobPosting.SOURCE_TYPE_CHOICES,
        'scraped_source_choices': scraped_source_choices,
        'skill_based_view': skill_based_view,
        'match_data': match_data,
        'recommended_skills': recommended_skills[:5] if recommended_skills else [],  # Show top 5 recommendations
        'view_type': view_type,
        'scraped_jobs': scraped_jobs if view_type == 'scraped' else None,
        'scraped_jobs_total_count': scraped_jobs_total_count,
        # Preference-related context
        'show_preference_modal': show_preference_modal,
        'preferences_configured': preferences_configured,
        'matching_jobs_count': matching_jobs_count,
        'active_filters': active_filters,
        'is_admin_or_hr': is_admin_or_hr_user,
        'show_all': request.GET.get('show_all', 'false').lower() == 'true',
    }
    return render(request, 'jobs/job_list.html', context)


@login_required
@ratelimit(key='user', rate='10/m', method='POST')
def save_preferences(request):
    """
    Save user job preferences.
    
    This view handles the submission of the job preference configuration form.
    It validates the form data, saves the preferences, and clears the user's
    preference cache to ensure updated results are shown.
    
    Rate limited to 10 requests per minute per user to prevent abuse.
    """
    # Check if request was rate limited
    if getattr(request, 'limited', False):
        logger.warning(f"Rate limit exceeded for user {request.user.id} on save_preferences")
        return JsonResponse({
            'success': False,
            'error': 'Too many requests. Please try again in a moment.'
        }, status=429)
    
    # Only accept POST requests
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Invalid request method. Only POST is allowed.'
        }, status=405)
    
    try:
        # Get or create user's job preferences
        from .models import JobPreference
        preferences, created = JobPreference.objects.get_or_create(user=request.user)
        
        # Instantiate form with POST data and existing preferences instance
        form = JobPreferenceForm(request.POST, instance=preferences)
        
        # Validate form
        if form.is_valid():
            try:
                # Use atomic transaction to ensure data consistency
                with transaction.atomic():
                    # Save the preferences
                    preferences = form.save(commit=False)
                    
                    # Mark preferences as configured
                    preferences.is_configured = True
                    
                    # Save to database
                    preferences.save()
                    
                    # Clear user's preference cache
                    cache_key = f"job_preferences_{request.user.id}"
                    cache.delete(cache_key)
                    
                    logger.info(f"User {request.user.id} successfully saved job preferences")
                    
                    # Return success response
                    return JsonResponse({
                        'success': True,
                        'message': 'Your job preferences have been saved successfully!'
                    })
                    
            except Exception as e:
                # Log the error with full traceback
                logger.error(
                    f"Error saving preferences for user {request.user.id}: {str(e)}",
                    exc_info=True
                )
                
                # Return error response
                return JsonResponse({
                    'success': False,
                    'error': 'An error occurred while saving your preferences. Please try again.'
                }, status=500)
        else:
            # Form validation failed - return errors
            logger.warning(
                f"Form validation failed for user {request.user.id}: {form.errors.as_json()}"
            )
            
            # Format errors for JSON response
            errors = {}
            for field, error_list in form.errors.items():
                errors[field] = [str(error) for error in error_list]
            
            return JsonResponse({
                'success': False,
                'error': 'Please correct the errors in the form.',
                'errors': errors
            }, status=400)
            
    except Exception as e:
        # Catch any unexpected errors
        logger.error(
            f"Unexpected error in save_preferences for user {request.user.id}: {str(e)}",
            exc_info=True
        )
        
        return JsonResponse({
            'success': False,
            'error': 'An unexpected error occurred. Please try again later.'
        }, status=500)


@login_required
def remind_later(request):
    """
    Handle "Remind me later" action for job preference modal.
    
    This view marks that the user was prompted to configure preferences
    and sets a session flag to prevent showing the modal again in the
    current session.
    """
    # Only accept POST requests
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Invalid request method. Only POST is allowed.'
        }, status=405)
    
    try:
        # Get or create user's job preferences
        from .models import JobPreference
        preferences, created = JobPreference.objects.get_or_create(user=request.user)
        
        # Mark that user was prompted
        preferences.was_prompted = True
        preferences.save()
        
        # Set session flag to prevent showing modal again in this session
        request.session['preference_modal_prompted'] = True
        
        logger.info(f"User {request.user.id} chose 'remind me later' for job preferences")
        
        # Return success response
        return JsonResponse({
            'success': True,
            'message': 'You can configure your preferences anytime from the job board.'
        })
        
    except Exception as e:
        # Log the error with full traceback
        logger.error(
            f"Error in remind_later for user {request.user.id}: {str(e)}",
            exc_info=True
        )
        
        # Return error response
        return JsonResponse({
            'success': False,
            'error': 'An error occurred. Please try again.'
        }, status=500)


@login_required
def remove_filter(request, filter_key):
    """
    Remove a specific filter from user's job preferences.
    
    This view clears a specific hard filter from the user's preferences,
    invalidates the cache, and redirects back to the job list.
    
    Args:
        filter_key: The preference field to clear (e.g., 'job_types', 'location_text', 
                   'remote_only', 'minimum_salary', 'source_type')
    """
    try:
        # Get user's job preferences
        from .models import JobPreference
        preferences = JobPreference.objects.get(user=request.user)
        
        # Map of valid filter keys to their default values
        filter_defaults = {
            'job_types': [],
            'location_text': '',
            'remote_only': False,
            'minimum_salary': None,
            'source_type': 'BOTH',
        }
        
        # Validate filter_key
        if filter_key not in filter_defaults:
            logger.warning(
                f"Invalid filter_key '{filter_key}' provided by user {request.user.id}"
            )
            messages.error(request, 'Invalid filter specified.')
            return redirect('jobs:job_list')
        
        # Clear the specific filter by setting it to default value
        setattr(preferences, filter_key, filter_defaults[filter_key])
        preferences.save()
        
        # Cache is automatically invalidated by the model's save method
        
        logger.info(
            f"User {request.user.id} removed filter '{filter_key}' from preferences"
        )
        
        messages.success(request, f'Filter removed successfully.')
        
    except JobPreference.DoesNotExist:
        logger.warning(
            f"User {request.user.id} attempted to remove filter but has no preferences"
        )
        messages.error(request, 'No preferences found.')
    
    except Exception as e:
        # Log the error with full traceback
        logger.error(
            f"Error removing filter '{filter_key}' for user {request.user.id}: {str(e)}",
            exc_info=True
        )
        messages.error(request, 'An error occurred while removing the filter.')
    
    # Redirect back to job list
    return redirect('jobs:job_list')


@login_required
def clear_all_filters(request):
    """
    Clear all hard filters from user's job preferences.
    
    This view resets all hard filter fields to their default values while
    preserving soft filter preferences (industries, experience_levels).
    The cache is automatically invalidated by the model's save method.
    """
    try:
        # Get user's job preferences
        from .models import JobPreference
        preferences = JobPreference.objects.get(user=request.user)
        
        # Clear all hard filters (preserve soft filters)
        preferences.job_types = []
        preferences.location_text = ''
        preferences.remote_only = False
        preferences.minimum_salary = None
        preferences.source_type = 'BOTH'
        preferences.skill_matching_enabled = False
        
        # Save preferences (cache is automatically invalidated by model's save method)
        preferences.save()
        
        logger.info(
            f"User {request.user.id} cleared all hard filters from preferences"
        )
        
        messages.success(request, 'All filters have been cleared successfully.')
        
    except JobPreference.DoesNotExist:
        logger.warning(
            f"User {request.user.id} attempted to clear filters but has no preferences"
        )
        messages.error(request, 'No preferences found.')
    
    except Exception as e:
        # Log the error with full traceback
        logger.error(
            f"Error clearing all filters for user {request.user.id}: {str(e)}",
            exc_info=True
        )
        messages.error(request, 'An error occurred while clearing filters.')
    
    # Redirect back to job list
    return redirect('jobs:job_list')


def job_detail(request, slug):
    job = get_object_or_404(JobPosting, slug=slug)
    user_application = None
    applicants = None
    
    if request.user.is_authenticated:
        user_application = JobApplication.objects.filter(job=job, applicant=request.user).first()
        
        # Add applicants list for admin/HR users
        if request.user.is_staff or is_hr_or_admin(request.user):
            applicants = job.applications.all().select_related('applicant__profile')
    
    context = {
        'job': job,
        'user_application': user_application,
        'applicants': applicants,
    }
    return render(request, 'jobs/job_detail.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def manage_jobs(request):
    jobs = JobPosting.objects.all()
    
    # Get statistics
    stats = {
        'active_jobs': jobs.filter(is_active=True).count(),
        'internal_jobs': jobs.filter(source_type='INTERNAL').count(),
        'total_applications': JobApplication.objects.count(),
        'pending_applications': JobApplication.objects.filter(status='PENDING').count(),
    }
    
    # Pagination
    paginator = Paginator(jobs, 10)
    page = request.GET.get('page')
    jobs = paginator.get_page(page)
    
    context = {
        'jobs': jobs,
        'stats': stats,
    }
    return render(request, 'jobs/manage_jobs.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def post_job(request):
    if request.method == 'POST':
        form = JobPostingForm(request.POST)
        document_formset = RequiredDocumentFormSet(request.POST)
        
        if form.is_valid() and document_formset.is_valid():
            try:
                with transaction.atomic():
                    # Save the job posting
                    job = form.save(commit=False)
                    job.posted_by = request.user
                    job.source = 'manual'
                    job.save()
                    
                    # Save document requirements for internal jobs
                    if job.source_type == 'INTERNAL':
                        document_formset.instance = job
                        document_formset.save()
                    
                    messages.success(request, 'Job posting created successfully!')
                    return redirect('jobs:job_detail', slug=job.slug)
            except Exception as e:
                messages.error(request, f'Error creating job posting: {str(e)}')
        else:
            # Add form errors to messages
            if form.errors:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
            if document_formset.errors:
                for form_errors in document_formset.errors:
                    for field, errors in form_errors.items():
                        for error in errors:
                            messages.error(request, f'Document requirement - {field}: {error}')
            if document_formset.non_form_errors():
                for error in document_formset.non_form_errors():
                    messages.error(request, f'Document requirements: {error}')
    else:
        form = JobPostingForm(initial={'is_active': True})  # Set is_active to True by default
        document_formset = RequiredDocumentFormSet()
    
    return render(request, 'jobs/post_job.html', {
        'form': form,
        'document_formset': document_formset
    })

@login_required
@user_passes_test(is_hr_or_admin)
def edit_job(request, slug):
    job = get_object_or_404(JobPosting, slug=slug)
    if request.method == 'POST':
        form = JobPostingForm(request.POST, instance=job)
        document_formset = RequiredDocumentFormSet(request.POST, instance=job)
        
        if form.is_valid() and document_formset.is_valid():
            try:
                with transaction.atomic():
                    job = form.save()
                    
                    if job.source_type == 'INTERNAL':
                        document_formset.save()
                    else:
                        # Remove all document requirements for external jobs
                        job.required_documents.all().delete()
                    
                    messages.success(request, 'Job posting updated successfully!')
                    return redirect('jobs:job_detail', slug=job.slug)
            except Exception as e:
                messages.error(request, f'Error updating job posting: {str(e)}')
    else:
        form = JobPostingForm(instance=job)
        document_formset = RequiredDocumentFormSet(instance=job)
    
    return render(request, 'jobs/post_job.html', {
        'form': form,
        'document_formset': document_formset,
        'job': job
    })

@login_required
@user_passes_test(is_hr_or_admin)
def delete_job(request, slug):
    job = get_object_or_404(JobPosting, slug=slug)
    
    # Block GET requests - deletion should only happen via POST with AJAX
    if request.method == 'GET':
        messages.error(request, 'Invalid request. Please use the delete button to delete job postings.')
        return redirect('jobs:manage_jobs')
    
    if request.method == 'POST':
        try:
            job_title = job.job_title
            job.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'message': f'Job posting "{job_title}" was deleted successfully.'
                })
            
            messages.success(request, f'Job posting "{job_title}" was deleted successfully!')
            return redirect('jobs:manage_jobs')
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'message': f'Error deleting job posting: {str(e)}'
                }, status=500)
            messages.error(request, f'Error deleting job posting: {str(e)}')
            return redirect('jobs:manage_jobs')
    
    return redirect('jobs:job_detail', slug=slug)

@login_required
def apply_for_job(request, slug):
    job = get_object_or_404(JobPosting, slug=slug, is_active=True)
    
    # Check if user has already applied
    if JobApplication.objects.filter(job=job, applicant=request.user).exists():
        messages.warning(request, 'You have already applied for this position.')
        return redirect('jobs:job_detail', slug=slug)
    
    # For internal jobs, show the application form
    if job.source_type == 'INTERNAL' or job.accepts_internal_applications:
        if request.method == 'POST':
            form = JobApplicationForm(request.POST, request.FILES, job=job)
            if form.is_valid():
                try:
                    with transaction.atomic():
                        application = form.save(commit=False)
                        application.job = job
                        application.applicant = request.user
                        application.save()
                        
                        # Handle required documents
                        for doc in job.required_documents.filter(is_required=True):
                            field_name = f'document_{doc.id}'
                            if field_name in request.FILES:
                                document = request.FILES[field_name]
                                # Store the document with the application
                                # You might want to create a separate model for this
                                # For now, we'll store it in additional_documents
                                application.additional_documents = document
                                application.save()
                        
                        messages.success(request, 'Your application has been submitted successfully!')
                        return redirect('jobs:job_detail', slug=slug)
                except Exception as e:
                    messages.error(request, f'Error submitting application: {str(e)}')
        else:
            form = JobApplicationForm(job=job)
        
        context = {
            'form': form,
            'job': job,
        }
        return render(request, 'jobs/apply.html', context)
    
    # For external jobs with application link, redirect to external site
    elif job.application_link:
        return redirect(job.application_link)
    
    # For external jobs without application link, redirect back with message
    else:
        messages.info(request, 'Please contact the organization directly to apply for this position.')
        return redirect('jobs:job_detail', slug=slug)

@login_required
@user_passes_test(is_hr_or_admin)
def manage_applications(request, slug):
    job = get_object_or_404(JobPosting, slug=slug)
    applications = job.applications.all()
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        applications = applications.filter(status=status)
    
    # Get application statistics
    stats = {
        'pending': applications.filter(status='PENDING').count(),
        'shortlisted': applications.filter(status='SHORTLISTED').count(),
        'interviewed': applications.filter(status='INTERVIEWED').count(),
        'accepted': applications.filter(status='ACCEPTED').count(),
    }
    
    # Pagination
    paginator = Paginator(applications, 20)
    page = request.GET.get('page')
    applications = paginator.get_page(page)
    
    context = {
        'job': job,
        'applications': applications,
        'stats': stats,
        'status': status,
        'status_choices': JobApplication.STATUS_CHOICES,
    }
    return render(request, 'jobs/manage_applications.html', context)

@login_required
@user_passes_test(is_hr_or_admin)
def update_application_status(request, application_id):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        application = get_object_or_404(JobApplication, id=application_id)
        data = json.loads(request.body)
        status = data.get('status')
        
        if status in dict(JobApplication.STATUS_CHOICES):
            application.status = status
            application.save()
            return JsonResponse({'success': True})
    
    return JsonResponse({'success': False}, status=400)

@login_required
@user_passes_test(is_hr_or_admin)
def application_details(request, application_id):
    application = get_object_or_404(JobApplication, id=application_id)
    context = {
        'application': application,
        'status_choices': JobApplication.STATUS_CHOICES,
    }
    html = render_to_string('jobs/application_details.html', context)
    return JsonResponse({'html': html})

@login_required
@user_passes_test(is_hr_or_admin)
def add_application_note(request, application_id):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        application = get_object_or_404(JobApplication, id=application_id)
        data = json.loads(request.body)
        note = data.get('note')
        
        if note:
            current_notes = application.notes or ''
            timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
            new_note = f"[{timestamp}] {request.user.get_full_name()}: {note}\n"
            application.notes = new_note + current_notes
            application.save()
            return JsonResponse({'success': True})
    
    return JsonResponse({'success': False}, status=400)

@login_required
@user_passes_test(is_hr_or_admin)
def send_application_email(request, application_id):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': False,
            'error': 'Invalid request'
        }, status=400)

    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Invalid method'
        }, status=405)

    try:
        application = get_object_or_404(JobApplication, id=application_id)
        data = json.loads(request.body)
        subject = data.get('subject')
        message = data.get('message')
        
        if not subject or not message:
            return JsonResponse({
                'success': False,
                'error': 'Subject and message are required.'
            }, status=400)
        
        # Send email using unified email system
        try:
            from core.email_utils import send_email_with_provider
            
            success = send_email_with_provider(
                subject=subject,
                message=message,
                recipient_list=[application.applicant.email],
                from_email=settings.DEFAULT_FROM_EMAIL,
                fail_silently=False
            )
            
            if not success:
                return JsonResponse({
                    'success': False,
                    'error': 'Failed to send email'
                }, status=500)
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Failed to send email: {str(e)}'
            }, status=500)
        
        # Log the email in notes
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
        note = f"[{timestamp}] Email sent by {request.user.get_full_name()}\nSubject: {subject}\nMessage: {message}\n\n"
        
        if application.notes:
            application.notes = note + application.notes
        else:
            application.notes = note
        
        application.save()
        
        return JsonResponse({'success': True})
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@user_passes_test(is_hr_or_admin)
def export_applicants(request, job_id):
    """Export job applicants to Excel file"""
    from core.export_utils import LogoHeaderService
    
    job = get_object_or_404(JobPosting, id=job_id)
    applicants = job.applications.all().select_related('applicant')
    
    # Create a workbook and add a worksheet
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=Applicants-{job.slug}-{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Applicants'
    
    # Add logo header and get starting row
    logo_path = LogoHeaderService.get_logo_path()
    start_row = LogoHeaderService.add_excel_header(
        worksheet, 
        logo_path,
        title=f"Job Applicants - {job.job_title}"
    )
    
    # Define the column headers
    columns = [
        'ID', 'Name', 'Email', 'Application Date', 'Status', 
        'Phone', 'Location', 'Skills', 'Experience', 'Notes'
    ]
    
    # Write the headers to the worksheet (at start_row instead of row 1)
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=start_row, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFCCCCCC', end_color='FFCCCCCC', fill_type='solid')
        cell.border = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000')
        )
    
    # Write the data to the worksheet (starting at start_row + 1 instead of row 2)
    row_num = start_row + 1
    for application in applicants:
        user = application.applicant
        
        # Get profile data if available
        try:
            profile = user.profile
            phone = profile.phone_number if hasattr(profile, 'phone_number') else 'N/A'
            location = profile.location if hasattr(profile, 'location') else 'N/A'
            skills = ", ".join([skill.name for skill in profile.skills.all()]) if hasattr(profile, 'skills') else 'N/A'
            experience = profile.years_of_experience if hasattr(profile, 'years_of_experience') else 'N/A'
        except:
            phone = location = skills = experience = 'N/A'
        
        row = [
            application.id,
            user.get_full_name(),
            user.email,
            application.application_date.strftime('%Y-%m-%d'),
            application.get_status_display(),
            phone,
            location,
            skills,
            experience,
            application.notes or 'N/A'
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            
        row_num += 1
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(response)
    return response



@login_required
@user_passes_test(is_hr_or_admin)
def bulk_update_jobs(request):
    """Admin function to perform bulk operations on job postings"""
    if request.method == 'POST':
        action = request.POST.get('action')
        job_ids = request.POST.getlist('job_ids')
        
        if not job_ids:
            messages.error(request, 'No jobs selected for bulk operation.')
            return redirect('jobs:manage_jobs')
        
        jobs = JobPosting.objects.filter(id__in=job_ids)
        
        try:
            if action == 'activate':
                jobs.update(is_active=True)
                messages.success(request, f'Successfully activated {jobs.count()} job postings.')
            elif action == 'deactivate':
                jobs.update(is_active=False)
                messages.success(request, f'Successfully deactivated {jobs.count()} job postings.')
            elif action == 'feature':
                jobs.update(is_featured=True)
                messages.success(request, f'Successfully featured {jobs.count()} job postings.')
            elif action == 'unfeature':
                jobs.update(is_featured=False)
                messages.success(request, f'Successfully unfeatured {jobs.count()} job postings.')
            elif action == 'delete':
                count = jobs.count()
                jobs.delete()
                messages.success(request, f'Successfully deleted {count} job postings.')
            else:
                messages.error(request, 'Invalid bulk action selected.')
        except Exception as e:
            messages.error(request, f'Error performing bulk operation: {str(e)}')
            logger.exception('Error in bulk update jobs')
        
        return redirect('jobs:manage_jobs')
    
    # GET request - redirect to manage jobs
    return redirect('jobs:manage_jobs')


@login_required
def job_scraper(request):
    """Job scraper main page with form"""
    form = JobScraperForm()
    return render(request, 'jobs/job_scraper.html', {
        'form': form,
        'page_title': 'Job Scraper - Find Jobs from BossJob.ph'
    })


@login_required
def job_scraper_results(request):
    """HTMX endpoint for job scraper results"""
    if request.method == 'POST':
        form = JobScraperForm(request.POST)
        
        if form.is_valid():
            keyword = form.cleaned_data['keyword']
            location = form.cleaned_data['location']
            
            try:
                # Perform the scraping
                results = scraper.search_jobs(keyword, location)
                
                # Log the scraping activity
                logger.info(f"User {request.user.username} scraped jobs for '{keyword}' in '{location}' - Found {results.get('total_found', 0)} jobs")
                
                # Save scraped data as JSON file
                try:
                    # Create directory if it doesn't exist
                    json_dir = os.path.join(settings.MEDIA_ROOT, 'scraped_jobs')
                    os.makedirs(json_dir, exist_ok=True)
                    
                    # Create filename with timestamp
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"scraped_jobs_{keyword.replace(' ', '_')}_{location.replace(' ', '_')}_{timestamp}.json"
                    filepath = os.path.join(json_dir, filename)
                    
                    # Prepare data for JSON export
                    export_data = {
                        'search_params': {
                            'keyword': keyword,
                            'location': location,
                            'scraped_by': request.user.username,
                            'scraped_at': datetime.now().isoformat()
                        },
                        'results': results
                    }
                    
                    # Write JSON file
                    with open(filepath, 'w', encoding='utf-8') as f:
                        json.dump(export_data, f, indent=2, ensure_ascii=False)
                    
                    logger.info(f"Scraped data saved to: {filepath}")
                    
                except Exception as e:
                    logger.error(f"Error saving scraped data to JSON: {str(e)}")
                
                # Save to database - deactivate old entries for same search
                try:
                    # Deactivate previous scraped jobs for the same keyword/location/user combination
                    ScrapedJob.objects.filter(
                        search_keyword=keyword,
                        search_location=location,
                        scraped_by=request.user,
                        source='BOSSJOB'
                    ).update(is_active=False)
                    
                    # Create new scraped job entry
                    scraped_job = ScrapedJob.objects.create(
                        search_keyword=keyword,
                        search_location=location,
                        source='BOSSJOB',
                        scraped_data=results,
                        total_found=results.get('total_found', 0),
                        scraped_by=request.user
                    )
                    logger.info(f"Scraped data saved to database with ID: {scraped_job.id}, previous entries deactivated")
                    
                except Exception as e:
                    logger.error(f"Error saving scraped data to database: {str(e)}")
                
                # Store job URLs in cache for redirect functionality
                for index, job in enumerate(results.get('jobs', [])):
                    if job.get('url') and job['url'] not in ['#', '#no-url-found', 'javascript:void(0)']:
                        cache_key = f"job_url_{index}"
                        cache.set(cache_key, job['url'], 3600)  # Cache for 1 hour
                
                # Return the results as HTML fragment for HTMX
                return render(request, 'jobs/partials/job_scraper_results.html', {
                    'results': results,
                    'form': form
                })
                
            except Exception as e:
                logger.error(f"Error in job scraper: {str(e)}")
                error_results = {
                    'success': False,
                    'jobs': [],
                    'total_found': 0,
                    'error': 'An unexpected error occurred while fetching jobs.',
                    'message': 'Please try again later.'
                }
                return render(request, 'jobs/partials/job_scraper_results.html', {
                    'results': error_results,
                    'form': form
                })
    else:
        # Form validation errors
        error_results = {
            'success': False,
            'jobs': [],
            'total_found': 0,
            'error': 'Please correct the form errors below.',
            'message': 'Invalid form data'
        }
        return render(request, 'jobs/partials/job_scraper_results.html', {
            'results': error_results,
            'form': form
        })
    
    # If not POST, redirect to main scraper page
    return redirect('jobs:job_scraper')


@login_required
def job_redirect(request, job_id):
    """Redirect to the actual job URL on BossJob.ph with improved error handling"""
    try:
        # Get the job URL from cache or database
        cache_key = f"job_url_{job_id}"
        job_url = cache.get(cache_key)
        
        if not job_url:
            # If not in cache, this might be an old or invalid job ID
            messages.error(request, "Job link not found or has expired. Please search again.")
            return redirect('jobs:job_scraper')
        
        # Handle special markers
        if job_url in ['#no-url-found', 'javascript:void(0)', '#']:
            messages.warning(request, "Direct job link is not available for this position. Redirecting to BossJob.ph job search.")
            return redirect('https://bossjob.ph/en-us/jobs-hiring')
        
        # Ensure the URL is BossJob.ph specific
        if 'bossjob.ph' not in job_url.lower():
            messages.info(request, "Redirecting to BossJob.ph main jobs page.")
            return redirect('https://bossjob.ph/en-us/jobs-hiring')
        
        # Test if the URL is accessible before redirecting
        try:
            import requests
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry
            
            # Create a session with retry strategy
            session = requests.Session()
            retry_strategy = Retry(
                total=2,
                status_forcelist=[429, 500, 502, 503, 504],
                method_whitelist=["HEAD", "GET", "OPTIONS"]
            )
            adapter = HTTPAdapter(max_retries=retry_strategy)
            session.mount("http://", adapter)
            session.mount("https://", adapter)
            
            # Test the URL with a HEAD request (faster than GET)
            response = session.head(job_url, timeout=5, allow_redirects=True)
            
            # If we get a 404 or other client error, try alternative URL patterns
            if response.status_code == 404:
                logger.warning(f"Job URL returned 404: {job_url}")
                
                # Try alternative URL patterns
                base_url = "https://bossjob.ph"
                job_id_match = None
                
                # Extract job ID from the URL
                import re
                patterns = [
                    r'/job/([^/?]+)',
                    r'/position/([^/?]+)',
                    r'/jobs/([^/?]+)',
                    r'job[_-]?id[=:]([^&/?]+)',
                    r'position[_-]?id[=:]([^&/?]+)'
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, job_url, re.IGNORECASE)
                    if match:
                        job_id_match = match.group(1)
                        break
                
                if job_id_match:
                    # Try different URL patterns
                    alternative_urls = [
                        f"{base_url}/job/{job_id_match}",
                        f"{base_url}/en-us/job/{job_id_match}",
                        f"{base_url}/position/{job_id_match}",
                        f"{base_url}/en-us/position/{job_id_match}",
                        f"{base_url}/jobs/{job_id_match}"
                    ]
                    
                    for alt_url in alternative_urls:
                        if alt_url != job_url:  # Don't test the same URL again
                            try:
                                alt_response = session.head(alt_url, timeout=3, allow_redirects=True)
                                if alt_response.status_code == 200:
                                    logger.info(f"Found working alternative URL: {alt_url}")
                                    # Update cache with working URL
                                    cache.set(cache_key, alt_url, 3600)  # Cache for 1 hour
                                    return redirect(alt_url)
                            except:
                                continue
                
                # If no alternative works, redirect to search page
                messages.warning(request, "The specific job page is no longer available. Redirecting to job search.")
                return redirect('https://bossjob.ph/en-us/jobs-hiring')
            
            elif response.status_code >= 400:
                logger.warning(f"Job URL returned status {response.status_code}: {job_url}")
                messages.warning(request, f"Job page is temporarily unavailable (Error {response.status_code}). Redirecting to job search.")
                return redirect('https://bossjob.ph/en-us/jobs-hiring')
            
        except requests.RequestException as e:
            logger.warning(f"Could not verify job URL {job_url}: {str(e)}")
            # If we can't verify, still try to redirect (might be a network issue)
            pass
        
        # Redirect to the actual job URL
        return redirect(job_url)
        
    except Exception as e:
        logger.error(f"Error in job redirect: {str(e)}")
        messages.error(request, "An error occurred while accessing the job link. Please try searching again.")
        return redirect('jobs:job_scraper')
